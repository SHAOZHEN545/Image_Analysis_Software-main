import copy
import csv
import datetime
import json
import os
import re
from contextlib import ExitStack, nullcontext
from dataclasses import dataclass, field
from functools import partial
from typing import List

import wx
import wx.lib.scrolledpanel
import numpy as np
import matplotlib.ticker as mticker
from matplotlib import cm, rcParams
from matplotlib.figure import Figure
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
from matplotlib.backends.backend_wx import NavigationToolbar2Wx
from matplotlib.font_manager import FontProperties
from matplotlib.lines import Line2D
from matplotlib.offsetbox import AnchoredOffsetbox, HPacker, TextArea, VPacker
from scipy.optimize import curve_fit

from fit_functions import FIT_FUNCTIONS, get_temp_psd_analysis, fit_multi_run_mot_lifetime
from imgFunc_v7 import readData
from openpyxl import Workbook

FIT_WINDOW_SETTINGS_FILE = os.path.join(
    os.path.dirname(__file__), "fitting_window_settings.json"
)

TEMP_PSD_DISPLAY_NAME = "Temperature, Density, and PSD"
MOT_LIFETIME_DISPLAY_NAME = "MOT Lifetime"
MOT_RINGDOWN_DISPLAY_NAME = "MOT Ringdown"
HEATMAP_VALUE_DISPLAY_NAME = "No Variable Fit"
HEATMAP_VALUE_FIT_KEY = "__heatmap_value__"
MULTI_FIT_OPTIONS = {
    TEMP_PSD_DISPLAY_NAME: TEMP_PSD_DISPLAY_NAME,
    MOT_LIFETIME_DISPLAY_NAME: MOT_LIFETIME_DISPLAY_NAME,
    HEATMAP_VALUE_DISPLAY_NAME: HEATMAP_VALUE_FIT_KEY,
}
MULTI_FIT_DISPLAY_ORDER = [
    TEMP_PSD_DISPLAY_NAME,
    MOT_LIFETIME_DISPLAY_NAME,
    HEATMAP_VALUE_DISPLAY_NAME,
]
TEMP_PSD_FIT_KEYS = {TEMP_PSD_DISPLAY_NAME}
SCALE_OPTIONS = ["Linear", "Semi-log", "Log"]
LINEAR_SCALE_ONLY_FITS = {
    TEMP_PSD_DISPLAY_NAME,
    MOT_LIFETIME_DISPLAY_NAME,
    MOT_RINGDOWN_DISPLAY_NAME,
}
PLOT_CAPTION_HEIGHT_RATIOS = (4, 1)
PLOT_RESIDUAL_CAPTION_HEIGHT_RATIOS = (4, 1.5, 1)


HOVER_SNAP_DISTANCE_SQ = 100.0
DEFAULT_PICK_TOLERANCE = 12.0
FILMSTRIP_SIZE_SCALE = 2.0 / 3.0
FILMSTRIP_PANEL_SCALE = 1.15
FILMSTRIP_MIN_HEIGHT = int(round(242 * FILMSTRIP_SIZE_SCALE))
FILMSTRIP_TARGET_HEIGHT = int(round(194 * FILMSTRIP_SIZE_SCALE))


def _compute_axis_label_fontsize():
    base = rcParams.get("axes.labelsize", 10)
    if isinstance(base, str):
        base = FontProperties(size=base).get_size_in_points()
    try:
        size = float(base)
    except (TypeError, ValueError):
        size = 10.0
    return size + 2.0


AXIS_LABEL_FONTSIZE = _compute_axis_label_fontsize()


def _normalise_fit_name(name):
    return name


@dataclass
class CaptionSpan:
    text: str
    bold: bool = False
    underline: bool = False


@dataclass
class CaptionLine:
    spans: List[CaptionSpan] = field(default_factory=list)

    def append(self, text, bold=False, underline=False):
        if text:
            self.spans.append(CaptionSpan(text, bold=bold, underline=underline))

    def plain_text(self):
        return "".join(span.text for span in self.spans)


FIT_PARAMETER_DIMENSIONS = {
    "Linear": ["y/x", "y"],
    "Quadratic": ["y/x^2", "y/x", "y"],
    "Exponential": ["y", "1/x", "y"],
    "Damped H.O.": ["y", "1/x", "1/x", "rad", "y"],
    "Gaussian": ["y", "x", "x", "y"],
    "Lorentzian": ["y", "x", "x", "y"],
    "Inverse": ["y*x", "x", "y"],
    TEMP_PSD_DISPLAY_NAME: ["y/x", "y"],
    MOT_LIFETIME_DISPLAY_NAME: ["y", "x", "y"],
    MOT_RINGDOWN_DISPLAY_NAME: ["y", "y", "1/x", "1/x", "rad"],
}


DERIVED_VALUE_DIMENSIONS = {
    ("Gaussian", "FWHM"): "x",
    ("Lorentzian", "FWHM"): "x",
}


PLAIN_FORMULAE = {
    "Linear": "y = m x + b",
    "Quadratic": "y = a x² + b x + c",
    "Exponential": "y = C + A e^{-B x}",
    "Damped H.O.": "y = C + A e^{-γ x} cos(ω x + φ)",
    "Gaussian": "y = A e^{-(x-μ)² / (2 σ²)} + C",
    "Lorentzian": "y = C + A / (1 + ((x - x₀)/(γ/2))²)",
    "Inverse": "y = A / (x - h) + B",
    TEMP_PSD_DISPLAY_NAME: "σ² = β t² + σ₀²",
    MOT_LIFETIME_DISPLAY_NAME: "N = N₀ e^{-t/τ} + C",
    MOT_RINGDOWN_DISPLAY_NAME: "z = z_eq + B e^{-γ t} cos(ω t + φ)",
}


class FittingWindow(wx.Frame):
    """Dedicated window for running fits and visualising results."""

    def __init__(self, parent):
        self.single_size = wx.Size(1265, 950)
        self.multi_size = wx.Size(
            int(self.single_size.GetWidth() * 1.1),
            int(self.single_size.GetHeight() * 0.9),
        )
        self.heatmap_size = wx.Size(
            int(self.single_size.GetWidth() * 1.3),
            int(self.single_size.GetHeight() * 0.9),
        )
        super().__init__(parent, title="Fitting v1.0", size=self.single_size)
        self._auto_resizing = False
        self._respect_user_bounds = False
        self._preferred_min_size = wx.Size(
            self.single_size.GetWidth(), self.single_size.GetHeight()
        )
        self.Bind(wx.EVT_SIZE, self._on_frame_size)
        self._apply_mode_size(self.single_size, force_resize=True)
        self.parent = parent

        # Capture the main window state so it can be restored when the fitting
        # window closes after loading a previous run.
        self._parent_original_settings = None
        self._parent_original_path = None
        self._parent_original_path_display = None
        self._parent_override_folder = None
        self._parent_context_restored = False
        self._parent_settings_modified = False

        if parent is not None:
            self._parent_original_path = getattr(parent, "path", None)
            path_ctrl = getattr(parent, "imageFolderPath", None)
            if path_ctrl is not None:
                try:
                    self._parent_original_path_display = path_ctrl.GetValue()
                except Exception:
                    self._parent_original_path_display = self._parent_original_path
            if self._parent_original_path_display is None:
                self._parent_original_path_display = (
                    self._parent_original_path if self._parent_original_path is not None else ""
                )

            collector = getattr(parent, "_collect_settings", None)
            if callable(collector):
                try:
                    self._parent_original_settings = copy.deepcopy(collector())
                except Exception as err:
                    print(f"Failed to capture image UI settings: {err}")

            saver = getattr(parent, "_save_settings", None)
            if callable(saver):
                try:
                    saver()
                except Exception as err:
                    print(
                        f"Failed to save image UI settings when opening fit window: {err}"
                    )

        self.results = None  # Processed data columns
        self.var_values = []
        self.param_values = []
        self.param2_values = []
        self.plot_data = {}
        self.var_scale = 1.0
        self.unit_scale = "s"
        self.param_scale = 1.0
        self.param_unit_scale = "s"
        self.param2_scale = 1.0
        self.param2_unit_scale = "s"
        self.subrun_results = []
        self.overlay_table_axes = []
        self.overlay_table_axis = None
        self.residual_axes = []
        self.overlay_right_axis = None
        self.caption_axes = []
        self.heatmap_axes_info = {}
        self.heatmap_index_grid = None
        self.heatmap_active_plots = set()
        self.heatmap_active_columns = {}
        self._current_axes_layout = None
        self._axis_plot_map = {}
        self._artist_plot_map = {}
        self._hover_annotations = {}
        self._hover_last = (None, None, None)
        self._hover_snap_distance_sq = HOVER_SNAP_DISTANCE_SQ
        self._font_sizes = {
            "label": AXIS_LABEL_FONTSIZE,
            "title": AXIS_LABEL_FONTSIZE + 4,
            "caption": 10.0,
            "tick": AXIS_LABEL_FONTSIZE,
        }
        self._has_unsaved_changes = False
        self._pick_tolerance = DEFAULT_PICK_TOLERANCE
        self.run_point_exclusions = {}
        self._session_point_exclusions = set()
        self._pending_session_point_exclusions = set()
        self._session_run_exclusions = {}
        self._pending_session_run_exclusions = {}

        self.unit_factors = {
            "time": {"s": 1.0, "ms": 1e-3, "us": 1e-6},
            "distance": {"m": 1.0, "cm": 1e-2, "mm": 1e-3, "um": 1e-6},
            "voltage": {"V": 1.0, "mV": 1e-3, "uV": 1e-6},
            "frequency": {"Hz": 1.0, "KHz": 1e3, "MHz": 1e6, "GHz": 1e9},
            "pixels": {"pix": 1.0},
            "arb.": {"arb.": 1.0},
        }
        self._unit_family_lookup = {}
        for family, mapping in self.unit_factors.items():
            for unit_name in mapping.keys():
                keys = {
                    self._normalise_unit_key(unit_name, lower=False),
                    self._normalise_unit_key(unit_name),
                }
                for key in keys:
                    if key:
                        self._unit_family_lookup[key] = family
        self.col_scales = {
            "Atom Number": ("M", 1e-6),
            "x-Atom Number": ("M", 1e-6),
            "y-Atom Number": ("M", 1e-6),
            "x-Center": ("pix", 1.0),
            "y-Center": ("pix", 1.0),
            "x-True Width": ("µm", 1e6),
            "y-True Width": ("µm", 1e6),
        }
        today = datetime.date.today()
        default_save = os.path.join(
            os.path.expanduser("~"),
            "Desktop",
            "Data",
            f"{today.year}",
            f"{today.month:02d}",
            f"{today.day:02d}",
        )
        os.makedirs(default_save, exist_ok=True)
        self.savePath = default_save

        panel = wx.Panel(self)
        self.panel = panel
        main_sizer = wx.BoxSizer(wx.HORIZONTAL)

        self.Bind(wx.EVT_CLOSE, self.on_close)

        control_panel = wx.Panel(panel)
        plot_panel = wx.Panel(panel)
        self.control_panel = control_panel
        self.plot_panel = plot_panel

        # Left column: all controls
        control_sizer = wx.BoxSizer(wx.VERTICAL)

        # Mode panel with radio buttons
        mode_box = wx.StaticBoxSizer(wx.HORIZONTAL, control_panel, "Mode")
        self.singleRadio = wx.RadioButton(
            control_panel,
            label="Variable Scan",
            style=wx.RB_GROUP,
        )
        self.singleRadio.SetToolTip(
            "Analyze one run: fit results vs. a scanned variable (e.g. atom number vs. hold time)."
        )
        self.multiRadio = wx.RadioButton(control_panel, label="Parameter Sweep")
        self.multiRadio.SetToolTip(
            "Compare multiple runs: plot fitted results (e.g. lifetime) vs. a changing parameter."
        )
        self.heatmapRadio = wx.RadioButton(control_panel, label="Parameter Heat Map")
        self.heatmapRadio.SetToolTip(
            "Scan two parameters: display the fitted result as a 2D color map."
        )
        mode_box.Add(self.singleRadio, 0, wx.ALL, 5)
        mode_box.Add(self.multiRadio, 0, wx.ALL, 5)
        mode_box.Add(self.heatmapRadio, 0, wx.ALL, 5)
        control_sizer.Add(mode_box, 0, wx.EXPAND | wx.ALL, 5)

        self._current_mode = "single"

        # Process panel
        process_box = wx.StaticBoxSizer(wx.VERTICAL, control_panel, "Process")
        start_sizer = wx.BoxSizer(wx.HORIZONTAL)
        start_sizer.Add(
            wx.StaticText(control_panel, label="Start from file:"),
            0,
            wx.ALIGN_CENTER_VERTICAL | wx.RIGHT,
            5,
        )
        default_start = (
            os.path.basename(self.parent.fileList[-1])
            if getattr(self.parent, "fileList", [])
            else ""
        )
        self.startFileCtrl = wx.TextCtrl(control_panel, value=default_start)
        self.startFileCtrl.SetToolTip("Name of the first image file to include when processing")
        start_sizer.Add(self.startFileCtrl, 1, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.startFileBtn = wx.Button(control_panel, label="Choose...")
        self.startFileBtn.SetToolTip("Browse for the starting image file")
        self.startFileBtn.Bind(wx.EVT_BUTTON, self.on_choose_start_file)
        start_sizer.Add(self.startFileBtn, 0, wx.ALIGN_CENTER_VERTICAL)
        process_box.Add(start_sizer, 0, wx.EXPAND | wx.ALL, 5)

        columns = wx.BoxSizer(wx.HORIZONTAL)

        # Parameter 2 column (for heat-map mode)
        param2_col = wx.BoxSizer(wx.VERTICAL)
        param2_label_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.param2LabelStatic = wx.StaticText(control_panel, label="Parameter 2:")
        param2_label_sizer.Add(self.param2LabelStatic, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.param2LabelCtrl = wx.TextCtrl(
            control_panel, value="Parameter 2", style=wx.TE_PROCESS_ENTER
        )
        self.param2LabelCtrl.SetToolTip("Caption to use for the second fit parameter axis")
        param2_label_sizer.Add(self.param2LabelCtrl, 1, wx.ALIGN_CENTER_VERTICAL)
        self.param2UnitTypeChoice = wx.Choice(control_panel, choices=["time", "distance", "voltage", "frequency", "pixels", "arb."])
        self.param2UnitTypeChoice.SetSelection(0)
        self.param2UnitTypeChoice.SetToolTip("Select the unit family for parameter 2 values")
        param2_label_sizer.Add(self.param2UnitTypeChoice, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 5)
        self.param2UnitScaleChoice = wx.Choice(control_panel)
        self.param2UnitScaleChoice.SetToolTip("Choose the scale used to display parameter 2")
        param2_label_sizer.Add(self.param2UnitScaleChoice, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 5)
        param2_col.Add(param2_label_sizer, 0, wx.EXPAND | wx.ALL, 5)
        self._update_param2_unit_scale_choices()

        self.param2FileLabel = wx.StaticText(control_panel, label="Parameter 2 List:")
        param2_col.Add(self.param2FileLabel, 0, wx.LEFT | wx.TOP, 5)
        param2_file_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.param2FileCtrl = wx.TextCtrl(control_panel)
        self.param2FileCtrl.SetToolTip("Path to a file containing parameter 2 values")
        param2_file_sizer.Add(self.param2FileCtrl, 1, wx.RIGHT, 5)
        self.chooseParam2FileBtn = wx.Button(control_panel, label="Choose File")
        self.chooseParam2FileBtn.SetToolTip("Browse for a parameter 2 list file")
        param2_file_sizer.Add(self.chooseParam2FileBtn, 0)
        param2_col.Add(param2_file_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)

        param2_offset_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.param2OffsetLabel = wx.StaticText(control_panel, label="Parameter 2 Offset:")
        param2_offset_sizer.Add(self.param2OffsetLabel, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.param2OffsetCtrl = wx.TextCtrl(control_panel, value="0", size=(60, -1))
        self.param2OffsetCtrl.SetToolTip("Constant offset applied to parameter 2 values")
        param2_offset_sizer.Add(self.param2OffsetCtrl, 0, wx.ALIGN_CENTER_VERTICAL)
        param2_col.Add(param2_offset_sizer, 0, wx.EXPAND | wx.ALL, 5)

        self.param2ListLabel = wx.StaticText(control_panel, label="Parameter 2 values:")
        param2_col.Add(self.param2ListLabel, 0, wx.LEFT | wx.TOP, 5)
        self.param2ListCtrl = wx.TextCtrl(control_panel, style=wx.TE_MULTILINE, size=(200, 100))
        self.param2ListCtrl.SetToolTip("Enter parameter 2 values manually, one per line")
        param2_col.Add(self.param2ListCtrl, 0, wx.EXPAND | wx.ALL, 5)

        self.param2_col_item = columns.Add(param2_col, 1, wx.EXPAND)

        # Parameter column
        param_col = wx.BoxSizer(wx.VERTICAL)
        param_label_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.paramLabelStatic = wx.StaticText(control_panel, label="Parameter:")
        param_label_sizer.Add(self.paramLabelStatic, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.paramLabelCtrl = wx.TextCtrl(
            control_panel, value="Parameter", style=wx.TE_PROCESS_ENTER
        )
        self.paramLabelCtrl.SetToolTip("Caption to use for the first fit parameter axis")
        param_label_sizer.Add(self.paramLabelCtrl, 1, wx.ALIGN_CENTER_VERTICAL)
        self.paramUnitTypeChoice = wx.Choice(control_panel, choices=["time", "distance", "voltage", "frequency", "pixels", "arb."])
        self.paramUnitTypeChoice.SetSelection(0)
        self.paramUnitTypeChoice.SetToolTip("Select the unit family for parameter values")
        param_label_sizer.Add(self.paramUnitTypeChoice, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 5)
        self.paramUnitScaleChoice = wx.Choice(control_panel)
        self.paramUnitScaleChoice.SetToolTip("Choose the scale used to display the parameter")
        param_label_sizer.Add(self.paramUnitScaleChoice, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 5)
        param_col.Add(param_label_sizer, 0, wx.EXPAND | wx.ALL, 5)
        self._update_param_unit_scale_choices()

        self.paramFileLabel = wx.StaticText(control_panel, label="Parameter List:")
        param_col.Add(self.paramFileLabel, 0, wx.LEFT | wx.TOP, 5)
        param_file_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.paramFileCtrl = wx.TextCtrl(control_panel)
        self.paramFileCtrl.SetToolTip("Path to a file containing parameter values")
        param_file_sizer.Add(self.paramFileCtrl, 1, wx.RIGHT, 5)
        self.chooseParamFileBtn = wx.Button(control_panel, label="Choose File")
        self.chooseParamFileBtn.SetToolTip("Browse for a parameter list file")
        param_file_sizer.Add(self.chooseParamFileBtn, 0)
        param_col.Add(param_file_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)

        param_offset_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.paramOffsetLabel = wx.StaticText(control_panel, label="Parameter Offset:")
        param_offset_sizer.Add(self.paramOffsetLabel, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.paramOffsetCtrl = wx.TextCtrl(control_panel, value="0", size=(60, -1))
        self.paramOffsetCtrl.SetToolTip("Constant offset applied to parameter values")
        param_offset_sizer.Add(self.paramOffsetCtrl, 0, wx.ALIGN_CENTER_VERTICAL)
        param_col.Add(param_offset_sizer, 0, wx.EXPAND | wx.ALL, 5)

        self.paramListLabel = wx.StaticText(control_panel, label="Parameter values:")
        param_col.Add(self.paramListLabel, 0, wx.LEFT | wx.TOP, 5)
        self.paramListCtrl = wx.TextCtrl(control_panel, style=wx.TE_MULTILINE, size=(200, 100))
        self.paramListCtrl.SetToolTip("Enter parameter values manually, one per line")
        param_col.Add(self.paramListCtrl, 0, wx.EXPAND | wx.ALL, 5)

        self.param_col_item = columns.Add(param_col, 1, wx.EXPAND)

        # Variable column
        var_col = wx.BoxSizer(wx.VERTICAL)
        var_label_sizer = wx.BoxSizer(wx.HORIZONTAL)
        var_label_sizer.Add(wx.StaticText(control_panel, label="Variable:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.varLabelCtrl = wx.TextCtrl(
            control_panel, value="TOF", style=wx.TE_PROCESS_ENTER
        )
        self.varLabelCtrl.SetToolTip("Label for the horizontal axis and processed variable")
        var_label_sizer.Add(self.varLabelCtrl, 1, wx.ALIGN_CENTER_VERTICAL)
        self.unitTypeChoice = wx.Choice(control_panel, choices=["time", "distance", "voltage", "frequency", "pixels", "arb."])
        self.unitTypeChoice.SetSelection(0)
        self.unitTypeChoice.SetToolTip("Select the unit family for the variable values")
        var_label_sizer.Add(self.unitTypeChoice, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 5)
        self.unitScaleChoice = wx.Choice(control_panel)
        self.unitScaleChoice.SetToolTip("Choose the scale used to display the variable")
        var_label_sizer.Add(self.unitScaleChoice, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 5)
        var_col.Add(var_label_sizer, 0, wx.EXPAND | wx.ALL, 5)
        self._update_unit_scale_choices()

        self.varFileLabel = wx.StaticText(control_panel, label="Variable List:")
        var_col.Add(self.varFileLabel, 0, wx.LEFT | wx.TOP, 5)
        file_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.varFileCtrl = wx.TextCtrl(control_panel)
        self.varFileCtrl.SetToolTip("Path to a file containing variable values")
        file_sizer.Add(self.varFileCtrl, 1, wx.RIGHT, 5)
        self.chooseVarFileBtn = wx.Button(control_panel, label="Choose File")
        self.chooseVarFileBtn.SetToolTip("Browse for a variable list file")
        file_sizer.Add(self.chooseVarFileBtn, 0)
        var_col.Add(file_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)

        var_offset_sizer = wx.BoxSizer(wx.HORIZONTAL)
        var_offset_sizer.Add(wx.StaticText(control_panel, label="Variable Offset:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.varOffsetCtrl = wx.TextCtrl(control_panel, value="0", size=(60, -1))
        self.varOffsetCtrl.SetToolTip("Constant offset applied to variable values")
        var_offset_sizer.Add(self.varOffsetCtrl, 0, wx.ALIGN_CENTER_VERTICAL)
        var_col.Add(var_offset_sizer, 0, wx.EXPAND | wx.ALL, 5)

        var_col.Add(wx.StaticText(control_panel, label="Variable values:"), 0, wx.LEFT | wx.TOP, 5)
        self.varListCtrl = wx.TextCtrl(control_panel, style=wx.TE_MULTILINE, size=(200, 100))
        self.varListCtrl.SetToolTip("Enter variable values manually, one per line")
        var_col.Add(self.varListCtrl, 0, wx.EXPAND | wx.ALL, 5)

        multi_fit_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.multiFitLabel = wx.StaticText(control_panel, label="Multi-fit:")
        multi_fit_sizer.Add(
            self.multiFitLabel, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5
        )
        self.multiFitChoice = wx.Choice(control_panel, choices=[])
        self.multiFitChoice.SetToolTip("Select the multi-run fitting routine to apply")
        self.multiFitChoice.Bind(wx.EVT_CHOICE, self.on_multi_fit_choice)
        multi_fit_sizer.Add(self.multiFitChoice, 0, wx.ALIGN_CENTER_VERTICAL)
        var_col.Add(multi_fit_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 5)

        columns.Add(var_col, 1, wx.EXPAND)

        process_box.Add(columns, 0, wx.EXPAND)

        self._variable_controls = [
            self.varLabelCtrl,
            self.unitTypeChoice,
            self.unitScaleChoice,
            self.varFileCtrl,
            self.chooseVarFileBtn,
            self.varOffsetCtrl,
            self.varListCtrl,
        ]

        self._refresh_multi_fit_choices()

        # Process button with progress bar
        proc_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.processBtn = wx.Button(control_panel, label="Process")
        self.processBtn.SetToolTip("Start or stop processing the selected data set")
        proc_sizer.Add(self.processBtn, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.progress = wx.Gauge(control_panel, range=1, size=(120, -1))
        self.progress.SetToolTip("Shows progress while files are being processed")
        proc_sizer.Add(self.progress, 1, wx.ALIGN_CENTER_VERTICAL)
        process_box.Add(proc_sizer, 0, wx.EXPAND | wx.ALL, 5)
        self._gauge_default_colour = self.progress.GetForegroundColour()
        self._processing = False
        self._stop_requested = False

        control_sizer.Add(process_box, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)

        # Hide parameter controls until multi-run mode enabled
        for w in [
            self.paramLabelStatic,
            self.paramLabelCtrl,
            self.paramUnitTypeChoice,
            self.paramUnitScaleChoice,
            self.paramFileLabel,
            self.paramFileCtrl,
            self.chooseParamFileBtn,
            self.paramOffsetLabel,
            self.paramOffsetCtrl,
            self.paramListLabel,
            self.paramListCtrl,
            self.param2LabelStatic,
            self.param2LabelCtrl,
            self.param2UnitTypeChoice,
            self.param2UnitScaleChoice,
            self.param2FileLabel,
            self.param2FileCtrl,
            self.chooseParam2FileBtn,
            self.param2OffsetLabel,
            self.param2OffsetCtrl,
            self.param2ListLabel,
            self.param2ListCtrl,
            self.multiFitLabel,
            self.multiFitChoice,
        ]:
            w.Hide()
        self.param_col_item.Show(False)
        self.param2_col_item.Show(False)
        self._update_multi_fit_controls()

        # Plot controls (for selecting data to plot)
        plot_ctrl_box = wx.StaticBoxSizer(wx.VERTICAL, control_panel, "Plots")
        plots_ctrl = wx.BoxSizer(wx.HORIZONTAL)

        plot1_box = wx.StaticBoxSizer(wx.VERTICAL, control_panel, "Plot 1")
        plot1_box.Add(wx.StaticText(control_panel, label="Data"), 0, wx.ALL, 5)
        self.colChoice1 = wx.Choice(control_panel)
        self.colChoice1.SetToolTip("Select which processed column to display in Plot 1")
        plot1_box.Add(self.colChoice1, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        self.scaleLabel1 = wx.StaticText(control_panel, label="Scale")
        plot1_box.Add(self.scaleLabel1, 0, wx.ALL, 5)
        self.scaleChoice1 = wx.Choice(control_panel, choices=SCALE_OPTIONS)
        self.scaleChoice1.SetSelection(0)
        self.scaleChoice1.SetToolTip("Choose the axis scale for Plot 1")
        plot1_box.Add(self.scaleChoice1, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        self.fitLabel1 = wx.StaticText(control_panel, label="Fit Function")
        plot1_box.Add(self.fitLabel1, 0, wx.ALL, 5)
        self._base_fit_choices = [""] + list(FIT_FUNCTIONS.keys())
        self.funcChoice1 = wx.Choice(control_panel, choices=list(self._base_fit_choices))
        self.funcChoice1.SetSelection(0)
        self.funcChoice1.SetToolTip("Select the fitting model applied to Plot 1")
        plot1_box.Add(self.funcChoice1, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        self.paramGuessPanel1 = wx.Panel(control_panel)
        self.paramSizer1 = wx.BoxSizer(wx.VERTICAL)
        self.paramGuessPanel1.SetSizer(self.paramSizer1)
        plot1_box.Add(self.paramGuessPanel1, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        self.paramGuessPanel1.Hide()
        self.paramGuessLabel1 = None
        caption_sizer1 = wx.BoxSizer(wx.HORIZONTAL)
        self.captionLbl1 = wx.StaticText(control_panel, label="Caption")
        caption_sizer1.Add(self.captionLbl1, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.captionCtrl1 = wx.TextCtrl(
            control_panel, style=wx.TE_MULTILINE | wx.TE_PROCESS_ENTER
        )
        self.captionCtrl1.SetToolTip(
            "Optional caption text displayed beneath Plot 1. Press Enter to apply; "
            "Shift+Enter inserts a new line."
        )
        caption_sizer1.Add(self.captionCtrl1, 1, wx.ALIGN_CENTER_VERTICAL)
        plot1_box.Add(caption_sizer1, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        self.plotBtn1 = wx.Button(control_panel, label="Plot 1")
        self.plotBtn1.SetToolTip("Generate or refresh Plot 1")
        self.closeBtn1 = wx.Button(control_panel, label="Close")
        self.closeBtn1.Disable()
        self.closeBtn1.SetToolTip("Remove Plot 1 from the canvas")
        btn_sizer1 = wx.BoxSizer(wx.HORIZONTAL)
        btn_sizer1.Add(self.plotBtn1, 0, wx.RIGHT, 5)
        btn_sizer1.Add(self.closeBtn1, 0)
        plot1_box.Add(btn_sizer1, 0, wx.ALL | wx.ALIGN_CENTER, 5)

        plot2_box = wx.StaticBoxSizer(wx.VERTICAL, control_panel, "Plot 2")
        plot2_box.Add(wx.StaticText(control_panel, label="Data"), 0, wx.ALL, 5)
        self.colChoice2 = wx.Choice(control_panel)
        self.colChoice2.SetToolTip("Select which processed column to display in Plot 2")
        plot2_box.Add(self.colChoice2, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        self.scaleLabel2 = wx.StaticText(control_panel, label="Scale")
        plot2_box.Add(self.scaleLabel2, 0, wx.ALL, 5)
        self.scaleChoice2 = wx.Choice(control_panel, choices=SCALE_OPTIONS)
        self.scaleChoice2.SetSelection(0)
        self.scaleChoice2.SetToolTip("Choose the axis scale for Plot 2")
        plot2_box.Add(self.scaleChoice2, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        self.fitLabel2 = wx.StaticText(control_panel, label="Fit Function")
        plot2_box.Add(self.fitLabel2, 0, wx.ALL, 5)
        self.funcChoice2 = wx.Choice(control_panel, choices=list(self._base_fit_choices))
        self.funcChoice2.SetSelection(0)
        self.funcChoice2.SetToolTip("Select the fitting model applied to Plot 2")
        plot2_box.Add(self.funcChoice2, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)

        fit_choice_best = self.funcChoice1.GetBestSize()
        min_width = fit_choice_best.GetWidth()
        min_height = fit_choice_best.GetHeight()
        min_size = wx.Size(min_width, min_height)
        self.funcChoice1.SetMinSize(min_size)
        self.funcChoice2.SetMinSize(min_size)
        self.funcChoice1.SetInitialSize(min_size)
        self.funcChoice2.SetInitialSize(min_size)
        self.paramGuessPanel2 = wx.Panel(control_panel)
        self.paramSizer2 = wx.BoxSizer(wx.VERTICAL)
        self.paramGuessPanel2.SetSizer(self.paramSizer2)
        plot2_box.Add(self.paramGuessPanel2, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        self.paramGuessPanel2.Hide()
        self.paramGuessLabel2 = None
        caption_sizer2 = wx.BoxSizer(wx.HORIZONTAL)
        self.captionLbl2 = wx.StaticText(control_panel, label="Caption")
        caption_sizer2.Add(self.captionLbl2, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.captionCtrl2 = wx.TextCtrl(
            control_panel, style=wx.TE_MULTILINE | wx.TE_PROCESS_ENTER
        )
        self.captionCtrl2.SetToolTip(
            "Optional caption text displayed beneath Plot 2. Press Enter to apply; "
            "Shift+Enter inserts a new line."
        )
        caption_sizer2.Add(self.captionCtrl2, 1, wx.ALIGN_CENTER_VERTICAL)
        plot2_box.Add(caption_sizer2, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        self.plotBtn2 = wx.Button(control_panel, label="Plot 2")
        self.plotBtn2.SetToolTip("Generate or refresh Plot 2")
        self.closeBtn2 = wx.Button(control_panel, label="Close")
        self.closeBtn2.Disable()
        self.closeBtn2.SetToolTip("Remove Plot 2 from the canvas")
        btn_sizer2 = wx.BoxSizer(wx.HORIZONTAL)
        btn_sizer2.Add(self.plotBtn2, 0, wx.RIGHT, 5)
        btn_sizer2.Add(self.closeBtn2, 0)
        plot2_box.Add(btn_sizer2, 0, wx.ALL | wx.ALIGN_CENTER, 5)

        plots_ctrl.Add(plot1_box, 0, wx.ALL, 5)
        self.plot2_box = plot2_box
        self.plot2_box_item = plots_ctrl.Add(plot2_box, 0, wx.ALL, 5)
        plot_ctrl_box.Add(plots_ctrl, 0, wx.ALIGN_LEFT)

        plot_options_box = wx.StaticBoxSizer(
            wx.StaticBox(control_panel, label="Plot Options"), wx.VERTICAL
        )
        options_row = wx.BoxSizer(wx.HORIZONTAL)
        self.paramBinCheck = wx.CheckBox(control_panel, label="Parameter Binning")
        self.paramBinCheck.SetToolTip(
            "Average neighbouring points on the aggregated parameter axis before fitting"
        )
        options_row.Add(self.paramBinCheck, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 8)
        self.varBinCheck = wx.CheckBox(control_panel, label="Variable Binning")
        self.varBinCheck.SetToolTip(
            "Average repeated variable list values within each run before analysis"
        )
        options_row.Add(self.varBinCheck, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 8)
        self.heatmapVarBinCheck = wx.CheckBox(
            control_panel, label="Heat Map Variable Binning"
        )
        self.heatmapVarBinCheck.SetToolTip(
            "Average repeated variable values within each heat map run before aggregation"
        )
        options_row.Add(
            self.heatmapVarBinCheck, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 8
        )
        self.heatmapVarBinCheck.Hide()
        self.overlayCheck = wx.CheckBox(control_panel, label="Overlay")
        self.overlayCheck.SetToolTip("Overlay Plot 2 on Plot 1 for comparison")
        options_row.Add(self.overlayCheck, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 8)
        self.overlayColorCheck = wx.CheckBox(control_panel, label="Colors")
        self.overlayColorCheck.SetToolTip(
            "Use distinct colors (blue circles, red squares) when overlaying plots"
        )
        options_row.Add(self.overlayColorCheck, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 8)
        self.overlayColorCheck.Enable(False)
        self.overlayColorCheck.Hide()
        self.residualCheck = wx.CheckBox(control_panel, label="Show Residuals")
        self.residualCheck.SetToolTip(
            "Display the residuals of the fitted curve directly below each plot"
        )
        options_row.Add(self.residualCheck, 0, wx.ALIGN_CENTER_VERTICAL)
        plot_options_box.Add(options_row, 0, wx.ALL, 5)
        plot_ctrl_box.Add(plot_options_box, 0, wx.ALL | wx.EXPAND, 5)

        control_sizer.Add(plot_ctrl_box, 0, wx.EXPAND | wx.ALL, 5)

        # Exclusion checkboxes (populated on pick events)
        self.excludePanel = wx.Panel(control_panel)
        exclude_box = wx.StaticBoxSizer(
            wx.StaticBox(self.excludePanel, label="Exclusions"), wx.VERTICAL
        )
        self.excludeScroll = wx.lib.scrolledpanel.ScrolledPanel(
            self.excludePanel, style=wx.TAB_TRAVERSAL | wx.BORDER_NONE
        )
        self.excludeScroll.SetScrollRate(0, 10)
        self.excludeSizer = wx.BoxSizer(wx.VERTICAL)
        self.excludeScroll.SetSizer(self.excludeSizer)
        self.excludeScroll.SetupScrolling(scroll_x=False, scroll_y=True)
        self._exclude_row_height = self._estimate_exclude_row_height(self.excludeScroll)
        visible_height = self._exclude_row_height * 4
        self.excludeScroll.SetMinSize(wx.Size(-1, visible_height))
        self.excludeScroll.SetMaxSize(wx.Size(-1, visible_height))
        exclude_box.Add(self.excludeScroll, 1, wx.EXPAND)
        self.excludePanel.SetSizer(exclude_box)
        self.excludePanel.Hide()
        control_sizer.Add(self.excludePanel, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)

        # Save controls
        save_box = wx.StaticBoxSizer(wx.VERTICAL, control_panel, "Save Data")
        save_box.Add(wx.StaticText(control_panel, label="Save Folder:"), 0, wx.LEFT | wx.TOP, 5)
        save_path_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.savePathCtrl = wx.TextCtrl(control_panel, value=self.savePath)
        self.savePathCtrl.Bind(wx.EVT_TEXT, self.on_save_path_text)
        self.savePathCtrl.SetToolTip("Folder where processed data will be exported")
        save_path_sizer.Add(self.savePathCtrl, 1, wx.RIGHT, 5)
        self.chooseSaveBtn = wx.Button(control_panel, label="Choose Folder")
        self.chooseSaveBtn.Bind(wx.EVT_BUTTON, self.on_choose_save_path)
        self.chooseSaveBtn.SetToolTip("Browse for a folder to store exported results")
        save_path_sizer.Add(self.chooseSaveBtn, 0)
        save_box.Add(save_path_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 5)
        action_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.saveBtn = wx.Button(control_panel, label="Save Data")
        self.saveBtn.SetToolTip("Export the current fit results to the save folder")
        action_sizer.Add(self.saveBtn, 0)
        self.loadSessionBtn = wx.Button(control_panel, label="Load Run")
        self.loadSessionBtn.SetToolTip(
            "Restore a previously saved fitting run from an export bundle"
        )
        action_sizer.Add(self.loadSessionBtn, 0, wx.LEFT, 5)
        save_box.Add(action_sizer, 0, wx.LEFT | wx.RIGHT | wx.TOP | wx.BOTTOM, 5)
        self._mark_unsaved(False)

        control_sizer.Add(save_box, 0, wx.EXPAND | wx.ALL, 5)

        control_panel.SetSizer(control_sizer)
        main_sizer.Add(control_panel, 0, wx.EXPAND)

        # Plot area in its own panel
        fig = Figure()
        self.canvas = FigureCanvas(plot_panel, -1, fig)
        self.axes = [fig.add_subplot(111)]
        self._reset_axes_box_aspect(self.axes)
        self.toolbar = NavigationToolbar2Wx(self.canvas)
        self.heatmap_cbar = []
        plot_sizer = wx.BoxSizer(wx.VERTICAL)
        plot_sizer.Add(self.canvas, 1, wx.EXPAND)
        plot_sizer.Add(self.toolbar, 0, wx.EXPAND)
        plot_panel.SetSizer(plot_sizer)
        main_sizer.Add(plot_panel, 1, wx.EXPAND)

        panel.SetSizer(main_sizer)

        default_cols = list(self.col_scales.keys())
        self.colChoice1.Set(default_cols)
        self.colChoice2.Set(default_cols)
        if default_cols:
            self.colChoice1.SetSelection(0)
            self.colChoice2.SetSelection(0)

        # Bind events
        self.processBtn.Bind(wx.EVT_BUTTON, self.on_process)
        self.saveBtn.Bind(wx.EVT_BUTTON, self.on_save)
        self.loadSessionBtn.Bind(wx.EVT_BUTTON, self.on_load_session)
        self.chooseVarFileBtn.Bind(wx.EVT_BUTTON, self.on_choose_var_file)
        self.varFileCtrl.Bind(wx.EVT_TEXT, self.on_var_file_text)
        self.plotBtn1.Bind(wx.EVT_BUTTON, lambda evt: self.update_plot(1))
        self.plotBtn2.Bind(wx.EVT_BUTTON, lambda evt: self.update_plot(2))
        self.closeBtn1.Bind(wx.EVT_BUTTON, self.on_close_plot1)
        self.closeBtn2.Bind(wx.EVT_BUTTON, self.on_close_plot2)
        self.paramBinCheck.Bind(wx.EVT_CHECKBOX, self.on_param_bin_toggle)
        self.varBinCheck.Bind(wx.EVT_CHECKBOX, self.on_var_bin_toggle)
        self.heatmapVarBinCheck.Bind(
            wx.EVT_CHECKBOX, self.on_heatmap_var_bin_toggle
        )
        self.overlayCheck.Bind(wx.EVT_CHECKBOX, self.on_overlay_toggle)
        self.overlayColorCheck.Bind(wx.EVT_CHECKBOX, self.on_overlay_color_toggle)
        self.residualCheck.Bind(wx.EVT_CHECKBOX, self.on_residual_toggle)
        self.funcChoice1.Bind(
            wx.EVT_CHOICE, lambda evt: self.on_func_change(1, event=evt)
        )
        self.funcChoice2.Bind(
            wx.EVT_CHOICE, lambda evt: self.on_func_change(2, event=evt)
        )
        self.colChoice1.Bind(wx.EVT_CHOICE, lambda evt: self.on_column_change(1, evt))
        self.colChoice2.Bind(wx.EVT_CHOICE, lambda evt: self.on_column_change(2, evt))
        self.scaleChoice1.Bind(wx.EVT_CHOICE, lambda evt: self.on_scale_change(1))
        self.scaleChoice2.Bind(wx.EVT_CHOICE, lambda evt: self.on_scale_change(2))
        self.unitTypeChoice.Bind(wx.EVT_CHOICE, self.on_unit_type_change)
        self.unitScaleChoice.Bind(wx.EVT_CHOICE, self.on_unit_scale_change)
        self.paramUnitTypeChoice.Bind(wx.EVT_CHOICE, self.on_param_unit_type_change)
        self.paramUnitScaleChoice.Bind(wx.EVT_CHOICE, self.on_param_unit_scale_change)
        self.param2UnitTypeChoice.Bind(wx.EVT_CHOICE, self.on_param2_unit_type_change)
        self.param2UnitScaleChoice.Bind(wx.EVT_CHOICE, self.on_param2_unit_scale_change)
        self.singleRadio.Bind(wx.EVT_RADIOBUTTON, self.on_mode_change)
        self.multiRadio.Bind(wx.EVT_RADIOBUTTON, self.on_mode_change)
        self.heatmapRadio.Bind(wx.EVT_RADIOBUTTON, self.on_mode_change)
        self.chooseParamFileBtn.Bind(wx.EVT_BUTTON, self.on_choose_param_file)
        self.paramFileCtrl.Bind(wx.EVT_TEXT, self.on_param_file_text)
        self.chooseParam2FileBtn.Bind(wx.EVT_BUTTON, self.on_choose_param2_file)
        self.param2FileCtrl.Bind(wx.EVT_TEXT, self.on_param2_file_text)
        self._bind_enter_action(self.varLabelCtrl, self.on_axis_label_enter)
        self._bind_enter_action(self.paramLabelCtrl, self.on_axis_label_enter)
        self._bind_enter_action(self.param2LabelCtrl, self.on_axis_label_enter)

        # Load default variable and parameter list files if present
        default_dir = os.path.join(
            os.path.expanduser("~"),
            "Desktop",
            "Image Network Folder",
        )
        default_var_file = os.path.join(default_dir, "Variable List.txt")
        if os.path.exists(default_var_file):
            self.varFileCtrl.SetValue(default_var_file)
            self.load_variable_file(default_var_file)
        default_param_file = os.path.join(default_dir, "Parameter List.txt")
        if os.path.exists(default_param_file):
            self.paramFileCtrl.SetValue(default_param_file)
            self.load_parameter_file(default_param_file)
        default_param2_file = os.path.join(default_dir, "Parameter 2 List.txt")
        if not os.path.exists(default_param2_file):
            alt = os.path.join(default_dir, "Parameter2 List.txt")
            if os.path.exists(alt):
                default_param2_file = alt
        if os.path.exists(default_param2_file):
            self.param2FileCtrl.SetValue(default_param2_file)
            self.load_parameter2_file(default_param2_file)

        self.paramCtrls1 = []
        self.paramCtrls2 = []
        self.exclude_checks = {}
        self.detail_checks = {}
        self.detail_windows = {}
        self.point_panels = {}
        self.point_label_info = {}
        self.variable_bin_groups = {}
        self.heatmap_detail_keys = set()
        self.current_multi_fit = None

        self.canvas.mpl_connect("pick_event", self.on_pick)
        self.canvas.mpl_connect("button_press_event", self.on_canvas_click)
        self.canvas.mpl_connect("motion_notify_event", self.on_canvas_motion)
        self._last_column_selection = {
            1: self.colChoice1.GetStringSelection(),
            2: self.colChoice2.GetStringSelection(),
        }
        self._last_func_selection = {
            1: _normalise_fit_name(self.funcChoice1.GetStringSelection()),
            2: _normalise_fit_name(self.funcChoice2.GetStringSelection()),
        }
        self._suspend_fit_visibility = False
        self._bind_enter_action(
            self.captionCtrl1,
            lambda: self.on_caption_change(1),
            allow_shift_newline=True,
        )
        self._bind_enter_action(
            self.captionCtrl2,
            lambda: self.on_caption_change(2),
            allow_shift_newline=True,
        )

        self._processing_sensitive_controls = [
            self.singleRadio,
            self.multiRadio,
            self.heatmapRadio,
            self.startFileCtrl,
            self.startFileBtn,
            self.paramLabelCtrl,
            self.paramUnitTypeChoice,
            self.paramUnitScaleChoice,
            self.paramFileCtrl,
            self.chooseParamFileBtn,
            self.paramOffsetCtrl,
            self.paramListCtrl,
            self.param2LabelCtrl,
            self.param2UnitTypeChoice,
            self.param2UnitScaleChoice,
            self.param2FileCtrl,
            self.chooseParam2FileBtn,
            self.param2OffsetCtrl,
            self.param2ListCtrl,
            self.varLabelCtrl,
            self.varFileLabel,
            self.unitTypeChoice,
            self.unitScaleChoice,
            self.varFileCtrl,
            self.chooseVarFileBtn,
            self.varOffsetCtrl,
            self.varListCtrl,
            self.multiFitChoice,
            self.savePathCtrl,
            self.chooseSaveBtn,
            self.loadSessionBtn,
            self.overlayColorCheck,
        ]

        self._set_controls_enabled(False)

        self._settings_path = FIT_WINDOW_SETTINGS_FILE
        self._user_settings = self._load_user_settings()
        self._apply_user_settings(self._user_settings)
        self._refresh_param_guess_visibility()

    def _get_display_client_size(self):
        display_index = wx.Display.GetFromWindow(self)
        if display_index == wx.NOT_FOUND:
            display_index = wx.Display.GetFromPoint(wx.GetMousePosition())
        if display_index == wx.NOT_FOUND:
            width, height = wx.DisplaySize()
            return wx.Size(width, height)
        display = wx.Display(display_index)
        area = display.GetClientArea()
        return wx.Size(area.GetWidth(), area.GetHeight())

    def _clamp_size_to_display(self, size):
        available = self._get_display_client_size()
        return wx.Size(
            min(size.GetWidth(), available.GetWidth()),
            min(size.GetHeight(), available.GetHeight()),
        )

    def _apply_mode_size(self, base_size, force_resize=False):
        clamped = self._clamp_size_to_display(base_size)
        self._preferred_min_size = wx.Size(clamped.GetWidth(), clamped.GetHeight())
        if force_resize:
            self._respect_user_bounds = False
        current_size = self.GetSize()
        if hasattr(current_size, "GetWidth"):
            current_width = current_size.GetWidth()
            current_height = current_size.GetHeight()
        else:
            current_width, current_height = current_size
        min_width = clamped.GetWidth()
        min_height = clamped.GetHeight()
        user_locked = (self._respect_user_bounds and not force_resize) or self.IsMaximized()
        target_min_width = min_width if not user_locked else min(min_width, current_width)
        target_min_height = min_height if not user_locked else min(min_height, current_height)
        target_min = wx.Size(target_min_width, target_min_height)
        self._auto_resizing = True
        try:
            self.SetMinSize(target_min)
            if not user_locked:
                new_width = max(current_width, target_min_width)
                new_height = max(current_height, target_min_height)
                if new_width != current_width or new_height != current_height:
                    self.SetSize((new_width, new_height))
        finally:
            self._auto_resizing = False
        return clamped

    def _on_frame_size(self, event):
        if event.GetEventObject() is self and not self._auto_resizing:
            size = event.GetSize()
            if hasattr(size, "GetWidth"):
                width, height = size.GetWidth(), size.GetHeight()
            else:
                width, height = size
            pref_width = self._preferred_min_size.GetWidth()
            pref_height = self._preferred_min_size.GetHeight()
            if self.IsMaximized():
                self._respect_user_bounds = True
            else:
                self._respect_user_bounds = width < pref_width or height < pref_height
        event.Skip()

    def _update_unit_scale_choices(self):
        unit_type = self.unitTypeChoice.GetStringSelection()
        scales = list(self.unit_factors[unit_type].keys())
        self.unitScaleChoice.Set(scales)
        index = 1 if unit_type == "time" and len(scales) > 1 else 0
        self.unitScaleChoice.SetSelection(index)
        self.unit_scale = scales[index]
        self.var_scale = self.unit_factors[unit_type][self.unit_scale]

    def on_unit_type_change(self, event):
        self._update_unit_scale_choices()
        self._refresh_plot_labels()

    def on_unit_scale_change(self, event):
        unit_type = self.unitTypeChoice.GetStringSelection()
        self.unit_scale = self.unitScaleChoice.GetStringSelection()
        self.var_scale = self.unit_factors[unit_type][self.unit_scale]
        self._refresh_plot_labels()

    def _update_param_unit_scale_choices(self):
        unit_type = self.paramUnitTypeChoice.GetStringSelection()
        scales = list(self.unit_factors[unit_type].keys())
        self.paramUnitScaleChoice.Set(scales)
        index = 1 if unit_type == "time" and len(scales) > 1 else 0
        self.paramUnitScaleChoice.SetSelection(index)
        self.param_unit_scale = scales[index]
        self.param_scale = self.unit_factors[unit_type][self.param_unit_scale]

    def _mark_unsaved(self, unsaved):
        results = self.results
        if isinstance(results, dict):
            has_results = len(results) > 0
        elif results is None:
            has_results = False
        else:
            try:
                has_results = len(results) > 0
            except TypeError:
                has_results = True
        self._has_unsaved_changes = bool(unsaved) and has_results
        if hasattr(self, "saveBtn"):
            if self._has_unsaved_changes:
                self.saveBtn.Enable()
            else:
                self.saveBtn.Disable()

    def _mark_results_modified(self):
        self._mark_unsaved(True)

    def on_param_unit_type_change(self, event):
        self._update_param_unit_scale_choices()
        self._refresh_plot_labels()

    def on_param_unit_scale_change(self, event):
        unit_type = self.paramUnitTypeChoice.GetStringSelection()
        self.param_unit_scale = self.paramUnitScaleChoice.GetStringSelection()
        self.param_scale = self.unit_factors[unit_type][self.param_unit_scale]
        self._refresh_plot_labels()

    def _update_param2_unit_scale_choices(self):
        unit_type = self.param2UnitTypeChoice.GetStringSelection()
        scales = list(self.unit_factors[unit_type].keys())
        self.param2UnitScaleChoice.Set(scales)
        index = 1 if unit_type == "time" and len(scales) > 1 else 0
        self.param2UnitScaleChoice.SetSelection(index)
        self.param2_unit_scale = scales[index]
        self.param2_scale = self.unit_factors[unit_type][self.param2_unit_scale]

    def on_param2_unit_type_change(self, event):
        self._update_param2_unit_scale_choices()
        self._refresh_plot_labels()

    def on_param2_unit_scale_change(self, event):
        unit_type = self.param2UnitTypeChoice.GetStringSelection()
        self.param2_unit_scale = self.param2UnitScaleChoice.GetStringSelection()
        self.param2_scale = self.unit_factors[unit_type][self.param2_unit_scale]
        self._refresh_plot_labels()

    def on_multi_fit_choice(self, event):
        self._update_multi_fit_controls()
        if event is not None:
            event.Skip()

    def on_mode_change(self, event):
        # Prevent intermediate visibility changes while we rebuild UI state
        self._suspend_fit_visibility = True
        freeze_succeeded = False
        try:
            # Freeze redraws while we mutate the UI
            try:
                self.Freeze()
                freeze_succeeded = True
            except Exception:
                pass
            # Proactively hide both guess panels so parent container toggles can't flash them
            self._force_hide_param_guesses()
            # Avoid a brief "enabled" caption state from prior mode
            try:
                self.captionCtrl1.Disable()
                self.captionCtrl2.Disable()
            except Exception:
                pass

            is_variable = self.singleRadio.GetValue()
            is_sweep = self.multiRadio.GetValue()
            is_heat = self.heatmapRadio.GetValue()

            new_mode = "heat" if is_heat else "multi" if is_sweep else "single"
            self.funcChoice1.SetSelection(0)
            self.funcChoice2.SetSelection(0)
            self._last_func_selection[1] = ""
            self._last_func_selection[2] = ""

            residual_checkbox = getattr(self, "residualCheck", None)
            if not is_variable and residual_checkbox and residual_checkbox.GetValue():
                residual_checkbox.SetValue(False)

            show_param1 = not is_variable
            show_param2 = is_heat
            for w in [
                self.paramLabelStatic,
                self.paramLabelCtrl,
                self.paramUnitTypeChoice,
                self.paramUnitScaleChoice,
                self.paramFileLabel,
                self.paramFileCtrl,
                self.chooseParamFileBtn,
                self.paramOffsetLabel,
                self.paramOffsetCtrl,
                self.paramListLabel,
                self.paramListCtrl,
            ]:
                w.Show(show_param1)
            for w in [
                self.param2LabelStatic,
                self.param2LabelCtrl,
                self.param2UnitTypeChoice,
                self.param2UnitScaleChoice,
                self.param2FileLabel,
                self.param2FileCtrl,
                self.chooseParam2FileBtn,
                self.param2OffsetLabel,
                self.param2OffsetCtrl,
                self.param2ListLabel,
                self.param2ListCtrl,
            ]:
                w.Show(show_param2)
            mf = show_param1 or show_param2
            self.multiFitLabel.SetLabel(
                "Variable Fit:" if (is_sweep or is_heat) else "Multi-fit:"
            )
            self.multiFitLabel.Show(mf)
            self.multiFitChoice.Show(mf)
            self.param_col_item.Show(show_param1)
            self.param2_col_item.Show(show_param2)
            self.plot2_box_item.Show(True)
            self.overlayCheck.Show(not is_heat)
            self.paramBinCheck.Show(not is_heat)
            show_var_bin = is_sweep and not is_heat
            self.varBinCheck.Show(show_var_bin)
            if hasattr(self, "heatmapVarBinCheck"):
                self.heatmapVarBinCheck.Show(is_heat)
            self.residualCheck.Show(not is_heat)
            show_fit_controls = not is_heat
            for ctrl in [
                self.scaleLabel1,
                self.scaleChoice1,
                self.fitLabel1,
                self.funcChoice1,
                self.scaleLabel2,
                self.scaleChoice2,
                self.fitLabel2,
                self.funcChoice2,
            ]:
                ctrl.Show(show_fit_controls)
            if is_heat:
                self.overlayCheck.SetValue(False)
                self.paramBinCheck.SetValue(False)
                self.varBinCheck.SetValue(False)
                self.residualCheck.SetValue(False)
                if hasattr(self, "heatmapVarBinCheck"):
                    self.heatmapVarBinCheck.SetValue(False)
            self.populate_param_controls(1)
            self.populate_param_controls(2)
            self._sync_overlay_color_control()
            self._refresh_multi_fit_choices()
            self._update_multi_fit_controls()
            base_size = (
                self.heatmap_size
                if is_heat
                else self.multi_size if not is_variable else self.single_size
            )
            self._apply_mode_size(base_size)

            self._current_mode = new_mode

            var_path = self.varFileCtrl.GetValue()
            if var_path:
                self.load_variable_file(var_path)
            param_path = self.paramFileCtrl.GetValue()
            if show_param1 and param_path:
                self.load_parameter_file(param_path)
            param2_path = self.param2FileCtrl.GetValue()
            if show_param2 and param2_path:
                self.load_parameter2_file(param2_path)

            # Reset processing state and plots when switching modes
            self.results = None
            self.var_values = []
            self.param_values = []
            self.param2_values = []
            self.subrun_results = []
            self.heatmap_axes_info = {}
            self.heatmap_index_grid = None
            self.heatmap_active_plots.clear()
            self.heatmap_active_columns.clear()
            self.plot_data.clear()
            self._clear_overlay_table_axes()
            self._clear_overlay_right_axis()
            self._clear_heatmap_colorbars()
            self._clear_heatmap_detail_panels()
            self._ensure_axes(1)
            self.axes[0].clear()
            self.canvas.draw()
            self.overlayCheck.SetValue(False)
            self.on_overlay_toggle(None)
            self.closeBtn1.Disable()
            self.closeBtn2.Disable()
            self.progress.SetForegroundColour(self._gauge_default_colour)
            self.progress.SetValue(0)
            self.processBtn.SetLabel("Process")
            self._processing = False
            self._stop_requested = False
            self.colChoice1.SetSelection(0)
            self.colChoice2.SetSelection(0)
            self.scaleChoice1.SetSelection(0)
            self.scaleChoice2.SetSelection(0)
            self._update_func_choices_for_plot(1)
            self._update_func_choices_for_plot(2)
            self.funcChoice1.SetSelection(0)
            self.funcChoice2.SetSelection(0)
            self._last_func_selection[1] = ""
            self._last_func_selection[2] = ""
            self.populate_param_controls(1)
            self.populate_param_controls(2)
            # Re-enable updates; keep guess panels hard-hidden after a mode switch
            self._suspend_fit_visibility = False
            self._force_hide_param_guesses()
            self._last_column_selection[1] = self.colChoice1.GetStringSelection()
            self._last_column_selection[2] = self.colChoice2.GetStringSelection()
            self._last_func_selection[1] = _normalise_fit_name(
                self.funcChoice1.GetStringSelection()
            )
            self._last_func_selection[2] = _normalise_fit_name(
                self.funcChoice2.GetStringSelection()
            )
            self._reset_caption_inputs(1, refresh_display=False)
            self._reset_caption_inputs(2, refresh_display=False)
            self.paramBinCheck.SetValue(False)
            self.varBinCheck.SetValue(False)
            self.variable_bin_groups = {}
            self._refresh_exclude_panel_visibility()
            self._set_controls_enabled(False)
            self._mark_unsaved(False)
            self._update_residual_control()
            # Final layout & redraw in one shot, then thaw
            self.control_panel.Layout()
            self.panel.Layout()
        finally:
            if self._suspend_fit_visibility:
                self._suspend_fit_visibility = False
            if freeze_succeeded:
                try:
                    self.Thaw()
                except Exception:
                    pass

    def _set_controls_enabled(self, enabled):
        processing = getattr(self, "_processing", False)
        base_enabled = bool(enabled) and not processing
        general_enabled = not processing

        for ctrl in getattr(self, "_processing_sensitive_controls", []):
            if ctrl is not None:
                try:
                    ctrl.Enable(general_enabled)
                except Exception:
                    pass

        ctrl_candidates = [
            self.colChoice1,
            self.colChoice2,
            self.funcChoice1,
            self.funcChoice2,
            self.scaleChoice1,
            self.scaleChoice2,
            self.plotBtn1,
            self.plotBtn2,
            self.captionCtrl1,
            self.captionCtrl2,
            self.captionLbl1,
            self.captionLbl2,
            self.paramBinCheck,
            self.varBinCheck,
            getattr(self, "heatmapVarBinCheck", None),
            self.overlayCheck,
            self.overlayColorCheck,
            self.saveBtn,
        ]
        for ctrl in ctrl_candidates:
            if ctrl is not None:
                ctrl.Enable(base_enabled)
        for ctrl in self.paramCtrls1 + self.paramCtrls2:
            ctrl.Enable(base_enabled)
        for chk in self.exclude_checks.values():
            chk.Enable(base_enabled)
        for chk in self.detail_checks.values():
            chk.Enable(base_enabled)

        is_variable = self.singleRadio.GetValue()
        is_heat = self.heatmapRadio.GetValue()
        if is_variable:
            has_results = self.results is not None
            for ctrl in [
                self.colChoice2,
                self.funcChoice2,
                self.scaleChoice2,
                self.plotBtn2,
                self.captionCtrl2,
                self.captionLbl2,
                self.overlayCheck,
            ]:
                ctrl.Enable(base_enabled and has_results)
            self.plot2_box_item.Show(True)
            self.overlayCheck.Show(True)
        elif is_heat:
            for ctrl in [
                self.funcChoice1,
                self.funcChoice2,
                self.scaleChoice1,
                self.scaleChoice2,
                self.overlayCheck,
                self.paramBinCheck,
                self.varBinCheck,
            ]:
                ctrl.Enable(False)
            heatmap_bin_ctrl = getattr(self, "heatmapVarBinCheck", None)
            if heatmap_bin_ctrl is not None:
                heatmap_bin_ctrl.Enable(base_enabled)
        else:
            for ctrl in [
                self.overlayCheck,
                self.paramBinCheck,
                self.varBinCheck,
                self.funcChoice1,
                self.funcChoice2,
                self.scaleChoice1,
                self.scaleChoice2,
                self.colChoice2,
                self.plotBtn2,
                self.captionCtrl2,
                self.captionLbl2,
            ]:
                ctrl.Enable(base_enabled)
            self.plot2_box_item.Show(True)
            self.overlayCheck.Show(True)
        allow_bin = base_enabled and not is_heat
        self._update_param_bin_control(allow_bin)
        self._update_var_bin_control(allow_bin)
        self._update_heatmap_var_bin_control(base_enabled)
        self.control_panel.Layout()
        self._sync_overlay_control()
        self._update_multi_fit_controls()
        self._update_residual_control()

    def _can_show_residuals(self):
        if getattr(self, "_processing", False):
            return False
        if self._heatmap_mode_active():
            return False
        if not getattr(self, "plot_data", None):
            return False
        for data in self.plot_data.values():
            if not data:
                continue
            x_vals = data.get("residual_x")
            y_vals = data.get("residual_y")
            if x_vals is None or y_vals is None:
                continue
            try:
                x_arr = np.asarray(x_vals, dtype=float)
                y_arr = np.asarray(y_vals, dtype=float)
            except Exception:
                continue
            if x_arr.size and y_arr.size:
                return True
        return False

    def _update_residual_control(self):
        checkbox = getattr(self, "residualCheck", None)
        if checkbox is None:
            return
        can_show = self._can_show_residuals()
        if not can_show and checkbox.GetValue():
            checkbox.SetValue(False)
        checkbox.Enable(can_show)

    def _has_processed_results(self):
        if isinstance(self.results, dict):
            for value in self.results.values():
                if isinstance(value, (list, tuple, np.ndarray)) and len(value):
                    return True
        if isinstance(self.subrun_results, (list, tuple)):
            for run in self.subrun_results:
                if not isinstance(run, dict):
                    continue
                results = run.get("results")
                if not isinstance(results, dict):
                    continue
                for value in results.values():
                    if isinstance(value, (list, tuple, np.ndarray)) and len(value):
                        return True
        if isinstance(self.plot_data, dict):
            for data in self.plot_data.values():
                if not isinstance(data, dict):
                    continue
                for key in ("plot_x", "plot_y", "plot_x_fit", "plot_y_fit"):
                    arr = data.get(key)
                    if isinstance(arr, (list, tuple, np.ndarray)) and len(arr):
                        return True
        return False

    def _has_repeated_values(self, values):
        if not values:
            return False
        try:
            arr = np.asarray(values, dtype=float)
        except Exception:
            return False
        if arr.size == 0:
            return False
        try:
            _, counts = np.unique(arr, return_counts=True)
        except Exception:
            return False
        return bool(np.any(counts > 1))

    def _heatmap_var_binning_enabled(self):
        checkbox = getattr(self, "heatmapVarBinCheck", None)
        if checkbox is None:
            return False
        return bool(self._heatmap_mode_active() and checkbox.GetValue())

    def _update_param_bin_control(self, base_enabled=True):
        if not hasattr(self, "paramBinCheck"):
            return
        can_bin = False
        if base_enabled and not self._processing and self.results is not None:
            if not self.singleRadio.GetValue() and self.param_values:
                can_bin = self._has_repeated_values(self.param_values)
            else:
                can_bin = self._has_repeated_values(self.var_values)
        label = "Variable Binning" if self.singleRadio.GetValue() else "Parameter Binning"
        if self.paramBinCheck.GetLabel() != label:
            self.paramBinCheck.SetLabel(label)
        if not can_bin and self.paramBinCheck.GetValue():
            self.paramBinCheck.SetValue(False)
            self.on_param_bin_toggle(None)
        self.paramBinCheck.Enable(can_bin)

    def _update_var_bin_control(self, base_enabled=True):
        if not hasattr(self, "varBinCheck"):
            return
        can_bin = False
        if (
            base_enabled
            and not self._processing
            and self.subrun_results
            and self.multiRadio.GetValue()
            and not self.heatmapRadio.GetValue()
        ):
            for run in self.subrun_results:
                if self._has_repeated_values(run.get("var_values", [])):
                    can_bin = True
                    break
        if not can_bin and self.varBinCheck.GetValue():
            self.varBinCheck.SetValue(False)
            self.on_var_bin_toggle(None)
        self.varBinCheck.Enable(can_bin)

    def _update_heatmap_var_bin_control(self, base_enabled=True):
        checkbox = getattr(self, "heatmapVarBinCheck", None)
        if checkbox is None:
            return
        can_bin = False
        if (
            base_enabled
            and not self._processing
            and self._heatmap_mode_active()
            and self.subrun_results
        ):
            for run in self.subrun_results:
                if self._has_repeated_values(run.get("var_values", [])):
                    can_bin = True
                    break
        if not can_bin and checkbox.GetValue():
            checkbox.SetValue(False)
            self.on_heatmap_var_bin_toggle(None)
        checkbox.Enable(can_bin)

    def _sync_overlay_control(self):
        can_overlay = (
            self.overlayCheck.IsShown()
            and len(self.plot_data) >= 2
            and self.results is not None
            and not self.heatmapRadio.GetValue()
            and self._plots_share_axis_context()
        )
        overlay_active = self.overlayCheck.GetValue()
        if not can_overlay:
            if overlay_active:
                self.overlayCheck.SetValue(False)
                self.on_overlay_toggle(None)
            self.overlayCheck.Enable(False)
        else:
            self.overlayCheck.Enable(not self._processing)
        overlay_active = self.overlayCheck.GetValue()
        self._sync_overlay_color_control(can_overlay=can_overlay, overlay_active=overlay_active)

    def _sync_overlay_color_control(self, *, can_overlay=None, overlay_active=None):
        if not hasattr(self, "overlayColorCheck"):
            return

        if can_overlay is None:
            can_overlay = (
                self.overlayCheck.IsShown()
                and len(self.plot_data) >= 2
                and self.results is not None
                and not self.heatmapRadio.GetValue()
                and self._plots_share_axis_context()
            )

        if overlay_active is None:
            overlay_active = self.overlayCheck.GetValue()

        should_show = can_overlay and overlay_active
        if should_show:
            self.overlayColorCheck.Show()
            self.overlayColorCheck.Enable(not self._processing)
        else:
            if self.overlayColorCheck.GetValue():
                self.overlayColorCheck.SetValue(False)
            self.overlayColorCheck.Enable(False)
            self.overlayColorCheck.Hide()

        sizer = self.overlayColorCheck.GetContainingSizer()
        if sizer is not None:
            sizer.Layout()
        if getattr(self, "control_panel", None) is not None:
            try:
                self.control_panel.Layout()
            except Exception:
                pass

    def _bind_enter_action(self, ctrl, callback, allow_shift_newline=False):
        """Invoke ``callback`` when Enter is pressed on ``ctrl``.

        ``wx.EVT_TEXT_ENTER`` is only generated for single-line text controls,
        so multi-line controls additionally listen for the key-down event. A
        plain Enter commits the change while Shift+Enter still allows new
        lines in captions.
        """

        def trigger():
            callback()

        multiline = bool(ctrl.GetWindowStyleFlag() & wx.TE_MULTILINE)

        if not (multiline and allow_shift_newline):

            def on_text_enter(event):
                trigger()
                event.Skip()

            ctrl.Bind(wx.EVT_TEXT_ENTER, on_text_enter)

        if multiline:

            def on_key_down(event):
                key = event.GetKeyCode()
                if key in (wx.WXK_RETURN, wx.WXK_NUMPAD_ENTER):
                    if allow_shift_newline:
                        meta_down = getattr(event, "MetaDown", lambda: False)()
                        if (
                            event.ShiftDown()
                            and not event.ControlDown()
                            and not event.AltDown()
                            and not meta_down
                        ):
                            event.Skip()
                            return
                    trigger()
                    return
                event.Skip()

            ctrl.Bind(wx.EVT_KEY_DOWN, on_key_down)

    def _get_caption_ctrl(self, plot_num):
        return self.captionCtrl1 if plot_num == 1 else self.captionCtrl2

    def _reset_caption_inputs(self, plot_num, refresh_display=True):
        ctrl = self._get_caption_ctrl(plot_num)
        ctrl.ChangeValue("")
        if plot_num in self.plot_data:
            self.plot_data[plot_num]["caption"] = ""
        if refresh_display and not self._heatmap_mode_active():
            self.redraw_plots()

    def _after_caption_update(self, plot_num):
        if self._heatmap_mode_active():
            self.update_heatmap()
        else:
            self.redraw_plots()
        self._mark_results_modified()

    def on_column_change(self, plot_num, event=None):
        if getattr(self, "_suspend_fit_visibility", False):
            return
        choice = self.colChoice1 if plot_num == 1 else self.colChoice2
        selection = choice.GetStringSelection()
        self._last_column_selection[plot_num] = selection
        self._update_func_choices_for_plot(plot_num)
        self.populate_param_controls(plot_num)
        if event is not None:
            event.Skip()

    def on_axis_label_enter(self, event=None):
        self._refresh_plot_labels()
        if event is not None:
            event.Skip()

    def _filter_fit_choices_for_plot(self, scale_name, column_name):
        scale_mode = (scale_name or SCALE_OPTIONS[0]).lower()
        disallowed = set()
        if getattr(self, "multiRadio", None) and self.multiRadio.GetValue():
            disallowed.update(
                {
                    _normalise_fit_name(TEMP_PSD_DISPLAY_NAME),
                    _normalise_fit_name(MOT_LIFETIME_DISPLAY_NAME),
                    _normalise_fit_name(MOT_RINGDOWN_DISPLAY_NAME),
                }
            )
        if scale_mode in ("semi-log", "log"):
            disallowed.update(
                {
                    _normalise_fit_name(TEMP_PSD_DISPLAY_NAME),
                    _normalise_fit_name(MOT_LIFETIME_DISPLAY_NAME),
                    _normalise_fit_name(MOT_RINGDOWN_DISPLAY_NAME),
                }
            )
        col_key = (column_name or "").lower()
        if col_key in {"atom number", "x-atom number", "y-atom number"}:
            disallowed.update(
                {
                    _normalise_fit_name(TEMP_PSD_DISPLAY_NAME),
                    _normalise_fit_name(MOT_RINGDOWN_DISPLAY_NAME),
                }
            )
        elif col_key in {"x-center", "y-center"}:
            disallowed.update(
                {
                    _normalise_fit_name(TEMP_PSD_DISPLAY_NAME),
                    _normalise_fit_name(MOT_LIFETIME_DISPLAY_NAME),
                    _normalise_fit_name(MOT_RINGDOWN_DISPLAY_NAME),
                }
            )
        elif col_key in {"x-true width", "y-true width"}:
            disallowed.update(
                {
                    _normalise_fit_name(MOT_LIFETIME_DISPLAY_NAME),
                    _normalise_fit_name(MOT_RINGDOWN_DISPLAY_NAME),
                }
            )
        if scale_mode == "linear":
            base_choices = list(self._base_fit_choices)
        else:
            base_choices = [
                name
                for name in self._base_fit_choices
                if _normalise_fit_name(name) not in LINEAR_SCALE_ONLY_FITS
            ]
            if not base_choices:
                base_choices = [""]
        filtered = [
            name
            for name in base_choices
            if _normalise_fit_name(name) not in disallowed
        ]
        return filtered if filtered else [""]

    def _update_func_choices_for_plot(self, plot_num):
        choice = self.funcChoice1 if plot_num == 1 else self.funcChoice2
        scale_choice = self.scaleChoice1 if plot_num == 1 else self.scaleChoice2
        scale_selection = scale_choice.GetStringSelection() or SCALE_OPTIONS[0]
        col_choice = self.colChoice1 if plot_num == 1 else self.colChoice2
        column_selection = col_choice.GetStringSelection()
        allowed = self._filter_fit_choices_for_plot(scale_selection, column_selection)
        current_label = choice.GetStringSelection()
        current_normalized = _normalise_fit_name(current_label)
        restricted = (scale_selection or "").lower() != "linear" and (
            current_normalized in LINEAR_SCALE_ONLY_FITS
        )
        options = [""] + [name for name in allowed if name]
        choice.Freeze()
        choice.Set(options)
        if current_label and current_label in options and not restricted:
            choice.SetStringSelection(current_label)
        else:
            choice.SetSelection(0)
            restricted = restricted or current_label != choice.GetStringSelection()
        choice.Thaw()
        if restricted:
            data = self.plot_data.get(plot_num)
            if data:
                data["func_name"] = None
                data["p0"] = tuple()
        if restricted:
            self.on_func_change(plot_num)
            if plot_num in self.plot_data:
                self._mark_results_modified()
        if not getattr(self, "_suspend_fit_visibility", False):
            self._set_param_guess_visibility(
                plot_num, self._should_show_param_guesses(plot_num)
            )
        self.control_panel.Layout()

    def on_scale_change(self, plot_num):
        if getattr(self, "_suspend_fit_visibility", False):
            return
        if self.heatmapRadio.GetValue():
            return
        choice = self.scaleChoice1 if plot_num == 1 else self.scaleChoice2
        selection = choice.GetStringSelection() or SCALE_OPTIONS[0]
        if plot_num in self.plot_data:
            self.plot_data[plot_num]["axis_scale"] = selection
            self.plot_data[plot_num]["axis_scale_mode"] = self._normalise_axis_scale(selection)
        self._update_func_choices_for_plot(plot_num)
        if plot_num in self.plot_data:
            self.redraw_plots()
            self._mark_results_modified()

    def on_caption_change(self, plot_num, _event=None):
        ctrl = self._get_caption_ctrl(plot_num)
        text = ctrl.GetValue()
        if plot_num in self.plot_data:
            self.plot_data[plot_num]["caption"] = text
        self._after_caption_update(plot_num)

    def on_save_path_text(self, event):
        self.savePath = event.GetEventObject().GetValue()

    def on_choose_save_path(self, event):
        dialog = wx.DirDialog(
            None,
            "Choose a directory:",
            defaultPath=self.savePath,
            style=wx.DD_DEFAULT_STYLE | wx.DD_NEW_DIR_BUTTON,
        )
        if dialog.ShowModal() == wx.ID_OK:
            self.savePath = dialog.GetPath()
            self.savePathCtrl.SetValue(self.savePath)
        dialog.Destroy()

    def on_load_session(self, event):
        """Restore a previously exported fitting session."""

        # Persist the current state so users can return to it later.
        try:
            self._save_user_settings()
        except Exception as err:
            print(f"Failed to save fitting window settings before load: {err}")
        parent = getattr(self, "parent", None)
        if parent is not None:
            parent_saver = getattr(parent, "_save_settings", None)
            if callable(parent_saver):
                try:
                    parent_saver()
                except Exception as err:
                    print(f"Failed to save image UI settings before load: {err}")

        dialog = wx.FileDialog(
            self,
            "Select a fitting session file",
            defaultDir=self.savePath or getattr(self.parent, "path", ""),
            wildcard=(
                "Session files (*.json;*.txt;*.csv;*.png;*.xlsx)|"
                "*.json;*.txt;*.csv;*.png;*.xlsx|All files (*.*)|*.*"
            ),
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST,
        )
        if dialog.ShowModal() != wx.ID_OK:
            dialog.Destroy()
            return

        artifact_path = dialog.GetPath()
        dialog.Destroy()
        base_name = self._deduce_session_base(artifact_path)
        if not base_name:
            wx.MessageBox(
                "Unable to determine the session name from the selected file.",
                "Load Run",
                style=wx.ICON_WARNING | wx.OK,
            )
            return

        folder = os.path.dirname(artifact_path)
        self.savePath = folder
        self.savePathCtrl.SetValue(folder)

        loaded = []
        missing = []
        errors = []

        session_payload = None
        session_json_paths = [
            os.path.join(folder, f"{base_name}.json"),
            os.path.join(folder, f"{base_name}-session.json"),
        ]
        for session_json_path in session_json_paths:
            if not os.path.isfile(session_json_path):
                continue
            try:
                with open(session_json_path, "r", encoding="utf-8") as fh:
                    candidate = json.load(fh)
            except Exception as err:
                errors.append(f"Session bundle: {err}")
                continue
            if not isinstance(candidate, dict):
                errors.append("Session bundle: unexpected format")
                continue
            session_payload = candidate
            break

        session_bundle_loaded = False
        if session_payload is not None:
            session_bundle_loaded = True
            loaded.append("Session bundle")

            def _apply_sequence(values, control, label):
                if control is None:
                    return False
                if values is None:
                    try:
                        control.ChangeValue("")
                    except Exception as err:
                        errors.append(f"{label}: {err}")
                    return False
                try:
                    normalised = []
                    for item in values:
                        normalised.append("" if item is None else str(item))
                    control.ChangeValue("\n".join(normalised))
                except Exception as err:
                    errors.append(f"{label}: {err}")
                    return False
                return True

            fitting_section = session_payload.get("fitting", {})
            settings = fitting_section.get("settings")
            if isinstance(settings, dict):
                try:
                    self._apply_user_settings(settings)
                except Exception as err:
                    errors.append(f"Fitting settings: {err}")
                else:
                    self._user_settings = settings
                    loaded.append("Fitting settings")
            else:
                missing.append("Fitting settings")

            if _apply_sequence(
                fitting_section.get("variable_sequence"),
                getattr(self, "varListCtrl", None),
                "Variable list",
            ):
                self.varFileCtrl.ChangeValue("")
                loaded.append("Variable list")
            else:
                missing.append("Variable list")

            if _apply_sequence(
                fitting_section.get("parameter_sequence"),
                getattr(self, "paramListCtrl", None),
                "Parameter list",
            ):
                self.paramFileCtrl.ChangeValue("")
                loaded.append("Parameter list")
            else:
                missing.append("Parameter list")

            if _apply_sequence(
                fitting_section.get("parameter2_sequence"),
                getattr(self, "param2ListCtrl", None),
                "Parameter 2 list",
            ):
                self.param2FileCtrl.ChangeValue("")
                loaded.append("Parameter 2 list")
            else:
                missing.append("Parameter 2 list")

            raw_exclusions = fitting_section.get("excluded_indices")
            if raw_exclusions is not None:
                try:
                    exclusions = {
                        int(val)
                        for val in raw_exclusions
                        if isinstance(val, (int, float, str)) and str(val).strip()
                    }
                except Exception:
                    exclusions = set()
                    for val in raw_exclusions:
                        try:
                            exclusions.add(int(val))
                        except Exception:
                            continue
                self._session_point_exclusions = set(exclusions)
                self._pending_session_point_exclusions = set(exclusions)
                if exclusions:
                    loaded.append("Point exclusions")
            else:
                self._session_point_exclusions = set()
                self._pending_session_point_exclusions = set()

            run_exclusions = fitting_section.get("run_exclusions")
            if isinstance(run_exclusions, dict):
                parsed_runs = {}
                for key, values in run_exclusions.items():
                    try:
                        run_idx = int(key)
                    except Exception:
                        continue
                    try:
                        parsed = {
                            int(val)
                            for val in values
                            if isinstance(val, (int, float, str)) and str(val).strip()
                        }
                    except Exception:
                        parsed = set()
                        for val in values or []:
                            try:
                                parsed.add(int(val))
                            except Exception:
                                continue
                    if parsed:
                        parsed_runs[run_idx] = parsed
                self._session_run_exclusions = {
                    idx: set(vals) for idx, vals in parsed_runs.items()
                }
                self._pending_session_run_exclusions = {
                    idx: set(vals) for idx, vals in parsed_runs.items()
                }
                if parsed_runs:
                    loaded.append("Run exclusions")
            else:
                self._session_run_exclusions = {}
                self._pending_session_run_exclusions = {}

            image_section = session_payload.get("image", {})
            if isinstance(image_section, dict) and image_section:
                image_settings = image_section.get("settings")
                applied_settings = False
                if isinstance(image_settings, dict) and parent is not None:
                    applier = getattr(parent, "apply_settings_snapshot", None)
                    if callable(applier):
                        try:
                            applier(copy.deepcopy(image_settings))
                        except Exception as err:
                            errors.append(f"Applying image settings: {err}")
                        else:
                            loaded.append("Image settings")
                            applied_settings = True
                    if not applied_settings:
                        fallback = getattr(parent, "_apply_settings", None)
                        if callable(fallback):
                            try:
                                parent._settings = copy.deepcopy(image_settings)
                                fallback()
                                saver = getattr(parent, "_save_settings", None)
                                if callable(saver):
                                    saver()
                            except Exception as err:
                                errors.append(f"Applying image settings: {err}")
                            else:
                                loaded.append("Image settings")
                                applied_settings = True
                    if applied_settings:
                        self._parent_settings_modified = True
                elif image_settings:
                    missing.append("Image settings (no active image window)")

                folder_hint = image_section.get("folder")
                display_hint = image_section.get("folder_display")
                if not folder_hint:
                    image_file_hint = image_section.get("image_file")
                    if image_file_hint:
                        folder_hint = self._normalise_image_folder(
                            image_file_hint, folder
                        )
                if folder_hint:
                    if self._apply_parent_image_folder(
                        folder_hint, display_value=display_hint
                    ):
                        loaded.append("Image folder")
                    else:
                        missing.append("Image folder (failed to update main window)")
                elif image_section:
                    missing.append("Image folder")

        if not session_bundle_loaded:
            self._session_point_exclusions = set()
            self._pending_session_point_exclusions = set()
            self._session_run_exclusions = {}
            self._pending_session_run_exclusions = {}
            settings_path = os.path.join(folder, f"{base_name}-fitting-settings.json")
            if os.path.isfile(settings_path):
                try:
                    with open(settings_path, "r", encoding="utf-8") as fh:
                        settings = json.load(fh)
                except Exception as err:
                    errors.append(f"Fitting settings: {err}")
                else:
                    self._apply_user_settings(settings)
                    start_file = settings.get("start_file")
                    if isinstance(start_file, str):
                        self.startFileCtrl.ChangeValue(start_file)
                    self._user_settings = settings
                    loaded.append("Fitting settings")
            else:
                missing.append("Fitting settings")

            var_path = os.path.join(folder, f"{base_name}-variable.txt")
            if os.path.isfile(var_path):
                self.varFileCtrl.SetValue(var_path)
                self.load_variable_file(var_path)
                loaded.append("Variable list")
            else:
                missing.append("Variable list")

            param_path = os.path.join(folder, f"{base_name}-parameter.txt")
            if os.path.isfile(param_path):
                self.paramFileCtrl.SetValue(param_path)
                self.load_parameter_file(param_path)
                loaded.append("Parameter list")
            else:
                missing.append("Parameter list")

            param2_path = os.path.join(folder, f"{base_name}-parameter2.txt")
            if os.path.isfile(param2_path):
                self.param2FileCtrl.SetValue(param2_path)
                self.load_parameter2_file(param2_path)
                loaded.append("Parameter 2 list")
            else:
                missing.append("Parameter 2 list")

            image_settings_path = os.path.join(
                folder, f"{base_name}-image-settings.json"
            )
            if os.path.isfile(image_settings_path):
                try:
                    with open(image_settings_path, "r", encoding="utf-8") as fh:
                        image_settings = json.load(fh)
                except Exception as err:
                    errors.append(f"Image settings: {err}")
                else:
                    applied = False
                    if parent is not None:
                        applier = getattr(parent, "apply_settings_snapshot", None)
                        if callable(applier):
                            try:
                                applier(image_settings)
                            except Exception as err:
                                errors.append(f"Applying image settings: {err}")
                            else:
                                loaded.append("Image settings")
                                applied = True
                        if not applied:
                            fallback = getattr(parent, "_apply_settings", None)
                            if callable(fallback):
                                try:
                                    parent._settings = copy.deepcopy(image_settings)
                                    fallback()
                                    saver = getattr(parent, "_save_settings", None)
                                    if callable(saver):
                                        saver()
                                except Exception as err:
                                    errors.append(f"Applying image settings: {err}")
                                else:
                                    loaded.append("Image settings")
                                    applied = True
                    if applied:
                        self._parent_settings_modified = True
                    if not applied:
                        missing.append("Image settings (no active image window)")
            else:
                missing.append("Image settings")

            csv_path = self._find_session_csv(base_name, folder)
            if csv_path:
                image_file = self._extract_image_file_from_csv(csv_path)
                if image_file:
                    image_folder = self._normalise_image_folder(image_file, folder)
                    if image_folder:
                        if self._apply_parent_image_folder(image_folder):
                            loaded.append("Image folder")
                        else:
                            missing.append(
                                "Image folder (failed to update main window)"
                            )
                    else:
                        missing.append("Image folder (unrecognised path)")
                else:
                    missing.append("Image file reference in CSV")
            else:
                missing.append("Session CSV")

        summary_lines = []
        if loaded:
            summary_lines.append("Loaded: " + ", ".join(sorted(set(loaded))))
        if missing:
            summary_lines.append("Missing: " + ", ".join(sorted(set(missing))))
        if errors:
            summary_lines.append("Errors: " + "; ".join(errors))

        if summary_lines:
            style = wx.ICON_INFORMATION if loaded and not errors else wx.ICON_WARNING
            wx.MessageBox("\n".join(summary_lines), "Load Run", style=style | wx.OK)

        if event is not None:
            event.Skip()

    def load_variable_file(self, path):
        """Load variable values from a text file into the list control."""
        try:
            with open(path, "r") as f:
                content = f.read().strip()
            self.varListCtrl.SetValue(content)
        except OSError:
            pass

    def on_var_file_text(self, event):
        path = event.GetEventObject().GetValue()
        if path:
            self.load_variable_file(path)

    def on_choose_var_file(self, event):
        dialog = wx.FileDialog(
            self,
            "Choose variable list file",
            wildcard="Text files (*.txt)|*.txt|All files (*.*)|*.*",
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST,
        )
        if dialog.ShowModal() == wx.ID_OK:
            path = dialog.GetPath()
            self.varFileCtrl.SetValue(path)
            self.load_variable_file(path)
        dialog.Destroy()

    def _find_session_csv(self, base_name, folder):
        candidates = [
            os.path.join(folder, f"{base_name}.csv"),
            os.path.join(folder, f"{base_name}-session.csv"),
            os.path.join(folder, f"{base_name}-multirun.csv"),
            os.path.join(folder, f"{base_name}-heatmap.csv"),
        ]
        for path in candidates:
            if os.path.isfile(path):
                return path
        return None

    @staticmethod
    def _strip_sequence_header(text):
        if not text:
            return ""

        lines = text.splitlines()
        first_idx = None
        for idx, line in enumerate(lines):
            if line.strip():
                first_idx = idx
                break
        if first_idx is None:
            return ""

        first_line = lines[first_idx].strip()
        tokens = [tok for tok in re.split(r"[\s,]+", first_line) if tok]
        numeric_first_line = False
        if tokens:
            numeric_first_line = True
            for tok in tokens:
                try:
                    float(tok)
                except ValueError:
                    numeric_first_line = False
                    break

        if numeric_first_line:
            relevant_lines = lines[first_idx:]
        else:
            relevant_lines = lines[first_idx + 1 :]

        return "\n".join(line.rstrip() for line in relevant_lines).strip()

    @staticmethod
    def _serialise_sequence_from_text(text):
        if not text:
            return []

        normalised = text.replace("\r\n", "\n").replace("\r", "\n").strip()
        if not normalised:
            return []
        return normalised.split("\n")

    @staticmethod
    def _deduce_session_base(path):
        if not path:
            return None
        name = os.path.basename(path)
        lowered = name.lower()
        suffixes = (
            "-session.json",
            "-fitting-settings.json",
            "-image-settings.json",
            "-variable.txt",
            "-parameter.txt",
            "-parameter2.txt",
        )
        for suffix in suffixes:
            if lowered.endswith(suffix):
                return name[: -len(suffix)]
        base, _ = os.path.splitext(name)
        return base or name

    @staticmethod
    def _extract_image_file_from_csv(path):
        try:
            with open(path, "r", encoding="utf-8", newline="") as fh:
                reader = csv.reader(fh)
                headers = next(reader, None)
                if not headers:
                    return None
                target_idx = None
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    key = header.strip().lower().replace(" ", "").replace("_", "")
                    if key == "imagefile":
                        target_idx = idx
                        break
                if target_idx is None:
                    return None
                for row in reader:
                    if target_idx >= len(row):
                        continue
                    value = (row[target_idx] or "").strip()
                    if value:
                        return value
        except Exception:
            return None
        return None

    @staticmethod
    def _normalise_image_folder(image_path, session_folder):
        if not image_path:
            return None
        cleaned = image_path.strip().strip('"')
        if not cleaned:
            return None
        if not os.path.isabs(cleaned):
            cleaned = os.path.normpath(os.path.join(session_folder, cleaned))
        directory = cleaned
        if not os.path.isdir(directory):
            parent_dir = os.path.dirname(directory)
            if parent_dir:
                directory = parent_dir
        if not directory:
            return None
        if directory.endswith(("\\", "/")):
            return directory
        if "\\" in cleaned and "/" not in cleaned:
            return directory + "\\"
        return directory + os.sep

    def _apply_parent_image_folder(
        self, folder, *, display_value=None, remember_override=True
    ):
        if not folder:
            return False

        parent = getattr(self, "parent", None)
        if parent is None:
            return False

        success = True
        path_updated = False

        try:
            parent.path = folder
            path_updated = True
        except Exception as err:
            print(f"Failed to update image folder path: {err}")
            success = False

        path_ctrl = getattr(parent, "imageFolderPath", None)
        if path_ctrl is not None:
            try:
                path_ctrl.SetValue(display_value if display_value is not None else folder)
            except Exception as err:
                print(f"Failed to update image folder text: {err}")
                success = False

        updater = getattr(parent, "updateImageListBox", None)
        if callable(updater):
            try:
                updater()
            except Exception as err:
                print(f"Failed to refresh image list: {err}")
                success = False

        monitor = getattr(parent, "monitor", None)
        if monitor is not None:
            try:
                if getattr(parent, "autoRunning", False) and hasattr(monitor, "changeFilePath"):
                    monitor.changeFilePath(folder)
                else:
                    setattr(monitor, "filePath", folder)
            except Exception as err:
                print(f"Failed to update auto-fit folder: {err}")
                success = False

        if remember_override and (path_updated or success):
            self._parent_override_folder = folder

        return success

    def _restore_session_exclusions(self):
        """Reapply saved measurement and run exclusions after processing."""

        pending_runs = getattr(self, "_pending_session_run_exclusions", None) or {}
        if pending_runs and self.subrun_results:
            for run_idx, indices in list(pending_runs.items()):
                try:
                    run_index = int(run_idx)
                except Exception:
                    continue
                try:
                    target = {int(val) for val in indices}
                except Exception:
                    target = set(indices or [])
                if not target:
                    continue
                if not (0 <= run_index < len(self.subrun_results)):
                    continue
                try:
                    self.set_run_exclusions_bulk(run_index, target)
                except Exception:
                    continue
            self._pending_session_run_exclusions = {}
            self._session_run_exclusions = {
                idx: set(values)
                for idx, values in self.run_point_exclusions.items()
            }

        pending_points = getattr(self, "_pending_session_point_exclusions", None)
        if pending_points:
            try:
                pending_points = {int(val) for val in pending_points}
            except Exception:
                pending_points = set(pending_points)
            if pending_points:
                applied = set()
                for key, info in list(self.point_label_info.items()):
                    indices = info.get("orig_indices")
                    if not indices:
                        continue
                    try:
                        index_set = {int(val) for val in np.atleast_1d(indices)}
                    except Exception:
                        index_set = set(np.atleast_1d(indices))
                    if not index_set:
                        continue
                    if not pending_points.intersection(index_set):
                        continue
                    chk = self.exclude_checks.get(key)
                    if chk is None:
                        continue
                    if not chk.GetValue():
                        self._set_checkbox_value(chk, True)
                        try:
                            self.on_exclude_toggle(key, indices)
                        except Exception:
                            continue
                    applied.update(index_set)
                if applied:
                    pending_points.difference_update(applied)
                if pending_points:
                    self._pending_session_point_exclusions = set(pending_points)
                else:
                    self._pending_session_point_exclusions = set()
            else:
                self._pending_session_point_exclusions = set()

        self._session_point_exclusions = set()
        for pdata in self.plot_data.values():
            try:
                self._session_point_exclusions.update(
                    int(idx) for idx in pdata.get("excluded", set())
                )
            except Exception:
                self._session_point_exclusions.update(pdata.get("excluded", set()))

    def _restore_parent_context(self):
        if self._parent_context_restored:
            return

        parent = getattr(self, "parent", None)
        if parent is None:
            self._parent_context_restored = True
            return

        restored = False
        if self._parent_settings_modified and isinstance(
            self._parent_original_settings, dict
        ):
            applier = getattr(parent, "apply_settings_snapshot", None)
            if callable(applier):
                try:
                    applier(copy.deepcopy(self._parent_original_settings))
                    restored = True
                except Exception as err:
                    print(f"Failed to restore image UI settings: {err}")
            if not restored:
                apply_fn = getattr(parent, "_apply_settings", None)
                if callable(apply_fn):
                    try:
                        parent._settings = copy.deepcopy(self._parent_original_settings)
                        apply_fn()
                        saver = getattr(parent, "_save_settings", None)
                        if callable(saver):
                            saver()
                        restored = True
                    except Exception as err:
                        print(f"Failed to reapply image UI settings directly: {err}")

        folder = self._parent_original_path
        display_value = self._parent_original_path_display
        if self._parent_override_folder is not None:
            if folder:
                self._apply_parent_image_folder(
                    folder,
                    display_value=display_value if display_value is not None else folder,
                    remember_override=False,
                )
            elif display_value:
                path_ctrl = getattr(parent, "imageFolderPath", None)
                if path_ctrl is not None:
                    try:
                        path_ctrl.SetValue(display_value)
                    except Exception as err:
                        print(f"Failed to restore image folder display text: {err}")

        self._parent_override_folder = None
        self._parent_settings_modified = False
        self._parent_context_restored = True

    def load_parameter_file(self, path):
        """Load parameter values from a text file into the list control."""
        try:
            with open(path, "r") as f:
                content = f.read()
            self.paramListCtrl.SetValue(self._strip_sequence_header(content))
        except OSError:
            pass

    def load_parameter2_file(self, path):
        """Load second parameter values from a text file."""
        try:
            with open(path, "r") as f:
                content = f.read()
            self.param2ListCtrl.SetValue(self._strip_sequence_header(content))
        except OSError:
            pass

    def on_param_file_text(self, event):
        path = event.GetEventObject().GetValue()
        if path:
            self.load_parameter_file(path)

    def on_choose_param_file(self, event):
        dialog = wx.FileDialog(
            self,
            "Choose parameter list file",
            wildcard="Text files (*.txt)|*.txt|All files (*.*)|*.*",
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST,
        )
        if dialog.ShowModal() == wx.ID_OK:
            path = dialog.GetPath()
            self.paramFileCtrl.SetValue(path)
            self.load_parameter_file(path)
        dialog.Destroy()

    def on_param2_file_text(self, event):
        path = event.GetEventObject().GetValue()
        if path:
            self.load_parameter2_file(path)

    def on_choose_param2_file(self, event):
        dialog = wx.FileDialog(
            self,
            "Choose parameter 2 list file",
            wildcard="Text files (*.txt)|*.txt|All files (*.*)|*.*",
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST,
        )
        if dialog.ShowModal() == wx.ID_OK:
            path = dialog.GetPath()
            self.param2FileCtrl.SetValue(path)
            self.load_parameter2_file(path)
        dialog.Destroy()

    def on_choose_start_file(self, event):
        dialog = wx.FileDialog(
            self,
            "Select start file",
            defaultDir=getattr(self.parent, "path", ""),
            wildcard="FITS files (*.fits)|*.fits|All files (*.*)|*.*",
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST,
        )
        if dialog.ShowModal() == wx.ID_OK:
            path = dialog.GetPath()
            directory = os.path.dirname(path)
            if not directory.endswith("\\"):
                directory += "\\"
            self.parent.path = directory
            try:
                self.parent.imageFolderPath.SetValue(directory)
            except Exception:
                pass
            self.parent.updateImageListBox()
            self.startFileCtrl.SetValue(os.path.basename(path))
        dialog.Destroy()

    def parse_variable_list(self):
        """Return list of floats from the variable list control with offset applied."""
        text = self.varListCtrl.GetValue()
        raw = []
        for line in text.replace(',', ' ').split():
            try:
                raw.append(float(line))
            except ValueError:
                continue
        try:
            offset = float(self.varOffsetCtrl.GetValue())
        except ValueError:
            offset = 0.0
        factor = self.var_scale
        return [(v + offset) * factor for v in raw]

    def parse_parameter_list(self):
        """Return list of floats from the parameter list control with offset."""
        text = self.paramListCtrl.GetValue()
        raw = []
        for line in text.replace(',', ' ').split():
            try:
                raw.append(float(line))
            except ValueError:
                continue
        try:
            offset = float(self.paramOffsetCtrl.GetValue())
        except ValueError:
            offset = 0.0
        factor = self.param_scale
        return [(v + offset) * factor for v in raw]

    def parse_parameter2_list(self):
        """Return list of floats from the second parameter list control."""
        text = self.param2ListCtrl.GetValue()
        raw = []
        for line in text.replace(',', ' ').split():
            try:
                raw.append(float(line))
            except ValueError:
                continue
        try:
            offset = float(self.param2OffsetCtrl.GetValue())
        except ValueError:
            offset = 0.0
        factor = self.param2_scale
        return [(v + offset) * factor for v in raw]

    def _find_error_key(self, column):
        """Locate matching error column for a given data column."""
        candidates = []
        # simple suffixes
        for suf in (" Std", " Err", " std", " err", "_Std", "_Err", "_std", "_err"):
            candidates.append(f"{column}{suf}")
        # insert before unit parentheses if present
        if "(" in column:
            base, rest = column.split("(", 1)
            base = base.strip()
            for suf in (" Std", " Err", " std", " err", "_Std", "_Err", "_std", "_err"):
                candidates.append(f"{base}{suf} ({rest}")
        for key in candidates:
            if key in self.results:
                return key
        return None

    def _process_file_batch(self, files, progress_offset=0):
        """Run the fitting workflow for a batch of files."""
        res = {
            "Atom Number": [],
            "x-Atom Number": [],
            "y-Atom Number": [],
            "x-Center": [],
            "x-Center Std": [],
            "y-Center": [],
            "y-Center Std": [],
            "x-True Center": [],
            "x-True Center Std": [],
            "y-True Center": [],
            "y-True Center Std": [],
            "x-True Width": [],
            "x-True Width Std": [],
            "y-True Width": [],
            "y-True Width Std": [],
            "image_file": [],
        }
        prev_tof = self.parent._tof_fit_running
        prev_layer = getattr(self.parent, "chosenLayerNumber", None)

        try:
            self.parent._tof_fit_running = True
            if prev_layer is not None:
                self.parent.chosenLayerNumber = 4
            for i, fname in enumerate(files):
                if self._stop_requested:
                    break
                self.parent.filename = fname
                self.parent.highlight_image_in_list(fname)
                try:
                    self.parent.fileIndex = self.parent.fileList.index(fname)
                except ValueError:
                    pass
                getattr(self.parent, "setFilenameText", lambda: None)()
                self.parent.fitImage(None)
                self.parent.updateFittingResults()
                atom_num = (
                    self.parent.rawAtomNumber * (self.parent.pixelToDistance ** 2) / self.parent.crossSection
                )
                res["Atom Number"].append(atom_num)
                res["x-Atom Number"].append(self.parent.atomNumFromGaussianX)
                res["y-Atom Number"].append(self.parent.atomNumFromGaussianY)
                res["x-Center"].append(self.parent.x_center)
                res["x-Center Std"].append(self.parent.x_center_std)
                res["y-Center"].append(self.parent.y_center)
                res["y-Center Std"].append(self.parent.y_center_std)
                res["x-True Center"].append(self.parent.true_x_center)
                res["x-True Center Std"].append(self.parent.true_x_center_std or 0.0)
                res["y-True Center"].append(self.parent.true_y_center)
                res["y-True Center Std"].append(self.parent.true_y_center_std or 0.0)
                res["x-True Width"].append(self.parent.true_x_width)
                res["x-True Width Std"].append(self.parent.true_x_width_std or 0.0)
                res["y-True Width"].append(self.parent.true_y_width)
                res["y-True Width Std"].append(self.parent.true_y_width_std or 0.0)
                res["image_file"].append(fname)
                self.progress.SetValue(progress_offset + i + 1)
                wx.YieldIfNeeded()
        finally:
            self.parent._tof_fit_running = prev_tof
            if prev_layer is not None:
                self.parent.chosenLayerNumber = prev_layer
        return res

    def process_files(self, files, progress_offset=0, restore_top=True):
        """Process a set of image files and return result dictionary."""

        watch_factory = getattr(self.parent, "suspend_auto_watch", None)
        proc_factory = getattr(self.parent, "suspend_processing_controls", None)

        with ExitStack() as stack:
            if callable(proc_factory):
                stack.enter_context(proc_factory())
            if callable(watch_factory):
                stack.enter_context(watch_factory())
            res = self._process_file_batch(files, progress_offset)

        if restore_top:
            highlight_top = getattr(self.parent, "highlight_top_image", None)
            if callable(highlight_top):
                highlight_top()

        return res

    def _process_single(self, sel1, sel2):
        """Handle standard single-run processing."""
        self.subrun_results = []
        self.current_multi_fit = None
        start_file = self.startFileCtrl.GetValue().strip()
        start_idx = len(self.parent.fileList) - 1
        if start_file:
            for i, f in enumerate(self.parent.fileList):
                if f == start_file or os.path.basename(f) == os.path.basename(start_file):
                    start_idx = i
                    break
        num = min(len(self.var_values), start_idx + 1)
        start_slice = max(0, start_idx - num + 1)
        files_to_process = list(reversed(self.parent.fileList[start_slice : start_idx + 1]))
        self.var_values = self.var_values[: len(files_to_process)]
        self.progress.SetRange(len(files_to_process))
        res = self.process_files(files_to_process)
        self.results = res
        self._mark_unsaved(True)
        cols = [k for k in res.keys() if k != "image_file" and not k.endswith("Std")]
        self.colChoice1.Set(cols)
        self.colChoice2.Set(cols)
        if cols:
            if sel1 in cols:
                self.colChoice1.SetStringSelection(sel1)
            else:
                self.colChoice1.SetSelection(0)
            if sel2 in cols:
                self.colChoice2.SetStringSelection(sel2)
            else:
                self.colChoice2.SetSelection(0)
            self.update_heatmap()
        self._update_func_choices_for_plot(1)
        self._update_func_choices_for_plot(2)
        self.populate_param_controls(1)
        self.populate_param_controls(2)
        self._last_column_selection[1] = self.colChoice1.GetStringSelection()
        self._last_column_selection[2] = self.colChoice2.GetStringSelection()
        self._last_func_selection[1] = _normalise_fit_name(
            self.funcChoice1.GetStringSelection()
        )
        self._last_func_selection[2] = _normalise_fit_name(
            self.funcChoice2.GetStringSelection()
        )

    def _build_heatmap_value_results(self, res):
        derived = {}
        if not isinstance(res, dict):
            return derived
        for key, values in res.items():
            if key == "image_file" or key.lower().endswith("std"):
                continue
            if not isinstance(values, (list, tuple, np.ndarray)):
                continue
            try:
                arr = np.asarray(values, dtype=float).ravel()
            except Exception:
                continue
            if arr.size == 0:
                continue
            value = arr[-1]
            try:
                derived[key] = float(value)
            except (TypeError, ValueError):
                continue
        return derived

    def _process_heatmap(self, sel1, sel2):
        """Process runs for two parameters and store 2-D results."""
        self.heatmap_index_grid = None
        self.heatmap_active_plots.clear()
        self.heatmap_active_columns.clear()
        p1_vals = list(reversed(self.parse_parameter_list()))
        p2_vals = list(reversed(self.parse_parameter2_list()))
        if not p1_vals or not p2_vals:
            return
        self.subrun_results = []
        start_file = self.startFileCtrl.GetValue().strip()
        start_idx = len(self.parent.fileList) - 1
        if start_file:
            for i, f in enumerate(self.parent.fileList):
                if f == start_file or os.path.basename(f) == os.path.basename(start_file):
                    start_idx = i
                    break
        run_vals = self.var_values[:]
        if not run_vals:
            run_vals = [0.0]
        run_len = len(run_vals)
        needed = run_len * len(p1_vals) * len(p2_vals)
        slice_start = max(0, start_idx - needed + 1)
        files = list(reversed(self.parent.fileList[slice_start : start_idx + 1]))
        max_runs = len(files) // run_len
        # limit param lists to available runs
        n1 = len(p1_vals)
        n2 = len(p2_vals)
        while n1 * n2 > max_runs and n2 > 0:
            n2 -= 1
        while n1 * n2 > max_runs and n1 > 0:
            n1 -= 1
        p1_vals = p1_vals[:n1]
        p2_vals = p2_vals[:n2]
        files = files[: run_len * n1 * n2]
        if not p1_vals or not p2_vals:
            return
        self.progress.SetRange(len(files))
        agg_results = {}
        offset = 0
        display_name = (
            self.multiFitChoice.GetStringSelection() or TEMP_PSD_DISPLAY_NAME
        )
        display_name = _normalise_fit_name(display_name)
        func_key = self._selected_multi_fit_key()
        self.current_multi_fit = func_key
        index_grid = np.empty((len(p2_vals), len(p1_vals)), dtype=object)
        for j in range(len(p2_vals)):
            for i in range(len(p1_vals)):
                index_grid[j, i] = tuple()
        for j, p2 in enumerate(p2_vals):
            for i, p1 in enumerate(p1_vals):
                idx = (j * n1 + i) * run_len
                run_files = files[idx : idx + run_len]
                res = self.process_files(run_files, progress_offset=offset)
                offset += len(run_files)
                run_data = {
                    "param": p1,
                    "param2": p2,
                    "var_values": list(run_vals),
                    "results": res,
                    "param_label": self.paramLabelCtrl.GetValue(),
                    "param_unit": self.param_unit_scale,
                    "param_scale": self.param_scale,
                    "param2_label": self.param2LabelCtrl.GetValue(),
                    "param2_unit": self.param2_unit_scale,
                    "param2_scale": self.param2_scale,
                    "var_label": self.varLabelCtrl.GetValue(),
                    "var_unit": self.unit_scale,
                    "var_scale": self.var_scale,
                }
                (
                    filtered_var,
                    filtered_results,
                    included_indices,
                    excluded_indices,
                    groups,
                ) = self._build_filtered_run_components(
                    run_data, set(), apply_binning=True
                )
                run_data["filtered_var_values"] = filtered_var
                run_data["filtered_results"] = filtered_results
                run_data["included_indices"] = included_indices
                run_data["excluded_indices"] = excluded_indices
                run_data["variable_groups"] = groups or {}
                analysis_var = filtered_var if filtered_var else list(run_vals)
                analysis_results = filtered_results if filtered_results else res
                self.subrun_results.append(run_data)
                run_idx = len(self.subrun_results) - 1
                if groups:
                    self.variable_bin_groups[run_idx] = groups
                else:
                    self.variable_bin_groups.pop(run_idx, None)
                if 0 <= j < index_grid.shape[0] and 0 <= i < index_grid.shape[1]:
                    existing = index_grid[j, i]
                    if not existing:
                        index_grid[j, i] = (run_idx,)
                    else:
                        index_grid[j, i] = tuple(existing) + (run_idx,)
                if func_key == HEATMAP_VALUE_FIT_KEY:
                    derived = self._build_heatmap_value_results(analysis_results)
                else:
                    fit_info = FIT_FUNCTIONS.get(func_key)
                    if fit_info is None:
                        derived = {}
                    else:
                        derived = fit_info["derived"](
                            None,
                            None,
                            analysis_var,
                            analysis_results,
                            self.parent,
                        )
                for k, v in derived.items():
                    if k not in agg_results:
                        agg_results[k] = np.full((n2, n1), np.nan)
                    agg_results[k][j, i] = v
                self.subrun_results[-1]["derived"] = dict(derived)
                if self._stop_requested:
                    break
            if self._stop_requested:
                break
        if p1_vals and p2_vals:
            def _compute_axis_order(values):
                if len(values) <= 1:
                    return np.arange(len(values), dtype=int)
                arr = np.asarray(values, dtype=float)
                if not np.all(np.isfinite(arr)):
                    return np.arange(len(values), dtype=int)
                return np.argsort(arr, kind="mergesort")

            x_order = _compute_axis_order(p1_vals)
            y_order = _compute_axis_order(p2_vals)

            if x_order.size and not np.array_equal(x_order, np.arange(len(p1_vals))):
                for key, arr in list(agg_results.items()):
                    agg_results[key] = np.take(arr, x_order, axis=1)
                index_grid = np.take(index_grid, x_order, axis=1)
                p1_vals = [p1_vals[idx] for idx in x_order]

            if y_order.size and not np.array_equal(y_order, np.arange(len(p2_vals))):
                for key, arr in list(agg_results.items()):
                    agg_results[key] = np.take(arr, y_order, axis=0)
                index_grid = np.take(index_grid, y_order, axis=0)
                p2_vals = [p2_vals[idx] for idx in y_order]

        p1_vals, agg_results, index_grid = self._collapse_heatmap_axis(
            p1_vals, agg_results, index_grid, axis=1
        )
        p2_vals, agg_results, index_grid = self._collapse_heatmap_axis(
            p2_vals, agg_results, index_grid, axis=0
        )

        self.param_values = p1_vals
        self.param2_values = p2_vals
        self.heatmap_index_grid = index_grid
        self.results = agg_results
        self._mark_unsaved(True)
        cols = [k for k in agg_results.keys() if "err" not in k.lower() and "std" not in k.lower()]
        self.colChoice1.Set(cols)
        self.colChoice2.Set(cols)
        if cols:
            if sel1 in cols:
                self.colChoice1.SetStringSelection(sel1)
            else:
                self.colChoice1.SetSelection(0)
            if sel2 in cols:
                self.colChoice2.SetStringSelection(sel2)
            else:
                self.colChoice2.SetSelection(0)
        self._update_func_choices_for_plot(1)
        self._update_func_choices_for_plot(2)
        self.populate_param_controls(1)
        self.populate_param_controls(2)
        self._last_column_selection[1] = self.colChoice1.GetStringSelection()
        self._last_column_selection[2] = self.colChoice2.GetStringSelection()
        self._last_func_selection[1] = _normalise_fit_name(
            self.funcChoice1.GetStringSelection()
        )
        self._last_func_selection[2] = _normalise_fit_name(
            self.funcChoice2.GetStringSelection()
        )
        self._update_heatmap_var_bin_control(not self._processing)

    def _collapse_heatmap_axis(self, values, agg_results, index_grid, axis):
        """Combine duplicate parameter values by averaging their heat map cells."""

        if not values:
            return list(values), agg_results, index_grid

        groups = []
        unique_vals = []
        seen = {}
        for idx, value in enumerate(values):
            if value in seen:
                groups[seen[value]].append(idx)
            else:
                seen[value] = len(groups)
                groups.append([idx])
                unique_vals.append(value)

        if all(len(g) == 1 for g in groups):
            return list(values), agg_results, index_grid

        collapsed = {}
        for key in agg_results.keys():
            data = np.asarray(agg_results[key], dtype=float)
            if axis == 1:
                parts = []
                for cols in groups:
                    slice_vals = data[:, cols]
                    if slice_vals.ndim == 1:
                        slice_vals = slice_vals[:, np.newaxis]
                    count = np.sum(~np.isnan(slice_vals), axis=1, keepdims=True)
                    total = np.nansum(slice_vals, axis=1, keepdims=True)
                    with np.errstate(invalid="ignore", divide="ignore"):
                        part = np.divide(
                            total,
                            count,
                            out=np.full_like(total, np.nan, dtype=float),
                            where=count > 0,
                        )
                    parts.append(part)
                collapsed[key] = np.concatenate(parts, axis=1)
            else:
                parts = []
                for rows_idx in groups:
                    slice_vals = data[rows_idx, :]
                    if slice_vals.ndim == 1:
                        slice_vals = slice_vals[np.newaxis, :]
                    count = np.sum(~np.isnan(slice_vals), axis=0, keepdims=True)
                    total = np.nansum(slice_vals, axis=0, keepdims=True)
                    with np.errstate(invalid="ignore", divide="ignore"):
                        part = np.divide(
                            total,
                            count,
                            out=np.full_like(total, np.nan, dtype=float),
                            where=count > 0,
                        )
                    parts.append(part)
                collapsed[key] = np.concatenate(parts, axis=0)

        if axis == 1:
            new_grid = np.empty((index_grid.shape[0], len(groups)), dtype=object)
            for row in range(index_grid.shape[0]):
                for new_idx, cols in enumerate(groups):
                    runs = []
                    for col in cols:
                        runs.extend(self._iter_heatmap_runs(index_grid[row, col]))
                    new_grid[row, new_idx] = tuple(dict.fromkeys(runs))
        else:
            new_grid = np.empty((len(groups), index_grid.shape[1]), dtype=object)
            for new_idx, rows_idx in enumerate(groups):
                for col in range(index_grid.shape[1]):
                    runs = []
                    for row in rows_idx:
                        runs.extend(self._iter_heatmap_runs(index_grid[row, col]))
                    new_grid[new_idx, col] = tuple(dict.fromkeys(runs))

        return list(unique_vals), collapsed, new_grid

    def _iter_heatmap_runs(self, cell_value):
        """Return a list of valid run indices stored in a heat map cell."""

        if cell_value is None:
            return []
        if isinstance(cell_value, np.ndarray):
            items = cell_value.tolist()
        elif isinstance(cell_value, (list, tuple, set)):
            items = list(cell_value)
        else:
            items = [cell_value]
        runs = []
        for item in items:
            try:
                idx = int(item)
            except (TypeError, ValueError):
                continue
            if idx >= 0:
                runs.append(idx)
        return runs

    def _process_multiple(self, sel1, sel2):
        """Process multiple runs for different parameter values."""
        param_values = list(reversed(self.parse_parameter_list()))
        if not param_values:
            return
        start_file = self.startFileCtrl.GetValue().strip()
        start_idx = len(self.parent.fileList) - 1
        if start_file:
            for i, f in enumerate(self.parent.fileList):
                if f == start_file or os.path.basename(f) == os.path.basename(start_file):
                    start_idx = i
                    break
        run_vals = self.var_values[:]
        if not run_vals:
            run_vals = [0.0]
        run_len = len(run_vals)
        total_needed = run_len * len(param_values)
        slice_start = max(0, start_idx - total_needed + 1)
        files = list(reversed(self.parent.fileList[slice_start : start_idx + 1]))
        max_runs = len(files) // run_len
        param_values = param_values[:max_runs]
        files = files[: run_len * max_runs]
        self.progress.SetRange(len(files))
        agg_results = {}
        self.subrun_results = []
        offset = 0
        display_name = (
            self.multiFitChoice.GetStringSelection() or TEMP_PSD_DISPLAY_NAME
        )
        display_name = _normalise_fit_name(display_name)
        func_key = self._selected_multi_fit_key()
        self.current_multi_fit = func_key
        watch_factory = getattr(self.parent, "suspend_auto_watch", None)
        proc_factory = getattr(self.parent, "suspend_processing_controls", None)
        with ExitStack() as stack:
            if callable(proc_factory):
                stack.enter_context(proc_factory())
            if callable(watch_factory):
                stack.enter_context(watch_factory())
            for i, pval in enumerate(param_values):
                run_files = files[i * run_len : (i + 1) * run_len]
                res = self._process_file_batch(run_files, progress_offset=offset)
                offset += len(run_files)
                if func_key == HEATMAP_VALUE_FIT_KEY:
                    derived = self._build_heatmap_value_results(res)
                else:
                    fit_info = FIT_FUNCTIONS.get(func_key)
                    if fit_info is None:
                        derived = {}
                    else:
                        derived = fit_info["derived"](None, None, run_vals, res, self.parent)
                for k, v in derived.items():
                    agg_results.setdefault(k, []).append(v)
                self.subrun_results.append(
                    {
                        "param": pval,
                        "var_values": list(run_vals),
                        "results": res,
                        "param_label": self.paramLabelCtrl.GetValue(),
                        "param_unit": self.param_unit_scale,
                        "param_scale": self.param_scale,
                        "var_label": self.varLabelCtrl.GetValue(),
                        "var_unit": self.unit_scale,
                        "var_scale": self.var_scale,
                    }
                )
                if self._stop_requested:
                    break
        highlight_top = getattr(self.parent, "highlight_top_image", None)
        if callable(highlight_top):
            highlight_top()
        self.var_values = param_values[: len(next(iter(agg_results.values()), []))]
        self.results = agg_results
        self._mark_unsaved(True)
        cols = [k for k in agg_results.keys() if "err" not in k.lower() and "std" not in k.lower()]
        self.colChoice1.Set(cols)
        self.colChoice2.Set(cols)
        if cols:
            if sel1 in cols:
                self.colChoice1.SetStringSelection(sel1)
            else:
                self.colChoice1.SetSelection(0)
            if sel2 in cols:
                self.colChoice2.SetStringSelection(sel2)
            else:
                self.colChoice2.SetSelection(0)
        self._update_func_choices_for_plot(1)
        self._update_func_choices_for_plot(2)
        self.populate_param_controls(1)
        self.populate_param_controls(2)
        self._last_column_selection[1] = self.colChoice1.GetStringSelection()
        self._last_column_selection[2] = self.colChoice2.GetStringSelection()
        self._last_func_selection[1] = _normalise_fit_name(
            self.funcChoice1.GetStringSelection()
        )
        self._last_func_selection[2] = _normalise_fit_name(
            self.funcChoice2.GetStringSelection()
        )

    def on_func_change(self, plot_num, event=None):
        if getattr(self, "_suspend_fit_visibility", False):
            return
        """Update parameter guess boxes when fit function changes."""
        choice = self.funcChoice1 if plot_num == 1 else self.funcChoice2
        raw = choice.GetStringSelection()
        selection = _normalise_fit_name(raw)
        self._last_func_selection[plot_num] = selection
        self.populate_param_controls(plot_num)
        if event is not None:
            event.Skip()

    def populate_param_controls(self, plot_num):
        panel = self.paramGuessPanel1 if plot_num == 1 else self.paramGuessPanel2
        sizer = self.paramSizer1 if plot_num == 1 else self.paramSizer2
        sizer.Clear(True)
        if plot_num == 1:
            self.paramCtrls1 = []
        else:
            self.paramCtrls2 = []
        if getattr(self, "_suspend_fit_visibility", False):
            panel.Layout()
            return
        func_choice = self.funcChoice1 if plot_num == 1 else self.funcChoice2
        selection_index = func_choice.GetSelection()
        if selection_index is None:
            selection_index = -1
        raw_func = func_choice.GetString(selection_index) if selection_index >= 0 else ""
        func_name = _normalise_fit_name(raw_func) if selection_index > 0 else ""
        ctrls = []
        if func_name and func_name in FIT_FUNCTIONS:
            fit_info = FIT_FUNCTIONS[func_name]
            guessed = list(fit_info.get("p0", []))
            # If data are available, use heuristic guesses based on current column
            if self.results is not None and self.var_values:
                choice = self.colChoice1 if plot_num == 1 else self.colChoice2
                column = choice.GetStringSelection()
                if column:
                    x = np.asarray(self.var_values)
                    y = np.asarray(self.results.get(column, []))
                    if self.paramBinCheck.GetValue():
                        x, y, _, _ = self._bin_data(x, y)
                    g = self._guess_initial_params(func_name, x, y)
                    if g and len(g) == len(guessed):
                        guessed = g
            names = fit_info.get("param_names", [f"p{i}" for i in range(len(guessed))])
            grid = wx.FlexGridSizer(rows=len(guessed), cols=2, hgap=5, vgap=2)
            label_width = max((panel.GetTextExtent(n)[0] for n in names), default=0) + 10
            for name, val in zip(names, guessed):
                lbl = wx.StaticText(panel, label=name, size=(label_width, -1))
                grid.Add(lbl, 0, wx.ALIGN_CENTER_VERTICAL)
                ctrl = wx.TextCtrl(panel, value=str(val), size=(60, -1), style=wx.TE_PROCESS_ENTER)
                ctrl.SetToolTip(f"Initial guess for {name} when fitting Plot {plot_num}")
                ctrl.Bind(wx.EVT_TEXT_ENTER, lambda evt, pn=plot_num: self.update_plot(pn))
                grid.Add(ctrl, 0, wx.ALIGN_CENTER_VERTICAL)
                ctrls.append(ctrl)
            sizer.Add(grid, 0, wx.TOP | wx.EXPAND, 2)
        should_show = self._should_show_param_guesses(plot_num)
        if self.heatmapRadio.GetValue():
            should_show = False
        if not getattr(self, "_suspend_fit_visibility", False):
            self._set_param_guess_visibility(plot_num, should_show)
        if plot_num == 1:
            self.paramCtrls1 = ctrls
        else:
            self.paramCtrls2 = ctrls
        panel.Layout()
        self.control_panel.Layout()

    def _should_show_param_guesses(self, plot_num):
        if getattr(self, "_processing", False) or getattr(
            self, "_suspend_fit_visibility", False
        ):
            return False
        if self.heatmapRadio.GetValue():
            return False
        func_choice = self.funcChoice1 if plot_num == 1 else self.funcChoice2
        selection_index = func_choice.GetSelection()
        if selection_index is None or selection_index <= 0:
            return False
        func_name = _normalise_fit_name(func_choice.GetString(selection_index))
        panel = self.paramGuessPanel1 if plot_num == 1 else self.paramGuessPanel2
        sizer = panel.GetSizer()
        children = sizer.GetChildren() if sizer else []
        child_count = len(children)
        return bool(func_name and func_name in FIT_FUNCTIONS and child_count)

    def _set_param_guess_visibility(self, plot_num, should_show):
        if getattr(self, "_suspend_fit_visibility", False) or getattr(
            self, "_processing", False
        ):
            return
        panel = self.paramGuessPanel1 if plot_num == 1 else self.paramGuessPanel2
        label = self.paramGuessLabel1 if plot_num == 1 else self.paramGuessLabel2
        if should_show:
            panel.Show()
        else:
            panel.Hide()

        if label is not None:
            label.Hide()
            label_sizer = label.GetContainingSizer()
            if label_sizer:
                label_sizer.Show(label, False)
        panel_sizer = panel.GetContainingSizer()
        if panel_sizer:
            panel_sizer.Show(panel, should_show)

        try:
            self.control_panel.Layout()
        except Exception:
            pass

    def _force_hide_param_guesses(self):
        try:
            for label, panel in (
                (self.paramGuessLabel1, self.paramGuessPanel1),
                (self.paramGuessLabel2, self.paramGuessPanel2),
            ):
                if label is not None:
                    label.Hide()
                    label_sizer = label.GetContainingSizer()
                    if label_sizer:
                        label_sizer.Show(label, False)
                panel.Hide()
                panel_sizer = panel.GetContainingSizer()
                if panel_sizer:
                    panel_sizer.Show(panel, False)
            self.control_panel.Layout()
        except Exception:
            pass

    def _refresh_param_guess_visibility(self):
        if getattr(self, "_suspend_fit_visibility", False) or getattr(
            self, "_processing", False
        ):
            return
        if self.heatmapRadio.GetValue():
            self._set_param_guess_visibility(1, False)
            self._set_param_guess_visibility(2, False)
            return
        self._set_param_guess_visibility(1, self._should_show_param_guesses(1))
        self._set_param_guess_visibility(2, self._should_show_param_guesses(2))

    def on_process(self, event):
        """Parse variables and process image files."""
        if self._processing:
            self._stop_requested = True
            return
        self._suspend_fit_visibility = True
        self._force_hide_param_guesses()
        self.processBtn.SetLabel("Stop")
        self._stop_requested = False
        self._processing = True
        self.results = None
        self._mark_unsaved(False)
        self._reset_caption_inputs(1, refresh_display=False)
        self._reset_caption_inputs(2, refresh_display=False)
        self._set_controls_enabled(False)
        sel1 = self.colChoice1.GetStringSelection()
        sel2 = self.colChoice2.GetStringSelection()
        self.var_values = list(reversed(self.parse_variable_list()))
        if not self._multi_fit_requires_variables():
            self.var_values = [0.0]
        self.plot_data.clear()
        self.heatmap_active_plots.clear()
        self.heatmap_active_columns.clear()
        self.param_values = []
        self.param2_values = []
        self.run_point_exclusions = {}
        if self._session_point_exclusions:
            self._pending_session_point_exclusions = set(self._session_point_exclusions)
        else:
            self._pending_session_point_exclusions = set()
        if self._session_run_exclusions:
            self._pending_session_run_exclusions = {
                int(idx): set(values)
                for idx, values in self._session_run_exclusions.items()
            }
        else:
            self._pending_session_run_exclusions = {}
        self.variable_bin_groups = {}
        self.overlayCheck.SetValue(False)
        self.on_overlay_toggle(None)
        self.closeBtn1.Disable()
        self.closeBtn2.Disable()
        self._ensure_axes(1)
        self.axes[0].clear()
        self.canvas.draw()
        self.progress.SetForegroundColour(self._gauge_default_colour)
        self.progress.SetValue(0)
        self._clear_excludes()
        if not self.var_values and self._multi_fit_requires_variables():
            self.processBtn.SetLabel("Process")
            self._processing = False
            return
        if self.heatmapRadio.GetValue():
            self._process_heatmap(sel1, sel2)
        elif self.multiRadio.GetValue():
            self._process_multiple(sel1, sel2)
        else:
            self._process_single(sel1, sel2)
        if self._stop_requested:
            self.progress.SetValue(0)
        else:
            self.progress.SetForegroundColour(wx.Colour(0, 200, 0))
        self.processBtn.SetLabel("Process")
        self._processing = False
        self._stop_requested = False
        has_results = self._has_processed_results()
        self._set_controls_enabled(has_results)
        if not has_results:
            self._refresh_exclude_panel_visibility()
        else:
            self._restore_session_exclusions()
        self._suspend_fit_visibility = False
        self._refresh_param_guess_visibility()

    def _compute_font_sizes(self):
        self._font_sizes = {
            "label": AXIS_LABEL_FONTSIZE,
            "title": AXIS_LABEL_FONTSIZE + 4,
            "caption": 10.0,
            "tick": AXIS_LABEL_FONTSIZE,
        }

    def _get_font_size(self, kind):
        if not self._font_sizes:
            self._compute_font_sizes()
        return self._font_sizes.get(kind, self._font_sizes.get("label", 12.0))

    def _reset_axes_box_aspect(self, axes):
        if not axes:
            return
        for axis in axes:
            if axis is None:
                continue
            try:
                axis.set_box_aspect(1)
            except Exception:
                try:
                    axis.set_aspect("auto")
                except Exception:
                    pass

    def _configure_caption_axis(self, ax):
        if ax is None:
            return
        ax.cla()
        ax.set_axis_off()
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)

    def _heatmap_mode_active(self):
        return self.heatmapRadio.GetValue()

    def _refresh_multi_fit_choices(self):
        if not hasattr(self, "multiFitChoice"):
            return

        previous = self.multiFitChoice.GetStringSelection()
        previous_key = _normalise_fit_name(previous)
        previous_key = MULTI_FIT_OPTIONS.get(previous_key, previous_key)
        heatmap_mode = self._heatmap_mode_active()
        available = [
            name
            for name in MULTI_FIT_DISPLAY_ORDER
            if heatmap_mode
            or MULTI_FIT_OPTIONS.get(name, name) != HEATMAP_VALUE_FIT_KEY
        ]

        self.multiFitChoice.Set(available)

        if previous in available:
            self.multiFitChoice.SetStringSelection(previous)
        elif available:
            self.multiFitChoice.SetSelection(0)
        else:
            self.multiFitChoice.SetSelection(wx.NOT_FOUND)

        if (
            previous
            and previous_key == HEATMAP_VALUE_FIT_KEY
            and not heatmap_mode
            and getattr(self, "current_multi_fit", None) == HEATMAP_VALUE_FIT_KEY
        ):
            self.current_multi_fit = None

    def _selected_multi_fit_key(self):
        display_name = self.multiFitChoice.GetStringSelection() or TEMP_PSD_DISPLAY_NAME
        display_name = _normalise_fit_name(display_name)
        return MULTI_FIT_OPTIONS.get(display_name, display_name)

    def _multi_fit_requires_variables(self):
        return self._selected_multi_fit_key() != HEATMAP_VALUE_FIT_KEY

    def _update_multi_fit_controls(self):
        disable_vars = self._heatmap_mode_active() and not self._multi_fit_requires_variables()
        processing = getattr(self, "_processing", False)
        allow_interaction = not disable_vars and not processing
        for ctrl in getattr(self, "_variable_controls", []):
            try:
                ctrl.Enable(allow_interaction)
            except Exception:
                continue
        if hasattr(self, "varListCtrl"):
            if disable_vars:
                self.varListCtrl.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_BTNFACE))
            else:
                self.varListCtrl.SetBackgroundColour(wx.NullColour)
            self.varListCtrl.Refresh()

    def _configure_plot_axis(self, ax, *, force_square=True):
        if ax is None:
            return
        if force_square:
            try:
                ax.set_box_aspect(1)
            except Exception:
                pass
        else:
            try:
                ax.set_box_aspect(None)
            except Exception:
                pass
        try:
            ax.set_aspect("auto")
        except Exception:
            pass
        try:
            ax.tick_params(labelsize=self._get_font_size("tick"))
        except Exception:
            pass

    def _ensure_axes(self, count, layout="single"):
        fig = self.canvas.figure
        checkbox = getattr(self, "residualCheck", None)
        show_residuals = bool(checkbox and checkbox.GetValue())
        if layout == "heatmap":
            if checkbox and checkbox.GetValue():
                checkbox.SetValue(False)
            show_residuals = False
        layout_key = (count, layout, show_residuals)
        if self._current_axes_layout == layout_key:
            return

        self._current_axes_layout = layout_key
        self._clear_overlay_right_axis()
        fig.clf()

        self.axes = []
        self.caption_axes = []
        self.overlay_table_axes = []
        self.overlay_table_axis = None
        self.residual_axes = []
        self._axis_plot_map = {}
        self._hover_annotations = {}
        self._hover_last = (None, None, None)

        if layout == "double":
            if show_residuals:
                gs = fig.add_gridspec(
                    3,
                    2,
                    height_ratios=PLOT_RESIDUAL_CAPTION_HEIGHT_RATIOS,
                    width_ratios=[1, 1],
                )
                ax1 = fig.add_subplot(gs[0, 0])
                ax2 = fig.add_subplot(gs[0, 1])
                res1 = fig.add_subplot(gs[1, 0], sharex=ax1)
                res2 = fig.add_subplot(gs[1, 1], sharex=ax2)
                cap1 = fig.add_subplot(gs[2, 0])
                cap2 = fig.add_subplot(gs[2, 1])
                self.residual_axes = [res1, res2]
                fig.subplots_adjust(
                    left=0.1,
                    right=0.95,
                    top=0.93,
                    bottom=0.07,
                    wspace=0.28,
                    hspace=0.18,
                )
            else:
                gs = fig.add_gridspec(
                    2,
                    2,
                    height_ratios=PLOT_CAPTION_HEIGHT_RATIOS,
                    width_ratios=[1, 1],
                )
                ax1 = fig.add_subplot(gs[0, 0])
                ax2 = fig.add_subplot(gs[0, 1])
                cap1 = fig.add_subplot(gs[1, 0])
                cap2 = fig.add_subplot(gs[1, 1])
                fig.subplots_adjust(
                    left=0.1,
                    right=0.95,
                    top=0.93,
                    bottom=0.08,
                    wspace=0.28,
                    hspace=0.32,
                )
            for cap in (cap1, cap2):
                self._configure_caption_axis(cap)
            self.axes = [ax1, ax2]
            self.caption_axes = [cap1, cap2]
        elif layout == "overlay":
            if show_residuals:
                gs = fig.add_gridspec(
                    3, 1, height_ratios=PLOT_RESIDUAL_CAPTION_HEIGHT_RATIOS
                )
                plot_ax = fig.add_subplot(gs[0, 0])
                residual_ax = fig.add_subplot(gs[1, 0], sharex=plot_ax)
                caption_ax = fig.add_subplot(gs[2, 0])
                self.residual_axes = [residual_ax]
                fig.subplots_adjust(
                    left=0.12,
                    right=0.9,
                    top=0.93,
                    bottom=0.07,
                    hspace=0.18,
                )
            else:
                gs = fig.add_gridspec(2, 1, height_ratios=PLOT_CAPTION_HEIGHT_RATIOS)
                plot_ax = fig.add_subplot(gs[0, 0])
                caption_ax = fig.add_subplot(gs[1, 0])
                fig.subplots_adjust(
                    left=0.12,
                    right=0.9,
                    top=0.93,
                    bottom=0.08,
                    hspace=0.32,
                )
            self._configure_caption_axis(caption_ax)
            self.axes = [plot_ax]
            self.caption_axes = [caption_ax]
            self.overlay_table_axis = None
            self.overlay_table_axes = []
        else:
            if show_residuals and layout != "heatmap":
                gs = fig.add_gridspec(
                    3, 1, height_ratios=PLOT_RESIDUAL_CAPTION_HEIGHT_RATIOS
                )
                plot_ax = fig.add_subplot(gs[0, 0])
                residual_ax = fig.add_subplot(gs[1, 0], sharex=plot_ax)
                caption_ax = fig.add_subplot(gs[2, 0])
                self.residual_axes = [residual_ax]
                fig.subplots_adjust(
                    left=0.12,
                    right=0.92,
                    top=0.93,
                    bottom=0.07,
                    hspace=0.18,
                )
            else:
                gs = fig.add_gridspec(2, 1, height_ratios=PLOT_CAPTION_HEIGHT_RATIOS)
                plot_ax = fig.add_subplot(gs[0, 0])
                caption_ax = fig.add_subplot(gs[1, 0])
                if layout == "heatmap":
                    fig.subplots_adjust(
                        left=0.1,
                        right=0.9,
                        top=0.93,
                        bottom=0.08,
                        hspace=0.32,
                    )
                else:
                    fig.subplots_adjust(
                        left=0.12,
                        right=0.92,
                        top=0.93,
                        bottom=0.08,
                        hspace=0.32,
                    )
            self._configure_caption_axis(caption_ax)
            self.axes = [plot_ax]
            self.caption_axes = [caption_ax]

        if layout == "double" and not self.multiRadio.GetValue():
            self._apply_mode_size(self.multi_size)
        elif layout == "heatmap":
            self._apply_mode_size(self.heatmap_size)
        elif not self.multiRadio.GetValue():
            self._apply_mode_size(self.single_size)

        self._compute_font_sizes()
        for axis in self.axes:
            self._configure_plot_axis(axis, force_square=not show_residuals)
        for axis in self.residual_axes:
            self._configure_plot_axis(axis, force_square=False)

    def _clear_overlay_table_axes(self):
        if self.overlay_table_axis is not None:
            try:
                self.overlay_table_axis.cla()
                self.overlay_table_axis.set_axis_off()
            except Exception:
                pass
        self.overlay_table_axes = []

    def _clear_overlay_right_axis(self):
        ax = getattr(self, "overlay_right_axis", None)
        if not ax:
            self.overlay_right_axis = None
            return
        self._hide_hover(axis=ax)
        self._axis_plot_map.pop(ax, None)
        self._hover_annotations.pop(ax, None)
        try:
            ax.remove()
        except Exception:
            pass
        self.overlay_right_axis = None

    def _normalise_unit_key(self, unit, *, lower=True):
        if unit is None:
            return ""
        try:
            text = str(unit)
        except Exception:
            return ""
        text = text.strip()
        if not text:
            return ""
        for source in ("µ", "μ"):
            if source in text:
                text = text.replace(source, "u")
        if lower:
            text = text.lower()
        return text

    def _resolve_unit_family(self, unit, selected_family=None):
        raw_key = self._normalise_unit_key(unit, lower=False)
        lower_key = self._normalise_unit_key(unit)
        if raw_key and raw_key in self._unit_family_lookup:
            return self._unit_family_lookup[raw_key]
        if lower_key and lower_key in self._unit_family_lookup:
            if not (
                raw_key
                and len(raw_key) == 1
                and raw_key.isalpha()
                and raw_key not in self._unit_family_lookup
            ):
                return self._unit_family_lookup[lower_key]
        if selected_family:
            return selected_family
        fallback = raw_key or lower_key
        if fallback:
            return f"unit:{fallback}"
        return None

    def _normalise_axis_scale(self, scale_value):
        if scale_value is None:
            return SCALE_OPTIONS[0].lower()
        try:
            text = str(scale_value).strip().lower()
        except Exception:
            return SCALE_OPTIONS[0].lower()
        return text or SCALE_OPTIONS[0].lower()

    def _plots_share_axis_context(self):
        indices = [idx for idx in (1, 2) if idx in getattr(self, "plot_data", {})]
        if len(indices) < 2:
            return False
        contexts = []
        for idx in indices:
            data = self.plot_data.get(idx) or {}
            x_family = data.get("x_unit_family")
            if not x_family:
                x_family = self._resolve_unit_family(data.get("x_unit"))
            y_family = data.get("y_unit_family")
            if not y_family:
                y_family = self._resolve_unit_family(data.get("y_unit"))
            scale = data.get("axis_scale_mode")
            if not scale:
                scale = self._normalise_axis_scale(data.get("axis_scale"))
            contexts.append((x_family, y_family, scale))
        base = contexts[0]
        return all(ctx == base for ctx in contexts[1:])

    def _compute_overlay_limits(self, arrays, scale):
        """Return shared y-limits for overlay axes when units match."""

        if not arrays:
            return None

        finite_chunks = []
        positive_only = (scale or "linear").lower() == "log"
        min_positive = None
        for arr in arrays:
            if arr is None:
                continue
            try:
                data = np.asarray(arr, dtype=float)
            except Exception:
                continue
            if data.size == 0:
                continue
            if positive_only:
                data = data[data > 0]
            data = data[np.isfinite(data)]
            if data.size:
                if positive_only:
                    current_min = float(np.min(data))
                    if min_positive is None or current_min < min_positive:
                        min_positive = current_min
                finite_chunks.append(data)
        if not finite_chunks:
            return None
        combined = np.concatenate(finite_chunks)
        if combined.size == 0:
            return None
        y_min = float(np.min(combined))
        y_max = float(np.max(combined))
        if not np.isfinite(y_min) or not np.isfinite(y_max):
            return None
        scale_mode = (scale or "linear").lower()
        if y_min == y_max:
            if scale_mode == "log":
                y_min /= 1.2
                y_max *= 1.2
                if y_min <= 0:
                    positive = combined[combined > 0]
                    if positive.size:
                        y_min = float(np.min(positive)) / 1.2
                        y_max = float(np.max(positive)) * 1.2
                    else:
                        return None
            else:
                pad = abs(y_min) * 0.05 if y_min != 0 else 1.0
                y_min -= pad
                y_max += pad
                if y_min == y_max:
                    y_max = y_min + 1.0
        else:
            if scale_mode == "log":
                lower = y_min / 1.1
                upper = y_max * 1.1
                if lower <= 0 and min_positive is not None:
                    lower = min_positive / 1.1
                y_min, y_max = lower, upper
            else:
                span = y_max - y_min
                pad = span * 0.05
                reference = max(abs(y_min), abs(y_max), 1.0)
                if pad <= 0:
                    pad = 0.05 * reference
                y_min -= pad
                y_max += pad
        return y_min, y_max

    def _overlay_error_bounds(self, values, errors):
        """Return arrays capturing the extents implied by error bars."""

        if values is None or errors is None:
            return []
        try:
            vals = np.asarray(values, dtype=float)
        except Exception:
            return []
        if vals.size == 0:
            return []
        try:
            err = np.asarray(errors, dtype=float)
        except Exception:
            return []
        if err.size == 0:
            return []

        def _match_shape(component):
            comp = np.asarray(component, dtype=float)
            if comp.size == 0:
                return np.zeros_like(vals, dtype=float)
            if comp.shape == vals.shape:
                return comp.astype(float)
            try:
                return np.broadcast_to(comp, vals.shape).astype(float)
            except ValueError:
                flat = comp.reshape(-1)
                if flat.size == 1:
                    return np.full(vals.shape, float(flat[0]))
                if flat.size == vals.size:
                    return flat.reshape(vals.shape)
                magnitude = float(np.max(np.abs(flat))) if flat.size else 0.0
                return np.full(vals.shape, magnitude)

        if err.ndim >= 2:
            lower = _match_shape(err[0])
            upper = _match_shape(err[1] if err.shape[0] > 1 else err[0])
        else:
            matched = _match_shape(err)
            lower = matched
            upper = matched

        return [vals + upper, vals - lower]

    def _choose_text_position(self, xs, ys):
        """Select a corner for text that minimizes overlap with data.

        Defaults to the top-right corner but moves to the corner with the
        fewest data points if necessary.
        """
        if len(xs) == 0 or len(ys) == 0:
            return 0.95, 0.95, "right", "top"
        xmin, xmax = np.min(xs), np.max(xs)
        ymin, ymax = np.min(ys), np.max(ys)
        if xmax == xmin:
            xnorm = np.zeros_like(xs)
        else:
            xnorm = (xs - xmin) / (xmax - xmin)
        if ymax == ymin:
            ynorm = np.zeros_like(ys)
        else:
            ynorm = (ys - ymin) / (ymax - ymin)
        corners = {
            "top_left": ((0.05, 0.95, "left", "top"), np.sum((xnorm < 0.2) & (ynorm > 0.8))),
            "top_right": ((0.95, 0.95, "right", "top"), np.sum((xnorm > 0.8) & (ynorm > 0.8))),
            "bottom_left": ((0.05, 0.05, "left", "bottom"), np.sum((xnorm < 0.2) & (ynorm < 0.2))),
            "bottom_right": ((0.95, 0.05, "right", "bottom"), np.sum((xnorm > 0.8) & (ynorm < 0.2))),
        }
        # Prefer the top-right corner unless it contains data
        if corners["top_right"][1] == 0:
            return corners["top_right"][0]
        best = min(corners.values(), key=lambda c: c[1])
        return best[0]

    def _extract_measure(self, column):
        """Return the base measurement name for a result column."""
        col = column.lower()
        if "-" in col:
            col = col.split("-", 1)[1]
        col = col.strip()
        if col.startswith("true "):
            col = col[5:]
        return col

    def _refresh_plot_labels(self):
        if self._heatmap_mode_active():
            return
        if not self.plot_data:
            return
        if self.multiRadio.GetValue():
            label = self.paramLabelCtrl.GetValue()
            unit = self.param_unit_scale
            scale = self.param_scale
            family = self.paramUnitTypeChoice.GetStringSelection()
        else:
            label = self.varLabelCtrl.GetValue()
            unit = self.unit_scale
            scale = self.var_scale
            family = self.unitTypeChoice.GetStringSelection()
        for data in self.plot_data.values():
            data["x_label"] = label
            data["x_unit"] = unit
            data["x_unit_display"] = unit
            data["x_scale"] = scale
            data["x_unit_family"] = self._resolve_unit_family(unit, family)
        self.redraw_plots()
        self._mark_results_modified()


    def redraw_plots(self):
        self._axis_plot_map = {}
        self._artist_plot_map = {}
        if self._heatmap_mode_active():
            self._hide_hover()
            self._update_residual_control()
            return

        if not self.plot_data:
            self._hide_hover()
            self._clear_overlay_table_axes()
            self._clear_overlay_right_axis()
            self._ensure_axes(1, layout="single")
            if self.axes:
                ax = self.axes[0]
                ax.clear()
                ax.set_title("")
                ax.set_xlabel("")
                ax.set_ylabel("")
                self._configure_plot_axis(ax, force_square=not bool(self.residual_axes))
            if self.caption_axes:
                self._configure_caption_axis(self.caption_axes[0])
            self.canvas.draw()
            self._update_residual_control()
            return

        self._compute_font_sizes()
        force_square = not bool(self.residual_axes)
        for axis in self.axes:
            self._configure_plot_axis(axis, force_square=force_square)

        overlay = self.overlayCheck.GetValue()
        if overlay:
            self._draw_overlay_view()
        else:
            self._draw_standard_view()
        self._update_residual_control()

    def _draw_standard_view(self):
        indices = [idx for idx in (1, 2) if idx in self.plot_data]
        if len(indices) >= 2:
            self._ensure_axes(2, layout="double")
            residual_axes = self.residual_axes if self.residual_axes else [None] * len(
                self.axes
            )
            for ax, residual_ax, caption_ax, idx in zip(
                self.axes, residual_axes, self.caption_axes, indices
            ):
                self._draw_single_plot(ax, caption_ax, idx, residual_ax=residual_ax)
        else:
            idx = indices[0] if indices else None
            self._ensure_axes(1, layout="single")
            residual_ax = self.residual_axes[0] if self.residual_axes else None
            self._draw_single_plot(
                self.axes[0], self.caption_axes[0], idx, residual_ax=residual_ax
            )
        self.canvas.draw()

    def _draw_single_plot(self, ax, caption_ax, plot_num, residual_ax=None):
        self._configure_caption_axis(caption_ax)
        self._hide_hover(axis=ax)
        if plot_num is None or plot_num not in self.plot_data:
            ax.clear()
            ax.set_title("")
            ax.set_xlabel("")
            ax.set_ylabel("")
            self._configure_plot_axis(ax, force_square=residual_ax is None)
            if residual_ax is not None:
                self._render_residual_axis(residual_ax, [], None)
            self._axis_plot_map.pop(ax, None)
            self._hide_hover(axis=ax)
            return

        data = self.plot_data[plot_num]
        ax.clear()
        self._plot_dataset(ax, data, plot_num=plot_num, set_title=True)
        combined_x = self._combine_plot_arrays(
            data.get("plot_x"), data.get("plot_x_fit")
        )
        combined_y = self._combine_plot_arrays(
            data.get("plot_y"), data.get("plot_y_fit")
        )
        self._apply_axis_scale(
            ax,
            data.get("axis_scale", SCALE_OPTIONS[0]),
            combined_x,
            combined_y,
        )
        ax.relim()
        ax.autoscale_view()
        ax.margins(y=0.05)
        self._configure_plot_axis(ax, force_square=residual_ax is None)
        x_axis_label = ax.get_xlabel()
        if residual_ax is not None:
            plot_color = None
            artist = data.get("artist")
            if artist is not None:
                try:
                    plot_color = artist.get_color()
                except Exception:
                    plot_color = None
            series = self._build_residual_series(data, color=plot_color)
            if x_axis_label:
                ax.set_xlabel("")
            ax.tick_params(labelbottom=False)
            self._render_residual_axis(
                residual_ax, [series] if series else [], x_axis_label
            )
        else:
            ax.tick_params(labelbottom=True)
        self._axis_plot_map[ax] = plot_num
        caption_lines = self._compose_single_caption(plot_num)
        self._apply_caption(ax, caption_ax, caption_lines)

    def _draw_overlay_view(self):
        indices = [idx for idx in (1, 2) if idx in self.plot_data]
        if len(indices) < 2:
            self.overlayCheck.SetValue(False)
            self._draw_standard_view()
            return

        self._ensure_axes(1, layout="overlay")
        ax_left = self.axes[0]
        caption_ax = self.caption_axes[0]
        self._configure_caption_axis(caption_ax)
        self._hide_hover(axis=ax_left)
        ax_left.clear()
        self._clear_overlay_right_axis()
        self.overlay_right_axis = None
        residual_ax = self.residual_axes[0] if self.residual_axes else None
        self._configure_plot_axis(ax_left, force_square=residual_ax is None)
        use_dual_colors = bool(getattr(self, "overlayColorCheck", None) and self.overlayColorCheck.GetValue())
        if use_dual_colors:
            colors_by_idx = {1: "tab:blue", 2: "tab:red"}
        else:
            colors_by_idx = {1: "black", 2: "black"}
        plot_x_vals = []
        y_chunks = []
        axis_sides = {}
        residual_series = []
        legend_handles = []

        for idx in indices:
            axis = ax_left
            color = colors_by_idx.get(idx, "black")
            data = self.plot_data[idx]
            label = data.get("column", f"Plot {idx}")
            axis_sides[idx] = "left"
            marker_style = "o" if idx == 1 else "s"
            legend_label = label or f"Plot {idx}"
            self._plot_dataset(
                axis,
                data,
                plot_num=idx,
                set_title=False,
                color=color,
                label=legend_label,
                marker=marker_style,
            )
            legend_handles.append(
                Line2D(
                    [],
                    [],
                    marker=marker_style,
                    linestyle="None",
                    markerfacecolor=color,
                    markeredgecolor=color,
                    color=color,
                    label=legend_label,
                )
            )
            px = data.get("plot_x")
            if px is not None and len(px):
                plot_x_vals.append(np.asarray(px))
            px_fit = data.get("plot_x_fit")
            if px_fit is not None and len(px_fit):
                plot_x_vals.append(np.asarray(px_fit))
            py = data.get("plot_y")
            if py is not None and len(py):
                py_arr = np.asarray(py)
                y_chunks.append(py_arr)
                for bounds in self._overlay_error_bounds(py_arr, data.get("plot_err")):
                    if bounds is not None and getattr(bounds, "size", 0):
                        arr = np.asarray(bounds)
                        y_chunks.append(arr)
            py_fit = data.get("plot_y_fit")
            if py_fit is not None and len(py_fit):
                arr = np.asarray(py_fit)
                y_chunks.append(arr)
            if residual_ax is not None:
                series = self._build_residual_series(data, color=color)
                if series:
                    residual_series.append(series)

        combined_x = self._combine_plot_arrays(*plot_x_vals)
        combined_y = self._combine_plot_arrays(*y_chunks)

        scale_choice = self.plot_data[indices[0]].get("axis_scale", SCALE_OPTIONS[0])
        self._apply_axis_scale(ax_left, scale_choice, combined_x, combined_y)

        shared_xscale = ax_left.get_xscale()

        ax_left.relim()

        x_limits = self._compute_overlay_limits(plot_x_vals, shared_xscale)
        y_scale_for_limits = "log" if ax_left.get_yscale() == "log" else "linear"
        combined_limits = self._compute_overlay_limits(y_chunks, y_scale_for_limits)

        if x_limits is None or combined_limits is None:
            ax_left.autoscale_view()
        if combined_limits is None:
            ax_left.margins(y=0.05)
        if x_limits is not None:
            ax_left.set_xlim(*x_limits)
        if combined_limits is not None:
            ax_left.set_ylim(*combined_limits)

        x_labels = {self.plot_data[idx].get("x_label") or "" for idx in indices}
        x_units = {
            self.plot_data[idx].get("x_unit_display")
            or self.plot_data[idx].get("x_unit")
            or ""
            for idx in indices
        }
        if len(x_labels) == 1 and len(x_units) == 1:
            xl = next(iter(x_labels))
            xu = next(iter(x_units))
            label_text = f"{xl} ({xu})" if xu else xl
            ax_left.set_xlabel(label_text, fontsize=self._get_font_size("label"))
        else:
            ax_left.set_xlabel("")

        y_labels = []
        y_units = []
        for idx in indices:
            data = self.plot_data[idx]
            base_label = data.get("y_label") or data.get("column") or f"Plot {idx}"
            if base_label and base_label not in y_labels:
                y_labels.append(base_label)
            unit = data.get("y_unit") or ""
            if isinstance(unit, str):
                unit = unit.strip()
            unit = unit or ""
            if unit and unit not in y_units:
                y_units.append(unit)

        label_text = " / ".join(y_labels)
        if y_units:
            unit_text = y_units[0] if len(y_units) == 1 else " / ".join(y_units)
        else:
            unit_text = ""
        if label_text and unit_text:
            y_label_text = f"{label_text} ({unit_text})"
        else:
            y_label_text = label_text or unit_text
        ax_left.set_ylabel(
            y_label_text or "",
            fontsize=self._get_font_size("label"),
        )

        self._axis_plot_map[ax_left] = tuple(indices)

        series_names = [self.plot_data[idx].get("column", f"Plot {idx}") for idx in indices]
        if series_names:
            title_text = " & ".join(series_names)
            start_file = self.startFileCtrl.GetValue().strip()
            subtitle = os.path.basename(start_file) if start_file else ""
            if subtitle:
                ax_left.set_title(
                    f"{title_text}\n{subtitle}",
                    fontsize=self._get_font_size("title"),
                )
            else:
                ax_left.set_title(
                    title_text, fontsize=self._get_font_size("title")
                )

        caption_lines = self._compose_overlay_caption(indices, axis_sides)
        self._apply_caption(ax_left, caption_ax, caption_lines)

        if legend_handles:
            legend = ax_left.legend(
                handles=legend_handles,
                loc="center left",
                bbox_to_anchor=(1.02, 0.5),
                frameon=False,
                borderaxespad=0.0,
                handletextpad=0.6,
            )
            if legend is not None and hasattr(legend, "_legend_box"):
                legend._legend_box.align = "left"

        if residual_ax is not None:
            x_label_text = ax_left.get_xlabel()
            if x_label_text:
                ax_left.set_xlabel("")
            ax_left.tick_params(labelbottom=False)
            self._render_residual_axis(residual_ax, residual_series, x_label_text)
        else:
            ax_left.tick_params(labelbottom=True)

        self._configure_plot_axis(ax_left, force_square=residual_ax is None)
        self.canvas.draw()

    def _guess_initial_params(self, func_name, x, y):
        """Heuristically estimate initial parameters for curve fitting."""
        if func_name not in FIT_FUNCTIONS:
            return []
        if len(x) == 0 or len(y) == 0:
            return list(FIT_FUNCTIONS[func_name].get("p0", []))
        x = np.asarray(x, dtype=float)
        y = np.asarray(y, dtype=float)
        order = np.argsort(x)
        x = x[order]
        y = y[order]
        if func_name == "Linear" or func_name in TEMP_PSD_FIT_KEYS:
            try:
                m, b = np.polyfit(x, y, 1)
            except Exception:
                m, b = FIT_FUNCTIONS[func_name]["p0"]
            return [m, b]
        if func_name == "Quadratic":
            try:
                a, b, c = np.polyfit(x, y, 2)
            except Exception:
                a, b, c = FIT_FUNCTIONS[func_name]["p0"]
            return [a, b, c]
        if func_name in ("Exponential", MOT_LIFETIME_DISPLAY_NAME):
            C = float(np.min(y))
            A = float(np.max(y) - C)
            if A == 0:
                A = 1.0
            if len(x) >= 2 and np.any(y > C):
                y1, y2 = y[0], y[-1]
                if y2 <= C:
                    y2 = C + 1e-9
                B = np.log((y1 - C) / (y2 - C)) / (x[-1] - x[0])
                if not np.isfinite(B) or B <= 0:
                    B = 1.0
            else:
                B = 1.0
            if func_name == MOT_LIFETIME_DISPLAY_NAME:
                tau = 1 / B if B != 0 else 1.0
                return [A, tau, C]
            return [A, B, C]
        if func_name == "Gaussian":
            C = float(np.min(y))
            A = float(np.max(y) - C)
            mu = float(x[np.argmax(y)])
            weights = y - C
            weights[weights < 0] = 0
            if np.sum(weights) > 0:
                sigma = np.sqrt(np.sum(weights * (x - mu) ** 2) / np.sum(weights))
            else:
                sigma = (np.max(x) - np.min(x)) / 6.0
            if sigma == 0:
                sigma = (np.max(x) - np.min(x)) / 6.0
            return [A, mu, sigma, C]
        if func_name == "Lorentzian":
            C = float(np.min(y))
            A = float(np.max(y) - C)
            x0 = float(x[np.argmax(y)])
            half = C + A / 2
            idx = np.argmax(y)
            left_side = y[: idx + 1][::-1]
            left_x = x[: idx + 1][::-1]
            right_side = y[idx:]
            right_x = x[idx:]
            try:
                left = np.interp(half, left_side, left_x)
                right = np.interp(half, right_side, right_x)
                gamma = abs(right - left)
            except Exception:
                span = float(np.max(x) - np.min(x))
                if not np.isfinite(span) or span <= 0:
                    span = 1.0
                gamma = span / 10.0
            if not np.isfinite(gamma) or gamma <= 0:
                gamma = 1.0
            return [A, x0, gamma, C]
        if func_name == "Inverse":
            default = list(FIT_FUNCTIONS[func_name].get("p0", []))
            if len(x) < 2:
                return default
            finite_y = y[np.isfinite(y)]
            if len(finite_y) == 0:
                return default
            tail_size = max(3, len(y) // 5)
            tail = y[-tail_size:]
            finite_tail = tail[np.isfinite(tail)]
            if len(finite_tail) == 0:
                B = float(np.median(finite_y))
            else:
                B = float(np.median(finite_tail))
            eps = np.finfo(float).eps
            valid_idx = np.where(np.isfinite(y) & (np.abs(y - B) > eps))[0]
            if len(valid_idx) >= 2:
                i0, i1 = valid_idx[0], valid_idx[-1]
                y0, y1 = float(y[i0]), float(y[i1])
                x0, x1 = float(x[i0]), float(x[i1])
                denom = (1.0 / (y1 - B)) - (1.0 / (y0 - B))
                if denom != 0 and np.isfinite(denom):
                    A = (x1 - x0) / denom
                    if np.isfinite(A) and A != 0:
                        h = x0 - A / (y0 - B)
                        if np.isfinite(h):
                            return [float(A), float(h), float(B)]
                try:
                    x_valid = x[valid_idx]
                    y_valid = y[valid_idx]
                    delta = y_valid - B
                    safe = np.isfinite(delta) & (np.abs(delta) > eps)
                    x_fit = x_valid[safe]
                    delta = delta[safe]
                    if len(x_fit) >= 2:
                        z = 1.0 / delta
                        if np.all(np.isfinite(z)):
                            m, b = np.polyfit(x_fit, z, 1)
                            if m != 0 and np.isfinite(m) and np.isfinite(b):
                                A = 1.0 / m
                                h = -b / m
                                if np.isfinite(A) and np.isfinite(h):
                                    return [float(A), float(h), float(B)]
                except Exception:
                    pass
            if len(default) == 3 and np.isfinite(B):
                default = [default[0], default[1], float(B)]
            return default
        if func_name in ("Damped H.O.", MOT_RINGDOWN_DISPLAY_NAME):
            C = float(np.mean(y))
            y_adj = y - C
            A = float((np.max(y_adj) - np.min(y_adj)) / 2)
            omega = 1.0
            if len(x) > 1:
                dt = np.mean(np.diff(x))
                freqs = np.fft.rfftfreq(len(y_adj), dt)
                fft_mag = np.abs(np.fft.rfft(y_adj))
                if len(fft_mag) > 1:
                    idx = np.argmax(fft_mag[1:]) + 1
                    omega = 2 * np.pi * freqs[idx]
            gamma = 0.1
            peaks = np.where((y_adj[1:-1] > y_adj[:-2]) & (y_adj[1:-1] > y_adj[2:]))[0] + 1
            if len(peaks) > 1:
                t0 = x[peaks[0]]
                t1 = x[peaks[-1]]
                amp0 = abs(y_adj[peaks[0]])
                amp1 = abs(y_adj[peaks[-1]])
                if amp0 > 0 and amp1 > 0 and t1 != t0:
                    gamma = np.log(amp0 / amp1) / (t1 - t0)
                    if gamma <= 0 or not np.isfinite(gamma):
                        gamma = 0.1
            phi = 0.0
            if func_name == MOT_RINGDOWN_DISPLAY_NAME:
                return [C, A, gamma, omega, phi]
            return [A, gamma, omega, phi, C]
        return list(FIT_FUNCTIONS[func_name].get("p0", []))

    def update_plot(self, plot_num):
        """Store selections for a plot and redraw."""
        if self.heatmapRadio.GetValue():
            if not self.results:
                return
            choice = self.colChoice1 if plot_num == 1 else self.colChoice2
            column = choice.GetStringSelection()
            if column and column in self.results:
                self.heatmap_active_columns[plot_num] = column
                self.heatmap_active_plots.add(plot_num)
                pdata = self.plot_data.setdefault(plot_num, {})
                pdata["column"] = column
                pdata.setdefault("excluded", set())
                pdata["heatmap"] = True
            else:
                self.heatmap_active_columns.pop(plot_num, None)
                self.heatmap_active_plots.discard(plot_num)
                self.plot_data.pop(plot_num, None)
            self._last_column_selection[plot_num] = column
            self.update_heatmap()
            self._mark_results_modified()
            return
        if not self.var_values or self.results is None:
            return
        choice = self.colChoice1 if plot_num == 1 else self.colChoice2
        func_choice = self.funcChoice1 if plot_num == 1 else self.funcChoice2
        column = choice.GetStringSelection()
        raw_func = func_choice.GetStringSelection()
        func_name = _normalise_fit_name(raw_func)
        existing = self.plot_data.get(plot_num)
        normalized_func = func_name or None
        selection_changed = False
        if existing and (
            existing.get("column") != column
            or existing.get("func_name") != normalized_func
        ):
            selection_changed = True
            self._clear_excludes(plot_num)
        if not column or column not in (self.results or {}):
            return
        self._last_column_selection[plot_num] = column
        self._last_func_selection[plot_num] = func_name
        x = np.asarray(self.var_values)
        y = np.asarray(self.results[column])
        y_err = None
        err_key = self._find_error_key(column)
        if err_key:
            try:
                y_err = np.asarray(self.results.get(err_key), dtype=float)
            except Exception:
                y_err = None

        groups = [np.array([i]) for i in range(len(x))]
        if self.paramBinCheck.GetValue():
            x, y, y_err, groups = self._bin_data(x, y, y_err)
        param_ctrls = self.paramCtrls1 if plot_num == 1 else self.paramCtrls2
        guessed = self._guess_initial_params(func_name, x, y) if func_name else []
        p0 = []
        for idx, ctrl in enumerate(param_ctrls):
            try:
                p0.append(float(ctrl.GetValue()))
            except ValueError:
                p0.append(guessed[idx] if idx < len(guessed) else 0.0)
        if not selection_changed:
            self._clear_excludes(plot_num)
        if self.multiRadio.GetValue():
            x_label = self.paramLabelCtrl.GetValue()
            x_unit = self.param_unit_scale
            x_scale = self.param_scale
            x_family = self.paramUnitTypeChoice.GetStringSelection()
        else:
            x_label = self.varLabelCtrl.GetValue()
            x_unit = self.unit_scale
            x_scale = self.var_scale
            x_family = self.unitTypeChoice.GetStringSelection()
        axis_scale_selection = (
            self.scaleChoice1.GetStringSelection()
            if plot_num == 1
            else self.scaleChoice2.GetStringSelection()
        ) or SCALE_OPTIONS[0]
        self.plot_data[plot_num] = {
            "x": x,
            "y": y,
            "y_err": y_err,
            "func_name": func_name or None,
            "column": column,
            "p0": tuple(p0),
            "excluded": set(),
            "orig_indices": groups,
            "caption": self._get_caption_ctrl(plot_num).GetValue(),
            "x_label": x_label,
            "x_unit": x_unit,
            "x_scale": x_scale,
            "x_unit_family": self._resolve_unit_family(x_unit, x_family),
            "axis_scale": axis_scale_selection,
            "axis_scale_mode": self._normalise_axis_scale(axis_scale_selection),
            "residual_x": None,
            "residual_y": None,
            "residual_axis_label": None,
            "residual_series_label": None,
        }
        self._update_func_choices_for_plot(plot_num)
        self.redraw_plots()
        if plot_num == 1:
            self.closeBtn1.Enable()
        elif plot_num == 2:
            self.closeBtn2.Enable()
        self._sync_overlay_control()
        self._mark_results_modified()

    def _clear_heatmap_colorbars(self):
        if not self.heatmap_cbar:
            return
        for cbar in self.heatmap_cbar:
            try:
                cbar.remove()
            except Exception:
                pass
        self.heatmap_cbar = []

    def _clear_heatmap_detail_panels(self):
        if not getattr(self, "heatmap_detail_keys", None):
            return
        removed = False
        for key in list(self.heatmap_detail_keys):
            panel = self.point_panels.pop(key, None)
            if panel is not None:
                try:
                    panel.Destroy()
                except Exception:
                    pass
                else:
                    removed = True
            self.exclude_checks.pop(key, None)
            self.detail_checks.pop(key, None)
            self.point_label_info.pop(key, None)
            window = self.detail_windows.pop(key, None)
            if window is not None:
                try:
                    window.Destroy()
                except Exception:
                    pass
        self.heatmap_detail_keys.clear()
        for pdata in self.plot_data.values():
            if isinstance(pdata, dict) and pdata.get("heatmap"):
                pdata.pop("heatmap_point_info", None)
        if removed:
            self.control_panel.Layout()
            self._refresh_exclude_panel_visibility()

    def update_heatmap(self, only_plot=None, preserve_details=False):
        if not preserve_details:
            self._clear_heatmap_detail_panels()
        self._axis_plot_map = {}
        self._hide_hover()
        self._update_residual_control()
        if not self.heatmapRadio.GetValue():
            self.heatmap_axes_info = {}
            self.heatmap_active_plots.clear()
            self.heatmap_active_columns.clear()
            return
        if not self.results:
            self.heatmap_axes_info = {}
            self.heatmap_active_plots.clear()
            self.heatmap_active_columns.clear()
            self._clear_heatmap_colorbars()
            if self.axes:
                for ax in self.axes:
                    ax.clear()
            if self.caption_axes:
                for cap_ax in self.caption_axes:
                    self._configure_caption_axis(cap_ax)
            self.closeBtn1.Disable()
            self.closeBtn2.Disable()
            self.canvas.draw()
            return

        if only_plot is not None:
            self.heatmap_active_plots.add(only_plot)

        valid_columns = {}
        still_active = set()
        for idx in (1, 2):
            if idx not in self.heatmap_active_plots:
                continue
            column = self.heatmap_active_columns.get(idx)
            if column and column in self.results:
                valid_columns[idx] = column
                still_active.add(idx)
            else:
                self.heatmap_active_columns.pop(idx, None)
        self.heatmap_active_plots = still_active
        for idx in list(self.heatmap_active_columns.keys()):
            if idx not in still_active:
                self.heatmap_active_columns.pop(idx, None)
        active_indices = list(valid_columns.keys())

        if not active_indices:
            self.heatmap_axes_info = {}
            self.heatmap_active_plots.clear()
            self.heatmap_active_columns.clear()
            self._clear_heatmap_colorbars()
            self._ensure_axes(1, layout="heatmap")
            if self.axes:
                ax = self.axes[0]
                ax.clear()
                ax.set_title("")
                ax.set_xlabel("")
                ax.set_ylabel("")
                self._configure_plot_axis(ax)
            if self.caption_axes:
                self._configure_caption_axis(self.caption_axes[0])
            self.canvas.draw()
            self.closeBtn1.Disable()
            self.closeBtn2.Disable()
            return

        self.heatmap_active_plots = set(active_indices)
        selections = [(idx, valid_columns[idx]) for idx in active_indices]
        self._clear_heatmap_colorbars()

        layout = "double" if len(selections) >= 2 else "heatmap"
        self._ensure_axes(len(selections), layout=layout)
        self.heatmap_axes_info = {}

        x_values = np.asarray(self.param_values, dtype=float)
        y_values = np.asarray(self.param2_values, dtype=float)

        def compute_edges(values, count):
            if count <= 0:
                return np.array([0.0, 1.0]), np.array([]), np.array([])
            arr = np.asarray(values, dtype=float)
            if arr.size != count:
                arr = np.linspace(0, count - 1, count, dtype=float)
            if arr.size == 1:
                center = float(arr[0])
                edges = np.array([center - 0.5, center + 0.5], dtype=float)
                return edges, np.array([center], dtype=float), arr
            diffs = np.diff(arr)
            edges = np.empty(arr.size + 1, dtype=float)
            edges[1:-1] = arr[:-1] + diffs / 2
            edges[0] = arr[0] - diffs[0] / 2
            edges[-1] = arr[-1] + diffs[-1] / 2
            centers = (edges[:-1] + edges[1:]) / 2
            return edges, centers, arr

        self.heatmap_cbar = []
        start_file = self.startFileCtrl.GetValue().strip()
        subtitle = os.path.basename(start_file) if start_file else None
        x_label = self.paramLabelCtrl.GetValue()
        y_label = self.param2LabelCtrl.GetValue()
        x_unit = self.param_unit_scale
        y_unit = self.param2_unit_scale
        excluded_runs = self._collect_heatmap_excluded_runs()

        for axis_idx, (plot_idx, column) in enumerate(selections):
            ax = self.axes[axis_idx]
            caption_ax = self.caption_axes[axis_idx]
            self._configure_caption_axis(caption_ax)
            ax.clear()
            data = np.asarray(self.results[column], dtype=float)
            if data.ndim > 2:
                data = np.squeeze(data)
            if data.ndim == 1:
                data = data.reshape(-1, 1)
            if data.ndim != 2:
                continue
            rows, cols = data.shape
            x_edges, x_centers, x_labels = compute_edges(x_values, cols)
            y_edges, y_centers, y_labels = compute_edges(y_values, rows)
            x_left, x_right = x_edges[0], x_edges[-1]
            y_bottom, y_top = y_edges[0], y_edges[-1]
            heatmap_kwargs = {
                "shading": "auto",
            }
            cmap_name = self._resolve_heatmap_colormap(column)
            if cmap_name:
                heatmap_kwargs["cmap"] = cm.get_cmap(cmap_name)
            im = ax.pcolormesh(
                x_edges,
                y_edges,
                data,
                **heatmap_kwargs,
            )
            x_min = np.nanmin(x_edges)
            x_max = np.nanmax(x_edges)
            y_min = np.nanmin(y_edges)
            y_max = np.nanmax(y_edges)
            if x_left <= x_right:
                ax.set_xlim(x_min, x_max)
            else:
                ax.set_xlim(x_max, x_min)
            if y_bottom <= y_top:
                ax.set_ylim(y_min, y_max)
            else:
                ax.set_ylim(y_max, y_min)
            if x_centers.size:
                finite_x = np.isfinite(x_centers) & np.isfinite(x_labels)
                if np.any(finite_x):
                    ax.set_xticks(x_centers[finite_x])
                    ax.set_xticklabels(
                        [
                            self._format_heatmap_value(val, None)
                            for val in np.asarray(x_labels)[finite_x]
                        ]
                    )
            if y_centers.size:
                finite_y = np.isfinite(y_centers) & np.isfinite(y_labels)
                if np.any(finite_y):
                    ax.set_yticks(y_centers[finite_y])
                    ax.set_yticklabels(
                        [
                            self._format_heatmap_value(val, None)
                            for val in np.asarray(y_labels)[finite_y]
                        ]
                    )
            ax.set_xlabel(
                f"{x_label} ({x_unit})" if x_unit else x_label,
                fontsize=self._get_font_size("label"),
            )
            ax.set_ylabel(
                f"{y_label} ({y_unit})" if y_unit else y_label,
                fontsize=self._get_font_size("label"),
            )
            title_parts = [column]
            if subtitle:
                title_parts.append(subtitle)
            ax.set_title("\n".join(title_parts), fontsize=self._get_font_size("title"))
            cbar = self.canvas.figure.colorbar(
                im, ax=ax, fraction=0.046, pad=0.04
            )
            self.heatmap_cbar.append(cbar)
            caption_ctrl = self.captionCtrl1 if plot_idx == 1 else self.captionCtrl2
            caption_lines = self._generate_user_caption_lines(caption_ctrl.GetValue())
            self._apply_caption(ax, caption_ax, caption_lines)
            self._configure_plot_axis(ax, force_square=False)
            self._draw_heatmap_cell_overlays(
                ax,
                column,
                x_edges,
                y_edges,
                data.shape,
                data=data,
                image=im,
                excluded_runs=excluded_runs,
                x_labels=x_labels,
                y_labels=y_labels,
            )
            self.heatmap_axes_info[ax] = {
                "plot_idx": plot_idx,
                "column": column,
                "x_edges": x_edges,
                "y_edges": y_edges,
                "x_centers": x_centers,
                "y_centers": y_centers,
                "x_values": x_labels,
                "y_values": y_labels,
                "shape": data.shape,
            }

        self.canvas.draw()
        self.closeBtn1.Enable(1 in self.heatmap_active_plots)
        self.closeBtn2.Enable(2 in self.heatmap_active_plots)
        self._update_residual_control()

    def _resolve_heatmap_colormap(self, column):
        if not column:
            return None
        lowered = column.lower()
        simplified = lowered.replace("_", " ").replace("-", " ")
        condensed = simplified.replace(" ", "")
        default_cmap = rcParams.get("image.cmap")

        if simplified in {"atom number", "x atom number", "y atom number"}:
            return default_cmap or None
        if "temperature" in simplified or simplified.startswith("temp") or " temp" in simplified:
            return "Blues_r"
        if "intercept" in simplified:
            return "Greys"
        if "psd" in condensed:
            return "inferno"
        if "density" in simplified:
            return "Greys"
        if "lifetime" in simplified:
            return "viridis"
        if "decay rate" in simplified or ("decay" in simplified and "rate" in simplified):
            return "viridis_r"
        if "n0" in condensed or "n_0" in condensed:
            return "Greys"
        if "offset" in simplified:
            return "Greys"
        if "atom number" in simplified:
            return "inferno"
        if "center" in simplified:
            return "seismic"
        if "true width" in simplified:
            return "Greys"
        return None

    def _resolve_heatmap_unit_and_factor(self, column):
        """Return the unit label and scale factor for a heat map result column."""

        if not column:
            return "", 1.0

        unit = ""
        factor = 1.0

        stored_unit, stored_factor = self.col_scales.get(column, ("", 1.0))
        if stored_unit:
            unit = stored_unit
        if stored_factor not in (None, 0):
            factor = stored_factor

        if not unit:
            for pdata in getattr(self, "plot_data", {}).values():
                if not isinstance(pdata, dict):
                    continue
                if pdata.get("column") != column:
                    continue
                candidate_unit = (pdata.get("y_unit") or "").strip()
                if candidate_unit:
                    unit = candidate_unit
                candidate_factor = pdata.get("y_display_factor")
                if isinstance(candidate_factor, (int, float)) and candidate_factor != 0:
                    factor = candidate_factor
                if unit:
                    break

        if not unit:
            match = re.search(r"\(([^)]+)\)\s*$", str(column))
            if match:
                unit = match.group(1).strip()

        if not unit:
            suffix_match = re.search(r"(?:_|\s)([A-Za-zµ][A-Za-zµ0-9/^-]{0,4})$", str(column))
            if suffix_match:
                candidate = suffix_match.group(1).strip()
                common_units = {
                    "uk",
                    "µk",
                    "nk",
                    "k",
                    "hz",
                    "khz",
                    "mhz",
                    "ghz",
                    "s",
                    "ms",
                    "us",
                    "ns",
                    "rad",
                    "pix",
                    "arb",
                    "arb.",
                    "m",
                    "cm",
                    "mm",
                    "µm",
                    "um",
                    "nm",
                    "v",
                    "mv",
                    "uv",
                }
                if candidate.lower() in common_units:
                    unit = candidate

        if unit:
            existing_unit, existing_factor = self.col_scales.get(column, (None, None))
            if not existing_unit or column not in self.col_scales:
                self.col_scales[column] = (unit, factor)
            elif existing_unit != unit or existing_factor != factor:
                self.col_scales[column] = (unit, factor if factor is not None else existing_factor)

        return unit, factor

    def _format_heatmap_value(
        self,
        value,
        unit,
        *,
        preferred_exponent=None,
        return_parts=False,
    ):
        if value is None:
            text = "N/A"
            exponent = None
        else:
            try:
                val = float(value)
            except (TypeError, ValueError):
                text = str(value)
                exponent = None
            else:
                if np.isnan(val):
                    text = "N/A"
                    exponent = None
                elif val == 0:
                    if preferred_exponent is not None:
                        text = f"{0.0:.3f}e{preferred_exponent:+03d}"
                        exponent = preferred_exponent
                    else:
                        text = "0"
                        exponent = None
                else:
                    abs_val = abs(val)
                    if preferred_exponent is not None and abs_val > 0:
                        exponent = preferred_exponent
                        mantissa = val / (10 ** exponent)
                        text = f"{mantissa:.3f}e{exponent:+03d}"
                    elif abs_val >= 1e4 or abs_val < 1e-3:
                        text = f"{val:.3e}"
                        try:
                            exponent = int(text.split("e")[1])
                        except (IndexError, ValueError):
                            exponent = None
                    else:
                        text = f"{val:.6g}"
                        exponent = None
        if unit and text not in ("", "N/A"):
            text_with_unit = f"{text} {unit}"
        else:
            text_with_unit = text
        if return_parts:
            return text_with_unit, exponent, text
        return text_with_unit

    def _format_heatmap_overlay_number(self, value):
        """Format a numeric value for heat map overlays with one decimal place."""

        if value is None or not np.isfinite(value):
            return "N/A"

        if value == 0:
            return "0.0"

        abs_val = abs(value)
        if abs_val >= 1e4 or abs_val < 1e-2:
            exponent = int(np.floor(np.log10(abs_val)))
            mantissa = value / (10 ** exponent)
            mantissa = round(mantissa, 1)
            if mantissa == 0:
                return "0.0"
            if abs(mantissa) >= 10:
                mantissa /= 10
                exponent += 1
            return f"{mantissa:.1f}e{exponent:+03d}"

        rounded = round(value, 1)
        return f"{rounded:.1f}"

    def _get_heatmap_cell_display_info(self, column, row, col, excluded_runs=None):
        mean_val, spread_val, count = self._compute_heatmap_cell_stats(
            column, row, col, excluded_runs=excluded_runs
        )
        unit, factor = self._resolve_heatmap_unit_and_factor(column)
        if factor is None:
            factor = 1.0
        if np.isfinite(mean_val):
            display_mean = mean_val * factor
        else:
            display_mean = np.nan
        if count == 0:
            display_spread = np.nan
        elif np.isfinite(spread_val):
            display_spread = spread_val * factor
        else:
            display_spread = np.nan
        (
            value_text,
            value_exponent,
            value_base_text,
        ) = self._format_heatmap_value(display_mean, unit, return_parts=True)
        if count == 0:
            spread_text = "N/A"
            spread_base_text = "N/A"
        else:
            spread_text, _, spread_base_text = self._format_heatmap_value(
                display_spread,
                unit,
                preferred_exponent=value_exponent,
                return_parts=True,
            )
        return {
            "mean": mean_val,
            "spread": spread_val,
            "count": count,
            "display_mean": display_mean,
            "display_spread": display_spread,
            "value_text": value_text,
            "value_base_text": value_base_text,
            "spread_text": spread_text,
            "spread_base_text": spread_base_text,
            "unit": unit,
        }

    def _draw_heatmap_cell_overlays(
        self,
        ax,
        column,
        x_edges,
        y_edges,
        shape,
        data=None,
        image=None,
        excluded_runs=None,
        x_labels=None,
        y_labels=None,
    ):
        if ax is None:
            return
        if not shape or len(shape) != 2:
            return
        rows, cols = shape
        if rows <= 0 or cols <= 0:
            return
        try:
            x_edge_arr = np.asarray(x_edges, dtype=float)
            y_edge_arr = np.asarray(y_edges, dtype=float)
        except Exception:
            return
        if x_edge_arr.size < cols + 1 or y_edge_arr.size < rows + 1:
            return
        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        x_dir = 1
        y_dir = 1
        if len(xlim) >= 2 and xlim[1] < xlim[0]:
            x_dir = -1
        if len(ylim) >= 2 and ylim[1] < ylim[0]:
            y_dir = -1
        margin_frac = 0.04
        fontsize = self._get_font_size("tick")
        for row_idx in range(rows):
            for col_idx in range(cols):
                display_info = self._get_heatmap_cell_display_info(
                    column, row_idx, col_idx, excluded_runs=excluded_runs
                )
                value_text = display_info.get("value_text", "")
                spread_text = display_info.get("spread_text", "")
                value_base_text = display_info.get(
                    "value_base_text", value_text
                )
                spread_base_text = display_info.get(
                    "spread_base_text", spread_text
                )
                unit = display_info.get("unit", "")
                try:
                    x0 = float(x_edge_arr[col_idx])
                    x1 = float(x_edge_arr[col_idx + 1])
                    y0 = float(y_edge_arr[row_idx])
                    y1 = float(y_edge_arr[row_idx + 1])
                except (IndexError, TypeError, ValueError):
                    continue
                if not (
                    np.isfinite(x0)
                    and np.isfinite(x1)
                    and np.isfinite(y0)
                    and np.isfinite(y1)
                ):
                    continue
                if x_labels is not None and col_idx < len(x_labels):
                    x_value = x_labels[col_idx]
                else:
                    x_value = (x0 + x1) / 2
                if y_labels is not None and row_idx < len(y_labels):
                    y_value = y_labels[row_idx]
                else:
                    y_value = (y0 + y1) / 2
                _, _, x_text = self._format_heatmap_value(
                    x_value, self.param_unit_scale, return_parts=True
                )
                _, _, y_text = self._format_heatmap_value(
                    y_value, self.param2_unit_scale, return_parts=True
                )
                coord_line = f"({x_text}, {y_text})"
                display_mean = display_info.get("display_mean")
                value_line = self._format_heatmap_overlay_number(display_mean)

                lines = [coord_line, value_line]

                display_spread = display_info.get("display_spread")
                error_line = self._format_heatmap_overlay_number(display_spread)
                if error_line not in ("", "N/A"):
                    lines.append(f"+/- {error_line}")

                text = "\n".join(lines)
                width = abs(x1 - x0)
                height = abs(y1 - y0)
                if x_dir == 1:
                    x_left = x0 if x0 <= x1 else x1
                else:
                    x_left = x0 if x0 >= x1 else x1
                if y_dir == 1:
                    y_bottom = y0 if y0 <= y1 else y1
                else:
                    y_bottom = y0 if y0 >= y1 else y1
                x_pos = x_left + x_dir * width * margin_frac
                y_pos = y_bottom + y_dir * height * margin_frac
                text_color = self._resolve_heatmap_text_color(
                    row_idx,
                    col_idx,
                    data=data,
                    image=image,
                )
                ax.text(
                    x_pos,
                    y_pos,
                    text,
                    color=text_color,
                    ha="left",
                    va="bottom",
                    fontsize=fontsize,
                    linespacing=1.1,
                    zorder=3,
                    clip_on=True,
                )

    def _resolve_heatmap_text_color(self, row, col, data=None, image=None):
        if data is None or image is None:
            return "black"
        try:
            value = data[row, col]
        except Exception:
            return "black"
        if not np.isfinite(value):
            return "black"
        cmap = getattr(image, "cmap", None)
        norm = getattr(image, "norm", None)
        if cmap is None:
            return "black"
        try:
            if norm is not None:
                rgba = cmap(norm(value))
            else:
                rgba = cmap(value)
        except Exception:
            return "black"
        if not rgba:
            return "black"
        try:
            r, g, b, a = rgba
        except Exception:
            return "black"
        if a < 0.2:
            return "black"
        luminance = 0.2126 * r + 0.7152 * g + 0.0722 * b
        return "black" if luminance >= 0.5 else "white"

    def _build_heatmap_point_label(self, plot_idx, row, col):
        param_label = self.paramLabelCtrl.GetValue() or "Parameter 1"
        param2_label = self.param2LabelCtrl.GetValue() or "Parameter 2"
        param_value = None
        if 0 <= col < len(self.param_values):
            param_value = self.param_values[col]
        param2_value = None
        if 0 <= row < len(self.param2_values):
            param2_value = self.param2_values[row]
        param_text = self._format_heatmap_value(param_value, self.param_unit_scale)
        param2_text = self._format_heatmap_value(param2_value, self.param2_unit_scale)
        return (
            f"Plot {plot_idx}: {param_label}={param_text}, "
            f"{param2_label}={param2_text}"
        )

    def _create_heatmap_detail_panel(self, key, row, col, run_idx, sequence=None):
        plot_idx, _ = key
        label_text = self._build_heatmap_point_label(plot_idx, row, col)
        if sequence is not None:
            current, total = sequence
            if total > 1:
                label_text = f"{label_text} (Run {current} of {total})"
        file_names = self._heatmap_point_file_names(run_idx)
        detail_tooltip = "Open a window with detailed plots for this heat map point"
        panel, exclude_chk, show_chk = self._ensure_point_selection_panel(
            key,
            label_text,
            (run_idx,),
            detail_callback=lambda evt, k=key, inds=(run_idx,): self.on_show_detail_toggle(
                evt, k, inds
            ),
            detail_tooltip=detail_tooltip,
            file_names=file_names,
        )
        self.heatmap_detail_keys.add(key)
        return panel, label_text, show_chk, exclude_chk

    def on_close_plot2(self, event):
        if self.heatmapRadio.GetValue():
            self.colChoice2.SetSelection(wx.NOT_FOUND)
            self.captionCtrl2.ChangeValue("")
            self.heatmap_active_plots.discard(2)
            self.heatmap_active_columns.pop(2, None)
            self._last_column_selection[2] = self.colChoice2.GetStringSelection()
            self._clear_plot_fit_selection(2)
            self.update_heatmap()
            self.closeBtn2.Disable()
            self._mark_results_modified()
            return
        if 2 in self.plot_data:
            self.plot_data.pop(2, None)
            self.overlayCheck.SetValue(False)
            self.closeBtn2.Disable()
            self._clear_plot_fit_selection(2)
            self.on_overlay_toggle(None)
            self._sync_overlay_control()
            self._mark_results_modified()

    def on_close_plot1(self, event):
        if self.heatmapRadio.GetValue():
            self.colChoice1.SetSelection(wx.NOT_FOUND)
            self.captionCtrl1.ChangeValue("")
            self.heatmap_active_plots.discard(1)
            self.heatmap_active_columns.pop(1, None)
            self._last_column_selection[1] = self.colChoice1.GetStringSelection()
            self._clear_plot_fit_selection(1)
            self.update_heatmap()
            self.closeBtn1.Disable()
            self._mark_results_modified()
            return
        if 1 in self.plot_data:
            self.plot_data.pop(1, None)
            self.overlayCheck.SetValue(False)
            self.closeBtn1.Disable()
            self._clear_plot_fit_selection(1)
            self.on_overlay_toggle(None)
            self._sync_overlay_control()
            self._mark_results_modified()

    def _clear_plot_fit_selection(self, plot_num):
        func_choice = self.funcChoice1 if plot_num == 1 else self.funcChoice2
        if func_choice.GetCount() > 0:
            func_choice.SetSelection(0)
        self._last_func_selection[plot_num] = ""
        self.populate_param_controls(plot_num)

    def on_param_bin_toggle(self, event):
        for idx in list(self.plot_data.keys()):
            self.update_plot(idx)

    def on_var_bin_toggle(self, event):
        if not self.subrun_results:
            return
        for run_idx in range(len(self.subrun_results)):
            self._recompute_run_metrics(run_idx)
        self._refresh_plot_data_values()
        self.redraw_plots()
        for key, window in list(self.detail_windows.items()):
            if isinstance(window, PointDetailFrame):
                runs = self._collect_detail_runs(window.orig_indices)
                image_paths = self._point_file_paths(window.orig_indices, window.column)
                window.update_context(
                    runs,
                    window.point_info,
                    image_paths,
                    window.fit_key,
                    window.display_name,
                    window.column,
                )
        self._mark_results_modified()
        self._update_var_bin_control(not self._processing)

    def on_heatmap_var_bin_toggle(self, event):
        if not self._heatmap_mode_active() or not self.subrun_results:
            return
        if getattr(self, "_processing", False):
            return
        self.variable_bin_groups = {}
        for run_idx in range(len(self.subrun_results)):
            self._recompute_run_metrics(run_idx)
        self.update_heatmap(preserve_details=True)
        self._mark_results_modified()
        self._update_heatmap_var_bin_control(not self._processing)

    def on_overlay_toggle(self, event):
        overlay = self.overlayCheck.GetValue()
        if overlay and not self._plots_share_axis_context():
            self.overlayCheck.SetValue(False)
            overlay = False
        enabled = (not overlay) and (not self._processing)
        self.captionCtrl2.Enable(enabled)
        self.captionLbl2.Enable(enabled)
        self._sync_overlay_color_control()
        self.redraw_plots()
        self._mark_results_modified()
        if not overlay:
            self._sync_overlay_control()

    def on_overlay_color_toggle(self, event):
        if not self.overlayCheck.GetValue():
            return
        self.redraw_plots()

    def on_residual_toggle(self, event):
        if self._heatmap_mode_active():
            self.residualCheck.SetValue(False)
            return
        self.redraw_plots()

    def _bin_data(self, x, y, y_err=None):
        unique = np.unique(x)
        x_b, y_b, err_b, groups = [], [], [], []
        for val in unique:
            mask = x == val
            idxs = np.where(mask)[0]
            ys = y[mask]
            mean = np.mean(ys)
            if len(ys) > 1:
                if y_err is not None:
                    se = np.sqrt(np.sum((ys - mean) ** 2) / len(ys))
                else:
                    se = np.std(ys, ddof=1) / np.sqrt(len(ys))
            else:
                if y_err is not None:
                    se = float(y_err[mask][0])
                else:
                    se = 0.0
            x_b.append(val)
            y_b.append(mean)
            err_b.append(se)
            groups.append(idxs)
        return np.asarray(x_b), np.asarray(y_b), np.asarray(err_b), groups

    def _clear_excludes(self, plot_num=None):
        """Remove exclusion checkboxes and reset exclusion sets."""
        for key, panel in list(self.point_panels.items()):
            pn, _ = key
            if plot_num is None or pn == plot_num:
                if key in self.detail_windows:
                    window = self.detail_windows.pop(key)
                    if window:
                        window.Destroy()
                if panel:
                    panel.Destroy()
                self.point_panels.pop(key, None)
                self.exclude_checks.pop(key, None)
                self.detail_checks.pop(key, None)
                self.point_label_info.pop(key, None)
                self.heatmap_detail_keys.discard(key)
        if plot_num is None:
            for data in self.plot_data.values():
                data.get("excluded", set()).clear()
        elif plot_num in self.plot_data:
            self.plot_data[plot_num]["excluded"].clear()
        # Re-layout the control column after removing checkboxes
        self.control_panel.Layout()
        self._refresh_exclude_panel_visibility()

    def _ensure_exclude_panel_visible(self):
        if not getattr(self, "excludePanel", None):
            return
        if self.excludePanel.IsShown():
            return
        self.excludePanel.Show()
        self.control_panel.Layout()

    def _estimate_exclude_row_height(self, parent):
        sample_panel = wx.Panel(parent)
        sample_sizer = wx.BoxSizer(wx.HORIZONTAL)
        sample_chk = wx.CheckBox(sample_panel, label="Sample")
        sample_sizer.Add(sample_chk, 0, wx.ALL, 4)
        sample_panel.SetSizer(sample_sizer)
        sample_sizer.Fit(sample_panel)
        height = sample_panel.GetBestSize().GetHeight()
        sample_panel.Destroy()
        return max(height, 24)

    def _refresh_exclude_panel_visibility(self):
        if not getattr(self, "excludePanel", None):
            return
        has_items = bool(self.excludeSizer and self.excludeSizer.GetItemCount())
        if has_items and not self.excludePanel.IsShown():
            self.excludePanel.Show()
            self.control_panel.Layout()
        elif not has_items and self.excludePanel.IsShown():
            self.excludePanel.Hide()
            self.control_panel.Layout()
        scroller = getattr(self, "excludeScroll", None)
        if scroller:
            scroller.FitInside()

    def _ensure_point_selection_panel(
        self,
        key,
        label,
        orig_indices,
        *,
        detail_callback=None,
        detail_enabled=True,
        detail_label="Show fit details",
        detail_tooltip=None,
        detail_disabled_tooltip=None,
        file_names=None,
        tooltip=None,
    ):
        if isinstance(orig_indices, (list, tuple, set, np.ndarray)):
            indices_tuple = tuple(orig_indices)
        else:
            indices_tuple = (orig_indices,)
        panel_exists = key in self.point_panels
        parent_panel = None
        if panel_exists:
            point_panel = self.point_panels[key]
            panel_sizer = point_panel.GetSizer()
            if panel_sizer is None:
                panel_sizer = wx.BoxSizer(wx.VERTICAL)
                point_panel.SetSizer(panel_sizer)
            if panel_sizer.GetItemCount():
                row_sizer_item = panel_sizer.GetItem(0)
                row_sizer = row_sizer_item.GetSizer() if row_sizer_item else None
                if row_sizer is None:
                    row_sizer = wx.BoxSizer(wx.HORIZONTAL)
                    panel_sizer.Insert(0, row_sizer, 0)
            else:
                row_sizer = wx.BoxSizer(wx.HORIZONTAL)
                panel_sizer.Add(row_sizer, 0)
            parent_panel = point_panel.GetParent()
        else:
            parent_panel = getattr(self, "excludeScroll", None)
            if not isinstance(parent_panel, wx.Window) or not parent_panel:
                parent_panel = getattr(self, "excludePanel", None)
            if not isinstance(parent_panel, wx.Window) or not parent_panel:
                parent_panel = self.control_panel
            point_panel = wx.Panel(parent_panel)
            panel_sizer = wx.BoxSizer(wx.VERTICAL)
            row_sizer = wx.BoxSizer(wx.HORIZONTAL)
            panel_sizer.Add(row_sizer, 0)
            point_panel.SetSizer(panel_sizer)

        exclude_chk = self.exclude_checks.get(key)
        if exclude_chk is None:
            exclude_chk = wx.CheckBox(point_panel, label=label)
            row_sizer.Add(exclude_chk, 0, wx.RIGHT, 8)
            exclude_chk.Bind(
                wx.EVT_CHECKBOX,
                lambda evt, k=key, inds=indices_tuple: self.on_exclude_toggle(k, inds),
            )
        else:
            exclude_chk.SetLabel(label)
        tooltip_text = tooltip
        if not tooltip_text and file_names:
            tooltip_text = (
                "Exclude this data point from fitting and plots ("
                + ", ".join(file_names)
                + ")"
            )
        if not tooltip_text:
            tooltip_text = "Exclude this data point from fitting and plots"
        exclude_chk.SetToolTip(tooltip_text)

        pending_points = getattr(self, "_pending_session_point_exclusions", None)
        if pending_points:
            try:
                pending_values = {int(val) for val in pending_points}
            except Exception:
                pending_values = set(pending_points)
            try:
                index_values = {int(val) for val in np.atleast_1d(indices_tuple)}
            except Exception:
                index_values = set(np.atleast_1d(indices_tuple))
            if (
                pending_values
                and index_values
                and pending_values.intersection(index_values)
                and not exclude_chk.GetValue()
            ):
                self._set_checkbox_value(exclude_chk, True)
                try:
                    self.on_exclude_toggle(key, indices_tuple)
                except Exception:
                    pass

        show_chk = None
        if detail_callback is not None:
            show_chk = self.detail_checks.get(key)
            if show_chk is None:
                show_chk = wx.CheckBox(point_panel, label=detail_label)
                if detail_tooltip:
                    show_chk.SetToolTip(detail_tooltip)
                row_sizer.Add(show_chk, 0)
                if detail_enabled:
                    show_chk.Bind(wx.EVT_CHECKBOX, detail_callback)
            else:
                if detail_tooltip:
                    show_chk.SetToolTip(detail_tooltip)
            if not detail_enabled:
                show_chk.Disable()
                if detail_disabled_tooltip:
                    show_chk.SetToolTip(detail_disabled_tooltip)
            else:
                show_chk.Enable()
        else:
            show_chk = self.detail_checks.get(key)
            if show_chk is not None:
                show_chk.Disable()

        if panel_exists:
            file_label = getattr(point_panel, "_file_label", None)
            if file_label is not None:
                file_label.SetLabel(self._format_point_files(file_names))
            elif file_names:
                file_label = wx.StaticText(
                    point_panel, label=self._format_point_files(file_names)
                )
                panel_sizer.Add(file_label, 0, wx.TOP, 2)
                point_panel._file_label = file_label
        else:
            file_label = None
            if file_names:
                file_label = wx.StaticText(
                    point_panel, label=self._format_point_files(file_names)
                )
                panel_sizer.Add(file_label, 0, wx.TOP, 2)
            point_panel._file_label = file_label
            target_sizer = None
            exclude_scroller = getattr(self, "excludeScroll", None)
            if (
                isinstance(exclude_scroller, wx.Window)
                and exclude_scroller
                and parent_panel is exclude_scroller
                and getattr(self, "excludeSizer", None)
            ):
                self._ensure_exclude_panel_visible()
                target_sizer = self.excludeSizer
                border = 2
            else:
                exclude_panel = getattr(self, "excludePanel", None)
                if (
                    isinstance(exclude_panel, wx.Window)
                    and exclude_panel
                    and parent_panel is exclude_panel
                    and getattr(self, "excludeSizer", None)
                ):
                    self._ensure_exclude_panel_visible()
                    target_sizer = self.excludeSizer
                    border = 2
                else:
                    target_sizer = parent_panel.GetSizer()
                    if target_sizer is None:
                        target_sizer = wx.BoxSizer(wx.VERTICAL)
                        parent_panel.SetSizer(target_sizer)
                    border = 0
            target_sizer.Add(point_panel, 0, wx.TOP | wx.EXPAND, border)
            self.point_panels[key] = point_panel

        exclude_scroller = getattr(self, "excludeScroll", None)
        if isinstance(exclude_scroller, wx.Window) and exclude_scroller:
            if point_panel.GetParent() is exclude_scroller:
                exclude_scroller.Layout()
                exclude_scroller.FitInside()

        self.exclude_checks[key] = exclude_chk
        if show_chk is not None:
            self.detail_checks[key] = show_chk

        point_panel.Layout()
        self.control_panel.Layout()
        return point_panel, exclude_chk, show_chk

    def _get_hover_annotation(self, ax):
        annotation = self._hover_annotations.get(ax)
        if (
            annotation is None
            or annotation.axes is not ax
            or annotation not in getattr(ax, "texts", [])
        ):
            annotation = ax.annotate(
                "",
                xy=(0, 0),
                xytext=(10, 10),
                textcoords="offset points",
                bbox={"boxstyle": "round,pad=0.3", "fc": "white", "alpha": 0.85},
                arrowprops={"arrowstyle": "->", "color": "black"},
                fontsize=self._get_font_size("tick"),
            )
            annotation.set_visible(False)
            self._hover_annotations[ax] = annotation
        return annotation

    def _format_hover_text(self, data, x_value, y_value, error=None):
        del data  # Only positional values are required for hover feedback.
        lines = [f"x: {x_value:.3g}", f"y: {y_value:.3g}"]
        if error is None:
            lines.append("err: n/a")
        elif isinstance(error, tuple):
            lower, upper = error
            lower_text = "n/a"
            upper_text = "n/a"
            if lower is not None and np.isfinite(lower):
                lower_text = f"{abs(lower):.3g}"
            if upper is not None and np.isfinite(upper):
                upper_text = f"{abs(upper):.3g}"
            lines.append(f"err: -{lower_text}/+{upper_text}")
        else:
            if np.isfinite(error):
                lines.append(f"err: {abs(error):.3g}")
            else:
                lines.append("err: n/a")
        return "\n".join(lines)

    def _hide_hover(self, axis=None):
        changed = False
        if axis is None:
            for annotation in self._hover_annotations.values():
                if annotation is not None and annotation.get_visible():
                    annotation.set_visible(False)
                    changed = True
            self._hover_last = (None, None, None)
        else:
            annotation = self._hover_annotations.get(axis)
            if annotation is not None and annotation.get_visible():
                annotation.set_visible(False)
                changed = True
            if self._hover_last[0] is axis:
                self._hover_last = (None, None, None)
        if changed and hasattr(self, "canvas"):
            self.canvas.draw_idle()

    def _resolve_hover_candidate(self, ax, data, event):
        if not data:
            return None
        plot_x = data.get("plot_x")
        plot_y = data.get("plot_y")
        if plot_x is None or plot_y is None:
            return None
        try:
            plot_x = np.asarray(plot_x, dtype=float)
            plot_y = np.asarray(plot_y, dtype=float)
        except Exception:
            return None
        if plot_x.size == 0 or plot_y.size == 0:
            return None
        length = min(plot_x.size, plot_y.size)
        if length <= 0:
            return None
        plot_x = plot_x[:length]
        plot_y = plot_y[:length]

        plot_err = data.get("plot_err")
        err_values = None
        err_lower = None
        err_upper = None
        if plot_err is not None:
            try:
                plot_err = np.asarray(plot_err, dtype=float)
            except Exception:
                plot_err = None
        if plot_err is not None:
            if plot_err.ndim == 1:
                err_values = plot_err[:length]
            elif plot_err.ndim == 2 and plot_err.shape[0] == 2:
                err_lower = plot_err[0][:length]
                err_upper = plot_err[1][:length]

        valid_mask = np.isfinite(plot_x) & np.isfinite(plot_y)
        if err_values is not None:
            valid_mask &= np.isfinite(err_values)
        if err_lower is not None and err_upper is not None:
            valid_mask &= np.isfinite(err_lower) & np.isfinite(err_upper)
        if not np.any(valid_mask):
            return None
        plot_x = plot_x[valid_mask]
        plot_y = plot_y[valid_mask]
        if err_values is not None:
            err_values = err_values[valid_mask]
        if err_lower is not None and err_upper is not None:
            err_lower = err_lower[valid_mask]
            err_upper = err_upper[valid_mask]
        if plot_x.size == 0 or plot_y.size == 0:
            return None

        xdata = getattr(event, "xdata", None)
        ydata = getattr(event, "ydata", None)
        if xdata is None or ydata is None:
            return None
        try:
            mouse_disp = ax.transData.transform((xdata, ydata))
            point_coords = np.column_stack((plot_x, plot_y))
            points_disp = ax.transData.transform(point_coords)
        except Exception:
            return None
        if points_disp.size == 0:
            return None
        deltas = points_disp - mouse_disp
        dist_sq = np.sum(deltas**2, axis=1)
        if dist_sq.size == 0 or not np.any(np.isfinite(dist_sq)):
            return None
        finite_mask = np.isfinite(dist_sq)
        if not np.all(finite_mask):
            dist_sq = dist_sq[finite_mask]
            plot_x = plot_x[finite_mask]
            plot_y = plot_y[finite_mask]
            if err_values is not None:
                err_values = err_values[finite_mask]
            if err_lower is not None and err_upper is not None:
                err_lower = err_lower[finite_mask]
                err_upper = err_upper[finite_mask]
        if dist_sq.size == 0:
            return None
        nearest = int(np.argmin(dist_sq))
        if nearest < 0 or nearest >= plot_x.size:
            return None

        hover_err = None
        if err_values is not None and nearest < err_values.size:
            hover_err = float(err_values[nearest])
        elif (
            err_lower is not None
            and err_upper is not None
            and nearest < err_lower.size
            and nearest < err_upper.size
        ):
            hover_err = (float(err_lower[nearest]), float(err_upper[nearest]))

        return {
            "index": nearest,
            "dist_sq": float(dist_sq[nearest]),
            "x": float(plot_x[nearest]),
            "y": float(plot_y[nearest]),
            "err": hover_err,
        }

    def on_canvas_motion(self, event):
        if self._heatmap_mode_active():
            self._hide_hover()
            return
        if getattr(self.toolbar, "mode", None):
            self._hide_hover()
            return
        ax = getattr(event, "inaxes", None)
        if ax is None:
            self._hide_hover()
            return
        plot_spec = self._axis_plot_map.get(ax)
        if plot_spec is None:
            self._hide_hover(axis=ax)
            return
        if isinstance(plot_spec, (list, tuple, set)):
            candidate_indices = [idx for idx in plot_spec if idx in self.plot_data]
        else:
            candidate_indices = [plot_spec] if plot_spec in self.plot_data else []
        if not candidate_indices:
            self._hide_hover(axis=ax)
            return

        best_result = None
        best_index = None
        best_plot = None
        for plot_idx in candidate_indices:
            data = self.plot_data.get(plot_idx)
            candidate = self._resolve_hover_candidate(ax, data, event)
            if candidate is None:
                continue
            dist_sq = candidate.get("dist_sq")
            if dist_sq is None or not np.isfinite(dist_sq):
                continue
            if best_result is None or dist_sq < best_result.get("dist_sq"):
                best_result = candidate
                best_index = candidate.get("index")
                best_plot = plot_idx
        if best_result is None or best_index is None or best_plot is None:
            self._hide_hover(axis=ax)
            return
        if best_result["dist_sq"] > self._hover_snap_distance_sq:
            self._hide_hover(axis=ax)
            return

        prev_axis = self._hover_last[0]
        prev_plot = self._hover_last[1] if len(self._hover_last) > 1 else None
        prev_point = self._hover_last[2] if len(self._hover_last) > 2 else None
        if prev_axis is ax and prev_plot == best_plot and prev_point == best_index:
            return

        annotation = self._get_hover_annotation(ax)
        hover_x = best_result["x"]
        hover_y = best_result["y"]
        hover_err = best_result.get("err")
        data = self.plot_data.get(best_plot)
        if data is None:
            self._hide_hover(axis=ax)
            return
        annotation.xy = (hover_x, hover_y)
        annotation.set_text(self._format_hover_text(data, hover_x, hover_y, hover_err))
        if not annotation.get_visible():
            annotation.set_visible(True)
        self._hover_last = (ax, best_plot, best_index)
        if hasattr(self, "canvas"):
            self.canvas.draw_idle()

    def _resolve_heatmap_cell(self, info, xdata, ydata):
        grid = self.heatmap_index_grid
        if grid is None or not isinstance(grid, np.ndarray):
            return None
        x_edges = info.get("x_edges")
        y_edges = info.get("y_edges")
        if x_edges is None or y_edges is None:
            return None
        x_min = min(x_edges[0], x_edges[-1])
        x_max = max(x_edges[0], x_edges[-1])
        y_min = min(y_edges[0], y_edges[-1])
        y_max = max(y_edges[0], y_edges[-1])
        if not (x_min <= xdata <= x_max and y_min <= ydata <= y_max):
            return None
        try:
            x_centers = np.asarray(info.get("x_centers"), dtype=float)
            y_centers = np.asarray(info.get("y_centers"), dtype=float)
        except (TypeError, ValueError):
            return None
        if x_centers.size == 0 or y_centers.size == 0:
            return None
        finite_x = np.where(np.isfinite(x_centers))[0]
        finite_y = np.where(np.isfinite(y_centers))[0]
        if finite_x.size == 0 or finite_y.size == 0:
            return None
        col_idx = np.argmin(np.abs(x_centers[finite_x] - xdata))
        row_idx = np.argmin(np.abs(y_centers[finite_y] - ydata))
        col = int(finite_x[col_idx])
        row = int(finite_y[row_idx])
        shape = info.get("shape", ())
        if not shape or len(shape) != 2:
            return None
        rows, cols = shape
        if rows <= 0 or cols <= 0:
            return None
        row = int(np.clip(row, 0, rows - 1))
        col = int(np.clip(col, 0, cols - 1))
        if row >= grid.shape[0] or col >= grid.shape[1]:
            return None
        return row, col

    def on_canvas_click(self, event):
        if not self.heatmapRadio.GetValue():
            return
        if getattr(self.toolbar, "mode", None):
            return
        if getattr(event, "button", None) != 1:
            return
        ax = getattr(event, "inaxes", None)
        if ax is None:
            return
        info = self.heatmap_axes_info.get(ax)
        if not info:
            return
        xdata = getattr(event, "xdata", None)
        ydata = getattr(event, "ydata", None)
        if xdata is None or ydata is None:
            return
        indices = self._resolve_heatmap_cell(info, xdata, ydata)
        if indices is None:
            return
        row, col = indices
        grid = self.heatmap_index_grid
        if grid is None or not isinstance(grid, np.ndarray):
            return
        cell_value = grid[row, col]
        run_indices = [
            idx
            for idx in self._iter_heatmap_runs(cell_value)
            if 0 <= idx < len(self.subrun_results)
        ]
        if not run_indices:
            return
        plot_idx = info.get("plot_idx", 1)
        column = info.get("column")
        heatmap_data = self.plot_data.setdefault(plot_idx, {})
        heatmap_data.setdefault("excluded", set())
        heatmap_data["heatmap"] = True
        if column:
            heatmap_data["column"] = column
        point_info_map = heatmap_data.setdefault("heatmap_point_info", {})

        param_label = self.paramLabelCtrl.GetValue() or "Parameter 1"
        param2_label = self.param2LabelCtrl.GetValue() or "Parameter 2"
        param_value = None
        if 0 <= col < len(self.param_values):
            param_value = self.param_values[col]
        param2_value = None
        if 0 <= row < len(self.param2_values):
            param2_value = self.param2_values[row]
        param_text = self._format_heatmap_value(param_value, self.param_unit_scale)
        param2_text = self._format_heatmap_value(param2_value, self.param2_unit_scale)

        created = False
        auto_open = None
        total = len(run_indices)
        for order, run_idx in enumerate(run_indices, start=1):
            key = (plot_idx, run_idx)
            existing_panel = key in self.point_panels
            panel, label_text, show_chk, exclude_chk = self._create_heatmap_detail_panel(
                key, row, col, run_idx, sequence=(order, total)
            )
            if not panel:
                continue
            self.point_panels[key] = panel
            if exclude_chk is not None:
                self.exclude_checks[key] = exclude_chk
            if show_chk is not None:
                self.detail_checks[key] = show_chk
            info_entry = self.point_label_info.setdefault(key, {})
            info_entry["plot"] = plot_idx
            info_entry["orig_indices"] = (run_idx,)
            info_entry["custom_label"] = label_text
            info_entry["heatmap_coords"] = (row, col)
            self._apply_point_label(key)
            point_info_map[run_idx] = {
                "x_text": f"{param_label} = {param_text}",
                "y_text": f"{param2_label} = {param2_text}",
                "excluded": bool(exclude_chk.GetValue()) if exclude_chk else False,
            }
            if not existing_panel and auto_open is None:
                auto_open = (key, show_chk)
            created = created or not existing_panel

        if created:
            self.control_panel.Layout()

        if auto_open and auto_open[1] is not None:
            key_to_open, chk = auto_open
            self._set_checkbox_value(chk, True)
            self.on_show_detail_toggle(None, key_to_open, (key_to_open[1],))
        elif auto_open and auto_open[0] in self.detail_windows:
            try:
                self.detail_windows[auto_open[0]].Raise()
            except Exception:
                pass
        elif not auto_open and run_indices:
            first_key = (plot_idx, run_indices[0])
            window = self.detail_windows.get(first_key)
            if window:
                try:
                    window.Raise()
                except Exception:
                    pass

    def on_pick(self, event):
        artist = event.artist
        indices = getattr(event, "ind", None)
        plot_num = self._artist_plot_map.get(artist)
        data = None
        if plot_num is not None:
            data = self.plot_data.get(plot_num)
        if data is None:
            for pn, pdata in self.plot_data.items():
                artists = pdata.get("artists") or [pdata.get("artist")]
                if artist in artists:
                    plot_num = pn
                    data = pdata
                    break
        if data is None or indices is None or len(indices) == 0:
            return
        display_idx = data.get("display_idx", np.arange(len(data.get("x", []))))
        try:
            display_idx = np.asarray(display_idx, dtype=int)
            pick_position = int(indices[0])
            if pick_position < 0 or pick_position >= display_idx.size:
                return
            orig_idx = int(display_idx[pick_position])
        except Exception:
            return
        groups = data.get("orig_indices")
        if groups and orig_idx < len(groups):
            orig_group = groups[orig_idx]
        else:
            orig_group = orig_idx
        orig_indices = tuple(int(i) for i in np.atleast_1d(orig_group))
        key = (plot_num, orig_idx)
        try:
            x_values = np.asarray(data.get("x"), dtype=float)
            if 0 <= orig_idx < x_values.size:
                x_display = float(x_values[orig_idx])
            else:
                x_display = None
        except Exception:
            x_display = None
        label = self._format_exclude_checkbox_label(plot_num, x_display, orig_indices)

        is_multi_mode = (
            self.multiRadio.GetValue()
            and not self.heatmapRadio.GetValue()
            and bool(self.subrun_results)
            and self.current_multi_fit is not None
        )
        detail_callback = None
        detail_tooltip = "Open a window with detailed plots for this data point"
        detail_enabled = False
        detail_disabled_tooltip = None
        if is_multi_mode:
            detail_callback = lambda evt, k=key, inds=orig_indices: self.on_show_detail_toggle(
                evt, k, inds
            )
            detail_enabled = all(0 <= idx < len(self.subrun_results) for idx in orig_indices)
            if not detail_enabled:
                detail_disabled_tooltip = "Detailed plots are unavailable for this point."

        file_names = self._point_file_names(orig_indices)
        self._ensure_point_selection_panel(
            key,
            label,
            orig_indices,
            detail_callback=detail_callback,
            detail_enabled=detail_enabled,
            detail_tooltip=detail_tooltip,
            detail_disabled_tooltip=detail_disabled_tooltip,
            file_names=file_names,
        )

        self.point_label_info.setdefault(key, {})
        self._apply_point_label(
            key,
            plot_num=plot_num,
            x_display=x_display,
            orig_indices=orig_indices,
        )

        self._ensure_exclude_panel_visible()
        window = self.detail_windows.get(key)
        if window:
            try:
                window.Raise()
            except Exception:
                pass

    def _set_checkbox_value(self, checkbox, value):
        if not checkbox:
            return
        if checkbox.GetValue() == value:
            return
        enabled = checkbox.GetEvtHandlerEnabled()
        checkbox.SetEvtHandlerEnabled(False)
        checkbox.SetValue(value)
        checkbox.SetEvtHandlerEnabled(enabled)

    def on_exclude_toggle(self, key, orig_indices):
        plot_num, _ = key
        data = self.plot_data.get(plot_num)
        if not data:
            return
        inds = set(int(i) for i in np.atleast_1d(orig_indices))
        if self.exclude_checks[key].GetValue():
            data.setdefault("excluded", set()).update(inds)
        else:
            data.setdefault("excluded", set()).difference_update(inds)
        if isinstance(self._session_point_exclusions, set):
            if self.exclude_checks[key].GetValue():
                self._session_point_exclusions.update(inds)
            else:
                self._session_point_exclusions.difference_update(inds)
            self._pending_session_point_exclusions = set(self._session_point_exclusions)
        heatmap_mode = self._heatmap_mode_active()
        if heatmap_mode:
            info_map = data.get("heatmap_point_info", {})
            for idx in inds:
                if idx in info_map:
                    info_map[idx]["excluded"] = self.exclude_checks[key].GetValue()
        window = self.detail_windows.get(key)
        if hasattr(window, "sync_exclude_state"):
            window.sync_exclude_state(self.exclude_checks[key].GetValue())
        self._apply_point_label(key)
        if heatmap_mode:
            for idx in inds:
                try:
                    self._update_heatmap_results_for_run(int(idx))
                except Exception:
                    continue
            self.update_heatmap(preserve_details=True)
        else:
            self.redraw_plots()
        self._mark_results_modified()

    def _point_has_sub_exclusions(self, orig_indices):
        if not orig_indices:
            return False
        if not isinstance(orig_indices, (list, tuple, set, np.ndarray)):
            indices = (orig_indices,)
        else:
            indices = orig_indices
        for idx in indices:
            try:
                run_idx = int(idx)
            except Exception:
                continue
            excluded = self.run_point_exclusions.get(run_idx)
            if isinstance(excluded, np.ndarray):
                if excluded.size > 0:
                    return True
            elif isinstance(excluded, (set, list, tuple)):
                if len(excluded):
                    return True
            elif excluded:
                return True
        return False

    def _format_exclude_checkbox_label(self, plot_num, x_display, orig_indices):
        if plot_num is None:
            label = "Exclude data point"
        else:
            label = f"Exclude Plot {plot_num}"
        if x_display is not None:
            try:
                label += f" x={float(x_display):.3g}"
            except Exception:
                label += f" x={x_display}"
        if self._point_has_sub_exclusions(orig_indices):
            label += " (subpoint excluded)"
        return label

    def _apply_point_label(self, key, plot_num=None, x_display=None, orig_indices=None):
        info = self.point_label_info.setdefault(key, {})
        if plot_num is not None:
            info["plot"] = plot_num
        if x_display is not None:
            info["x"] = x_display
        if orig_indices is not None:
            normalized = []
            try:
                iterable = np.atleast_1d(orig_indices)
            except Exception:
                if isinstance(orig_indices, (list, tuple, set)):
                    iterable = orig_indices
                else:
                    iterable = (orig_indices,)
            if hasattr(iterable, "tolist"):
                iterable = iterable.tolist()
            for val in iterable:
                try:
                    normalized.append(int(val))
                except Exception:
                    normalized.append(val)
            info["orig_indices"] = tuple(normalized)
        plot_val = info.get("plot")
        if plot_val is None and isinstance(key, (tuple, list)) and key:
            plot_val = key[0]
            info["plot"] = plot_val
        indices = info.get("orig_indices")
        custom_label = info.get("custom_label")
        if custom_label:
            label = custom_label
            if self._point_has_sub_exclusions(indices):
                if "(subpoint excluded)" not in label:
                    label = f"{label} (subpoint excluded)"
        else:
            label = self._format_exclude_checkbox_label(plot_val, info.get("x"), indices)
        chk = self.exclude_checks.get(key)
        if chk:
            chk.SetLabel(label)
        panel = self.point_panels.get(key)
        if panel:
            panel.Layout()
        self.control_panel.Layout()

    def _refresh_point_label_for_run(self, run_idx):
        try:
            target = int(run_idx)
        except Exception:
            return
        for key, info in list(self.point_label_info.items()):
            indices = info.get("orig_indices") or ()
            for val in indices:
                try:
                    if int(val) == target:
                        self._apply_point_label(key)
                        break
                except Exception:
                    continue

    def apply_point_exclusion(self, key, orig_indices, excluded):
        chk = self.exclude_checks.get(key)
        desired = bool(excluded)
        if chk is not None:
            current = bool(chk.GetValue())
            if current != desired:
                self._set_checkbox_value(chk, desired)
                self.on_exclude_toggle(key, orig_indices)
        self._apply_point_label(key, orig_indices=orig_indices)

    def note_subpoint_exclusions(self, key, orig_indices):
        self._apply_point_label(key, orig_indices=orig_indices)

    def _point_file_names(self, orig_indices):
        if not isinstance(self.results, dict):
            return []
        image_files = self.results.get("image_file")
        if image_files is None:
            return []
        try:
            files = list(image_files)
        except TypeError:
            return []
        names = []
        for idx in orig_indices:
            if 0 <= idx < len(files):
                name = files[idx]
                if isinstance(name, str):
                    names.append(os.path.basename(name))
                else:
                    names.append(str(name))
        return names

    def _heatmap_point_file_names(self, run_idx):
        if not self.subrun_results or not (0 <= run_idx < len(self.subrun_results)):
            return []
        run = self.subrun_results[run_idx] or {}
        results = run.get("results") or {}
        files = results.get("image_file")
        if not files:
            return []
        try:
            iterator = list(files)
        except TypeError:
            return []
        names = []
        for entry in iterator:
            if isinstance(entry, str):
                names.append(os.path.basename(entry))
            else:
                names.append(str(entry))
        return names

    def _point_file_paths(self, orig_indices, column=None):
        entries = []
        seen = set()
        column = column or ""
        y_unit, y_factor = self.col_scales.get(column, ("", 1.0))
        if self.subrun_results:
            for idx in orig_indices:
                if not (0 <= idx < len(self.subrun_results)):
                    continue
                run = self.subrun_results[idx]
                files = run.get("results", {}).get("image_file")
                if not files:
                    continue
                try:
                    iterator = list(files)
                except TypeError:
                    continue
                var_values = run.get("var_values", [])
                run_results = run.get("results", {}) or {}
                column_values = run_results.get(column) if column else None
                var_scale = run.get("var_scale") or self.var_scale or 1.0
                for pos, path in enumerate(iterator):
                    if not isinstance(path, str) or path in seen:
                        continue
                    seen.add(path)
                    x_val = None
                    if pos < len(var_values):
                        try:
                            raw_x = float(var_values[pos])
                        except (TypeError, ValueError):
                            raw_x = var_values[pos]
                        if isinstance(raw_x, (int, float)):
                            scale = var_scale if var_scale else 1.0
                            x_val = raw_x / scale if scale else raw_x
                        else:
                            x_val = raw_x
                    y_val = None
                    if isinstance(column_values, (list, tuple)) and pos < len(column_values):
                        try:
                            raw_y = float(column_values[pos])
                        except (TypeError, ValueError):
                            raw_y = column_values[pos]
                        if isinstance(raw_y, (int, float)):
                            y_val = raw_y * (y_factor if y_factor else 1.0)
                        else:
                            y_val = raw_y
                    entry = {"path": path, "x": x_val, "y": y_val, "y_unit": y_unit}
                    entries.append(entry)
            if entries:
                return entries
        if not isinstance(self.results, dict):
            return entries
        image_files = self.results.get("image_file")
        if image_files is None:
            return entries
        try:
            files = list(image_files)
        except TypeError:
            return entries
        var_values = list(self.var_values) if isinstance(self.var_values, (list, tuple, np.ndarray)) else []
        column_values = None
        if column and column in self.results:
            try:
                column_values = list(self.results.get(column))
            except TypeError:
                column_values = None
        for idx in orig_indices:
            if not (0 <= idx < len(files)):
                continue
            path = files[idx]
            if not isinstance(path, str) or path in seen:
                continue
            seen.add(path)
            x_val = None
            if idx < len(var_values):
                try:
                    raw_x = float(var_values[idx])
                except (TypeError, ValueError):
                    raw_x = var_values[idx]
                scale = self.var_scale or 1.0
                if isinstance(raw_x, (int, float)):
                    x_val = raw_x / (scale if scale else 1.0)
                else:
                    x_val = raw_x
            y_val = None
            if isinstance(column_values, list) and idx < len(column_values):
                try:
                    raw_y = float(column_values[idx])
                except (TypeError, ValueError):
                    raw_y = column_values[idx]
                if isinstance(raw_y, (int, float)):
                    y_val = raw_y * (y_factor if y_factor else 1.0)
                else:
                    y_val = raw_y
            entries.append({"path": path, "x": x_val, "y": y_val, "y_unit": y_unit})
        return entries

    def _format_point_files(self, names):
        if not names:
            return ""
        if len(names) == 1:
            return f"File: {names[0]}"
        if len(names) <= 3:
            return "Files: " + ", ".join(names)
        remaining = len(names) - 3
        return "Files: " + ", ".join(names[:3]) + f", +{remaining} more"

    def _build_filtered_run_components(self, run, exclusions, apply_binning=True):
        if not run:
            return [], {}, [], [], {}
        try:
            var_values = list(run.get("var_values", []))
        except Exception:
            var_values = []
        total = len(var_values)
        mask = [True] * total
        try:
            exclusion_set = {int(val) for val in exclusions or []}
        except Exception:
            exclusion_set = set(exclusions or [])
        for idx in sorted(exclusion_set):
            if 0 <= idx < total:
                mask[idx] = False
        filtered_var = [val for i, val in enumerate(var_values) if i >= len(mask) or mask[i]]
        filtered_results = {}
        results = run.get("results", {}) or {}
        for key, values in results.items():
            if isinstance(values, (list, tuple)):
                filtered_results[key] = [
                    val for i, val in enumerate(values) if i >= len(mask) or mask[i]
                ]
            elif isinstance(values, np.ndarray):
                if not total:
                    filtered_results[key] = np.asarray(values).tolist()
                else:
                    base_mask = list(mask)
                    if len(base_mask) < len(values):
                        base_mask.extend([True] * (len(values) - len(base_mask)))
                    bool_mask = np.array(base_mask[: len(values)], dtype=bool)
                    try:
                        filtered_array = values[bool_mask]
                    except Exception:
                        filtered_array = values
                    filtered_results[key] = np.asarray(filtered_array).tolist()
            else:
                filtered_results[key] = values
        included_indices = [i for i, include in enumerate(mask) if include]
        excluded_indices = [i for i, include in enumerate(mask) if not include]
        groups = {}
        var_bin_active = getattr(self, "varBinCheck", None)
        var_bin_active = bool(var_bin_active and var_bin_active.GetValue())
        heatmap_bin_active = self._heatmap_var_binning_enabled()
        if apply_binning and included_indices and (var_bin_active or heatmap_bin_active):
            (
                filtered_var,
                filtered_results,
                included_indices,
                groups,
            ) = self._apply_variable_binning_preview(
                filtered_var, filtered_results, included_indices
            )
        return filtered_var, filtered_results, included_indices, excluded_indices, groups

    def _filtered_run_data(self, run_idx):
        if not self.subrun_results or not (0 <= run_idx < len(self.subrun_results)):
            self.variable_bin_groups.pop(run_idx, None)
            return [], {}, [], []
        run = self.subrun_results[run_idx]
        (
            filtered_var,
            filtered_results,
            included_indices,
            excluded_indices,
            groups,
        ) = self._build_filtered_run_components(
            run, self.run_point_exclusions.get(run_idx, set()), apply_binning=True
        )
        if groups:
            self.variable_bin_groups[run_idx] = groups
        else:
            self.variable_bin_groups.pop(run_idx, None)
        return filtered_var, filtered_results, included_indices, excluded_indices

    def _aggregate_samples(self, samples):
        if not samples:
            return np.nan
        numeric = []
        for sample in samples:
            try:
                value = float(sample)
            except (TypeError, ValueError):
                numeric = None
                break
            if not np.isfinite(value):
                numeric = None
                break
            numeric.append(value)
        if numeric:
            return float(np.mean(numeric))
        return samples[0]

    def _build_param_bin_analysis_context(self, groups, base_var, base_results):
        if not groups:
            return None
        try:
            base_array = np.asarray(base_var, dtype=float)
        except Exception:
            base_array = np.asarray(base_var)
        if base_array.ndim != 1 or base_array.size == 0:
            return None

        aggregated_var = []
        for grp in groups:
            idxs = np.asarray(np.atleast_1d(grp), dtype=int)
            idxs = idxs[(idxs >= 0) & (idxs < base_array.size)]
            if idxs.size == 0:
                aggregated_var.append(np.nan)
                continue
            samples = base_array[idxs]
            try:
                aggregated_var.append(float(np.mean(samples)))
            except Exception:
                aggregated_var.append(float(samples[0]) if samples.size else np.nan)
        aggregated_var = np.asarray(aggregated_var, dtype=float)

        base_len = base_array.size
        aggregated_results = {}
        for key, values in base_results.items():
            arr = None
            if isinstance(values, np.ndarray) and values.shape and values.shape[0] == base_len:
                arr = values
            elif isinstance(values, (list, tuple)) and len(values) == base_len:
                arr = np.asarray(values)
            if arr is None:
                aggregated_results[key] = values
                continue
            aggregated_series = []
            for grp in groups:
                idxs = np.asarray(np.atleast_1d(grp), dtype=int)
                idxs = idxs[(idxs >= 0) & (idxs < base_len)]
                if idxs.size == 0:
                    aggregated_series.append(np.nan)
                    continue
                try:
                    samples = arr[idxs]
                except Exception:
                    samples = np.asarray([], dtype=float)
                if np.size(samples) == 0:
                    aggregated_series.append(np.nan)
                    continue
                aggregated_series.append(
                    self._aggregate_samples(np.asarray(samples).ravel().tolist())
                )
            aggregated_results[key] = np.asarray(aggregated_series)

        return aggregated_var, aggregated_results

    def _apply_variable_binning(
        self, run_idx, var_values, results, included_indices
    ):
        try:
            value_array = np.asarray(var_values, dtype=float)
        except Exception:
            value_array = np.array(var_values, dtype=float)
        if value_array.size == 0:
            return var_values, results, included_indices, {}
        groups = []
        if value_array.size:
            try:
                base_tol = float(np.nanmax(np.abs(value_array)))
            except (ValueError, TypeError):
                base_tol = 0.0
        else:
            base_tol = 0.0
        if not np.isfinite(base_tol):
            base_tol = 0.0
        tolerance = max(base_tol * 1e-9, 1e-12)
        for pos, value in enumerate(var_values):
            placed = False
            for group in groups:
                if abs(value - group["value"]) <= tolerance:
                    group["positions"].append(pos)
                    group["values"].append(value)
                    group["orig_indices"].append(included_indices[pos])
                    placed = True
                    break
            if not placed:
                groups.append(
                    {
                        "value": float(value),
                        "positions": [pos],
                        "values": [value],
                        "orig_indices": [included_indices[pos]],
                    }
                )
        if not groups:
            return var_values, results, included_indices, {}
        aggregated_var = []
        aggregated_results = {key: [] for key in results.keys()}
        representative_indices = []
        group_mapping = {}
        for group in groups:
            rep_idx = int(min(group["orig_indices"]))
            representative_indices.append(rep_idx)
            group_mapping[rep_idx] = tuple(sorted(int(idx) for idx in group["orig_indices"]))
            try:
                mean_val = float(np.mean(group["values"]))
            except Exception:
                mean_val = group["values"][0]
            aggregated_var.append(mean_val)
            for key, values in results.items():
                seq = list(values) if isinstance(values, (list, tuple)) else [values]
                samples = [seq[pos] for pos in group["positions"] if pos < len(seq)]
                if not samples:
                    aggregated_results[key].append(np.nan)
                else:
                    aggregated_results[key].append(self._aggregate_samples(samples))
        return aggregated_var, aggregated_results, representative_indices, group_mapping

    def _apply_variable_binning_preview(self, var_values, results, included_indices):
        try:
            value_array = np.asarray(var_values, dtype=float)
        except Exception:
            value_array = np.array(var_values, dtype=float)
        if value_array.size == 0:
            return var_values, results, included_indices, {}
        groups = []
        if value_array.size:
            try:
                base_tol = float(np.nanmax(np.abs(value_array)))
            except (ValueError, TypeError):
                base_tol = 0.0
        else:
            base_tol = 0.0
        if not np.isfinite(base_tol):
            base_tol = 0.0
        tolerance = max(base_tol * 1e-9, 1e-12)
        for pos, value in enumerate(var_values):
            placed = False
            for group in groups:
                if abs(value - group["value"]) <= tolerance:
                    group["positions"].append(pos)
                    group["values"].append(value)
                    group["orig_indices"].append(included_indices[pos])
                    placed = True
                    break
            if not placed:
                groups.append(
                    {
                        "value": float(value),
                        "positions": [pos],
                        "values": [value],
                        "orig_indices": [included_indices[pos]],
                    }
                )
        if not groups:
            return var_values, results, included_indices, {}
        aggregated_var = []
        aggregated_results = {key: [] for key in results.keys()}
        representative_indices = []
        group_mapping = {}
        for group in groups:
            if not group["orig_indices"]:
                continue
            rep_idx = int(min(group["orig_indices"]))
            representative_indices.append(rep_idx)
            group_mapping[rep_idx] = tuple(sorted(int(idx) for idx in group["orig_indices"]))
            try:
                mean_val = float(np.mean(group["values"]))
            except Exception:
                mean_val = group["values"][0]
            aggregated_var.append(mean_val)
            for key, values in results.items():
                if isinstance(values, (list, tuple)):
                    seq = list(values)
                else:
                    seq = [values]
                samples = [seq[pos] for pos in group["positions"] if pos < len(seq)]
                if not samples:
                    aggregated_results[key].append(np.nan)
                else:
                    aggregated_results[key].append(self._aggregate_samples(samples))
        return aggregated_var, aggregated_results, representative_indices, group_mapping

    def _collect_detail_runs(self, orig_indices):
        runs = []
        if not self.subrun_results:
            return runs
        for idx in orig_indices:
            if 0 <= idx < len(self.subrun_results):
                run_data = copy.deepcopy(self.subrun_results[idx])
                (
                    filtered_var,
                    filtered_results,
                    included,
                    excluded,
                ) = self._filtered_run_data(idx)
                run_data["index"] = idx
                run_data["filtered_var_values"] = filtered_var
                run_data["filtered_results"] = filtered_results
                run_data["included_indices"] = included
                run_data["excluded_indices"] = excluded
                if idx in self.variable_bin_groups:
                    run_data["variable_groups"] = self.variable_bin_groups[idx]
                else:
                    run_data["variable_groups"] = {}
                runs.append(run_data)
        return runs

    def collect_detail_runs_preview(self, orig_indices, preview_exclusions):
        runs = []
        if not self.subrun_results:
            return runs
        preview_map = preview_exclusions or {}
        for idx in orig_indices:
            if not (0 <= idx < len(self.subrun_results)):
                continue
            run_copy = copy.deepcopy(self.subrun_results[idx])
            if idx in preview_map:
                try:
                    exclusions = {int(val) for val in preview_map[idx]}
                except Exception:
                    exclusions = set(preview_map[idx])
            else:
                exclusions = set(self.run_point_exclusions.get(idx, set()))
            (
                filtered_var,
                filtered_results,
                included,
                excluded,
                groups,
            ) = self._build_filtered_run_components(
                run_copy, exclusions, apply_binning=True
            )
            run_copy["filtered_var_values"] = filtered_var
            run_copy["filtered_results"] = filtered_results
            run_copy["included_indices"] = included
            run_copy["excluded_indices"] = excluded
            run_copy["variable_groups"] = groups
            run_copy["index"] = idx
            runs.append(run_copy)
        return runs

    def set_run_measurement_excluded(self, run_idx, measurement_idx, excluded):
        if not (0 <= run_idx < len(self.subrun_results)):
            return False
        group_map = self.variable_bin_groups.get(run_idx, {})
        target_indices = None
        if measurement_idx in group_map:
            target_indices = group_map[measurement_idx]
        else:
            for members in group_map.values():
                if measurement_idx in members:
                    target_indices = members
                    break
        if not target_indices:
            target_indices = (measurement_idx,)
        exclusions = self.run_point_exclusions.setdefault(run_idx, set())
        changed = False
        if excluded:
            for idx in target_indices:
                if idx not in exclusions:
                    exclusions.add(idx)
                    changed = True
        else:
            for idx in target_indices:
                if idx in exclusions:
                    exclusions.remove(idx)
                    changed = True
            if not exclusions and run_idx in self.run_point_exclusions:
                self.run_point_exclusions.pop(run_idx, None)
        if not changed:
            return False
        heatmap_mode = self._heatmap_mode_active()
        self._recompute_run_metrics(run_idx)
        if heatmap_mode:
            self.update_heatmap(preserve_details=True)
        else:
            self._refresh_plot_data_values()
            self.redraw_plots()
        self._broadcast_run_update(run_idx)
        self._mark_results_modified()
        self._refresh_point_label_for_run(run_idx)
        self._session_run_exclusions = {
            idx: set(vals) for idx, vals in self.run_point_exclusions.items()
        }
        self._pending_session_run_exclusions = {
            idx: set(vals) for idx, vals in self.run_point_exclusions.items()
        }
        return True

    def set_run_exclusions_bulk(self, run_idx, excluded_indices):
        if not (0 <= run_idx < len(self.subrun_results)):
            return False
        try:
            target = {int(idx) for idx in excluded_indices or []}
        except Exception:
            target = set(excluded_indices or [])
        current = set(self.run_point_exclusions.get(run_idx, set()))
        if current == target:
            self._refresh_point_label_for_run(run_idx)
            return False
        if target:
            self.run_point_exclusions[run_idx] = target
        else:
            self.run_point_exclusions.pop(run_idx, None)
        heatmap_mode = self._heatmap_mode_active()
        self._recompute_run_metrics(run_idx)
        if heatmap_mode:
            self.update_heatmap(preserve_details=True)
        else:
            self._refresh_plot_data_values()
            self.redraw_plots()
        self._broadcast_run_update(run_idx)
        self._mark_results_modified()
        self._refresh_point_label_for_run(run_idx)
        self._session_run_exclusions = {
            idx: set(vals) for idx, vals in self.run_point_exclusions.items()
        }
        self._pending_session_run_exclusions = {
            idx: set(vals) for idx, vals in self.run_point_exclusions.items()
        }
        return True

    def _recompute_run_metrics(self, run_idx):
        if not (0 <= run_idx < len(self.subrun_results)):
            return
        base_run = self.subrun_results[run_idx]
        (
            filtered_var,
            filtered_results,
            included,
            excluded,
        ) = self._filtered_run_data(run_idx)
        base_run["filtered_var_values"] = filtered_var
        base_run["filtered_results"] = filtered_results
        base_run["included_indices"] = included
        base_run["excluded_indices"] = excluded
        base_run["variable_groups"] = self.variable_bin_groups.get(run_idx, {})
        func_key = self.current_multi_fit or self._selected_multi_fit_key()
        if not func_key:
            return
        if func_key == HEATMAP_VALUE_FIT_KEY:
            derived = self._build_heatmap_value_results(filtered_results)
        else:
            fit_info = FIT_FUNCTIONS.get(func_key)
            if fit_info is None:
                derived = {}
            else:
                derived = fit_info["derived"](
                    None,
                    None,
                    filtered_var,
                    filtered_results,
                    self.parent,
                )
        if self._heatmap_mode_active():
            base_run["derived"] = dict(derived)
            self._update_heatmap_results_for_run(run_idx)
            return
        if not isinstance(self.results, dict):
            return
        for key, values in self.results.items():
            if not isinstance(values, list):
                try:
                    seq = list(values)
                except TypeError:
                    continue
                else:
                    self.results[key] = seq
                    values = self.results[key]
            if run_idx >= len(values):
                continue
            values[run_idx] = derived.get(key, np.nan)

    def _update_heatmap_results_for_run(self, run_idx):
        if not isinstance(self.results, dict):
            return
        grid = getattr(self, "heatmap_index_grid", None)
        if grid is None or not isinstance(grid, np.ndarray):
            return
        try:
            total_rows, total_cols = grid.shape
        except Exception:
            return
        excluded_runs = self._collect_heatmap_excluded_runs()
        targets = []
        for r in range(total_rows):
            for c in range(total_cols):
                cell_runs = grid[r, c]
                if not cell_runs:
                    continue
                try:
                    members = tuple(int(val) for val in np.atleast_1d(cell_runs))
                except Exception:
                    try:
                        members = tuple(cell_runs)
                    except Exception:
                        members = (cell_runs,)
                if run_idx in members:
                    targets.append((r, c, members))
        if not targets:
            return
        for row, col, members in targets:
            self._recalculate_heatmap_cell(row, col, members, excluded_runs)

    def _collect_heatmap_excluded_runs(self):
        excluded_runs = set()
        for pdata in self.plot_data.values():
            if not isinstance(pdata, dict):
                continue
            excluded = pdata.get("excluded", set())
            if not excluded:
                continue
            for val in excluded:
                try:
                    excluded_runs.add(int(val))
                except Exception:
                    continue
        return excluded_runs

    def _collect_heatmap_cell_values(self, column, members, excluded_runs=None):
        if not column:
            return np.asarray([], dtype=float)
        if excluded_runs is None:
            excluded_runs = set()
        values = []
        for idx in self._iter_heatmap_runs(members):
            if idx in excluded_runs:
                continue
            if not (0 <= idx < len(self.subrun_results)):
                continue
            run = self.subrun_results[idx] or {}
            derived = run.get("derived", {}) or {}
            val = derived.get(column)
            if val is None:
                val = np.nan
            try:
                values.append(float(val))
            except (TypeError, ValueError):
                values.append(np.nan)
        return np.asarray(values, dtype=float)

    def _extract_heatmap_cell_value(self, column, row, col):
        if not column:
            return np.nan
        if not isinstance(self.results, dict) or column not in self.results:
            return np.nan
        try:
            arr = np.asarray(self.results[column], dtype=float)
        except Exception:
            return np.nan
        if arr.ndim != 2:
            return np.nan
        if not (0 <= row < arr.shape[0] and 0 <= col < arr.shape[1]):
            return np.nan
        try:
            return float(arr[row, col])
        except (TypeError, ValueError):
            return np.nan

    def _extract_heatmap_cell_error(self, column, row, col, members, excluded_runs=None):
        if not column or self._is_error_key(column):
            return np.nan
        if excluded_runs is None:
            excluded_runs = set()
        error_val = np.nan
        results = self.results if isinstance(self.results, dict) else {}
        error_key = self._find_matching_error_key(column, results)
        if error_key:
            error_val = self._extract_heatmap_cell_value(error_key, row, col)
            if np.isfinite(error_val):
                return error_val
        for idx in self._iter_heatmap_runs(members):
            if idx in excluded_runs or not (0 <= idx < len(self.subrun_results)):
                continue
            run = self.subrun_results[idx] or {}
            derived = run.get("derived", {}) or {}
            if not derived:
                continue
            run_error_key = self._find_matching_error_key(column, derived)
            if not run_error_key:
                continue
            try:
                candidate = float(derived.get(run_error_key, np.nan))
            except (TypeError, ValueError):
                continue
            if np.isfinite(candidate):
                return candidate
        return error_val

    def _compute_heatmap_cell_stats(self, column, row, col, excluded_runs=None):
        grid = getattr(self, "heatmap_index_grid", None)
        if grid is None or not isinstance(grid, np.ndarray):
            return (np.nan, np.nan, 0)
        if not (0 <= row < grid.shape[0] and 0 <= col < grid.shape[1]):
            return (np.nan, np.nan, 0)
        if excluded_runs is None:
            excluded_runs = self._collect_heatmap_excluded_runs()
        members = grid[row, col]
        values = self._collect_heatmap_cell_values(column, members, excluded_runs)
        if values.size == 0:
            mean_val = self._extract_heatmap_cell_value(column, row, col)
            return (mean_val, np.nan, 0)
        finite = values[np.isfinite(values)]
        count = int(finite.size)
        mean_val = self._extract_heatmap_cell_value(column, row, col)
        if count == 0:
            return (mean_val, np.nan, 0)
        if not np.isfinite(mean_val):
            mean_val = float(np.mean(finite)) if count else np.nan
        if count >= 2:
            spread = float(np.std(finite, ddof=1))
        elif count == 1:
            spread = self._extract_heatmap_cell_error(
                column, row, col, members, excluded_runs
            )
            if not np.isfinite(spread):
                spread = np.nan
        else:
            spread = np.nan
        return (mean_val, spread, count)

    def _recalculate_heatmap_cell(self, row, col, members, excluded_runs=None):
        if excluded_runs is None:
            excluded_runs = set()
        if not isinstance(members, (list, tuple, set, np.ndarray)):
            members = (members,)
        for key, values in list(self.results.items()):
            try:
                arr = np.asarray(values, dtype=float)
            except Exception:
                continue
            if arr.ndim != 2:
                continue
            if row >= arr.shape[0] or col >= arr.shape[1]:
                continue
            arr_vals = self._collect_heatmap_cell_values(key, members, excluded_runs)
            if arr_vals.size == 0:
                new_val = np.nan
            else:
                finite = arr_vals[np.isfinite(arr_vals)]
                if finite.size == 0:
                    new_val = np.nan
                else:
                    new_val = float(np.mean(finite))
            arr[row, col] = new_val
            self.results[key] = arr

    def _refresh_plot_data_values(self):
        for plot_num, pdata in self.plot_data.items():
            column = pdata.get("column")
            if not column or not isinstance(self.results, dict):
                continue
            if column not in self.results:
                continue
            x = np.asarray(self.var_values)
            y = np.asarray(self.results.get(column, []), dtype=float)
            err_key = self._find_error_key(column)
            if err_key and err_key in self.results:
                try:
                    y_err = np.asarray(self.results.get(err_key), dtype=float)
                except Exception:
                    y_err = None
            else:
                y_err = None

            groups = [np.array([idx]) for idx in range(len(x))]
            if self.paramBinCheck.GetValue():
                x, y, y_err, groups = self._bin_data(x, y, y_err)

            pdata["x"] = x
            pdata["y"] = y
            pdata["y_err"] = y_err
            pdata["orig_indices"] = groups

    def _broadcast_run_update(self, run_idx):
        for key, window in list(self.detail_windows.items()):
            if not isinstance(window, PointDetailFrame):
                continue
            if run_idx not in window.orig_indices:
                continue
            runs = self._collect_detail_runs(window.orig_indices)
            image_paths = self._point_file_paths(window.orig_indices, window.column)
            window.update_context(
                runs,
                window.point_info,
                image_paths,
                window.fit_key,
                window.display_name,
                window.column,
            )

    def _open_point_inspector(self, key, orig_indices, data, orig_idx):
        runs = self._collect_detail_runs(orig_indices)
        plot_num, _ = key
        column = None
        if data:
            column = data.get("column")
        if self.multiRadio.GetValue() and not self.heatmapRadio.GetValue():
            display_name = (
                self.multiFitChoice.GetStringSelection() or TEMP_PSD_DISPLAY_NAME
            )
            display_name = _normalise_fit_name(display_name)
            fit_key = self.current_multi_fit
        else:
            raw_name = data.get("func_name") if data else None
            display_name = _normalise_fit_name(raw_name) if raw_name else "Data point"
            fit_key = raw_name

        x_label = data.get("x_label") if data else ""
        x_unit = data.get("x_unit") if data else ""
        x_scale = data.get("x_scale") if data else 1.0
        x_text = ""
        try:
            x_values = np.asarray(data.get("x"), dtype=float)
        except Exception:
            x_values = None
        if x_values is not None and 0 <= orig_idx < x_values.size:
            raw_val = float(x_values[orig_idx])
            denom = x_scale if x_scale else 1.0
            display_val = raw_val / denom
            label = x_label or "x"
            unit_str = f" {x_unit}" if x_unit else ""
            x_text = f"{label} = {display_val:.6g}{unit_str}".strip()

        y_text = ""
        y_values = None
        if data:
            try:
                y_values = np.asarray(data.get("y"), dtype=float)
            except Exception:
                y_values = None
        if y_values is not None and 0 <= orig_idx < y_values.size:
            raw_y = float(y_values[orig_idx])
            y_unit, y_factor = self.col_scales.get(column, ("", 1.0))
            if y_factor is None:
                y_factor = 1.0
            display_y = raw_y * y_factor
            label = column or "y"
            unit_str = f" {y_unit}" if y_unit else ""
            y_text = f"{label} = {display_y:.6g}{unit_str}".strip()

        excluded = False
        chk = self.exclude_checks.get(key)
        if chk:
            excluded = chk.GetValue()

        point_info = {
            "plot": plot_num,
            "x_text": x_text,
            "y_text": y_text,
            "excluded": excluded,
        }
        if data:
            heatmap_info = data.get("heatmap_point_info")
            if heatmap_info and orig_idx in heatmap_info:
                point_info.update(heatmap_info[orig_idx])
        image_paths = self._point_file_paths(orig_indices, column)

        window = self.detail_windows.get(key)
        if isinstance(window, PointDetailFrame):
            window.update_context(runs, point_info, image_paths, fit_key, display_name, column)
            return window

        try:
            inspector = PointDetailFrame(
                self,
                key,
                orig_indices,
                runs,
                fit_key,
                display_name,
                column,
                point_info,
                image_paths,
            )
        except Exception:
            return None
        self.detail_windows[key] = inspector
        inspector.Bind(
            wx.EVT_CLOSE,
            lambda evt, k=key: self._on_detail_window_closed(k, evt),
        )
        inspector.Show()
        return inspector

    def on_show_detail_toggle(self, event, key, orig_indices):
        chk = self.detail_checks.get(key)
        if not chk:
            return
        if chk.GetValue():
            if key in self.detail_windows and self.detail_windows[key]:
                self.detail_windows[key].Raise()
                return
            if key not in getattr(self, "heatmap_detail_keys", set()):
                data = self.plot_data.get(key[0])
                window = self._open_point_inspector(key, orig_indices, data, key[1])
                if window:
                    return
            try:
                window = self._create_detail_window(key, orig_indices)
            except Exception as exc:  # pragma: no cover - UI safeguard
                chk.SetValue(False)
                wx.MessageBox(
                    f"Unable to open detail plot:\n{exc}",
                    "Detail Plot",
                    style=wx.ICON_WARNING | wx.OK,
                )
                return
            self.detail_windows[key] = window
            window.Bind(
                wx.EVT_CLOSE,
                lambda evt, k=key: self._on_detail_window_closed(k, evt),
            )
            window.Show()
        else:
            window = self.detail_windows.pop(key, None)
            if window:
                window.Destroy()

    def _create_detail_window(self, key, orig_indices):
        orig_indices = tuple(orig_indices or ())
        runs = self._collect_detail_runs(orig_indices)
        if not runs:
            raise ValueError("No run data available for the selected point.")
        display_name = (
            self.multiFitChoice.GetStringSelection() or TEMP_PSD_DISPLAY_NAME
        )
        display_name = _normalise_fit_name(display_name)
        plot_num, _ = key
        column = None
        data = self.plot_data.get(plot_num)
        if data:
            column = data.get("column")
        point_info = {"plot": plot_num}
        chk = self.exclude_checks.get(key)
        if chk is not None:
            point_info["excluded"] = bool(chk.GetValue())
        else:
            point_info["excluded"] = False
        run_idx = None
        if orig_indices:
            run_idx = orig_indices[0]
        heatmap_map = data.get("heatmap_point_info") if data else None
        if heatmap_map is not None and run_idx in heatmap_map:
            extra = heatmap_map.get(run_idx) or {}
            if isinstance(extra, dict):
                point_info.update(extra)
        image_paths = self._point_file_paths(orig_indices, column)
        return PointDetailFrame(
            self,
            key,
            orig_indices,
            runs,
            self.current_multi_fit,
            display_name,
            column,
            point_info,
            image_paths,
        )

    def _on_detail_window_closed(self, key, event):
        self.detail_windows.pop(key, None)
        chk = self.detail_checks.get(key)
        if chk and chk.GetValue():
            chk.SetValue(False)
        event.Skip()

    def on_save(self, event=None):
        if not self.results:
            return
        folder = self.savePathCtrl.GetValue()
        if not folder:
            return
        os.makedirs(folder, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        excluded_union = set()
        for pdata in self.plot_data.values():
            excluded_union.update(pdata.get("excluded", set()))
        self._session_point_exclusions = set(excluded_union)
        if self.heatmapRadio.GetValue() and self.subrun_results:
            base_name = f"{timestamp}-heatmap"
            self._write_session_bundle(base_name, folder, excluded_union)
            plot_path = os.path.join(folder, f"{base_name}.png")
            self.update_heatmap()
            self.canvas.draw()
            self.canvas.figure.savefig(plot_path, bbox_inches="tight")
        elif self.multiRadio.GetValue() and self.subrun_results:
            base_name = f"{timestamp}-multirun"
            self._write_session_bundle(base_name, folder, excluded_union)
            plot_path = os.path.join(folder, f"{base_name}.png")
            self.redraw_plots()
            self.canvas.draw()
            self.canvas.figure.savefig(plot_path, bbox_inches="tight")
        else:
            parts = []
            for idx in sorted(self.plot_data.keys()):
                data = self.plot_data[idx]
                col = data.get("column", "data")
                func = data.get("func_name") or "NoFit"
                col_clean = "".join(ch for ch in col if ch.isalnum())
                func_clean = "".join(ch for ch in func if ch.isalnum())
                parts.append(f"{col_clean}-{func_clean}")
            if not parts:
                parts.append("data")
            base_name = f"{timestamp}-{'-'.join(parts)}"
            self._write_session_bundle(base_name, folder, excluded_union)
            plot_path = os.path.join(folder, f"{base_name}.png")
            self.redraw_plots()
            self.canvas.draw()
            self.canvas.figure.savefig(plot_path, bbox_inches="tight")
        self._mark_unsaved(False)

    def _write_session_bundle(self, base_name, folder, excluded_union):
        if not base_name or not folder:
            return
        excluded_union = set(excluded_union or ())

        def _to_plain(value):
            if isinstance(value, np.generic):
                return value.item()
            if isinstance(value, np.ndarray):
                try:
                    return value.tolist()
                except Exception:
                    return value
            return value

        def _coerce_scalar(value):
            if value is None:
                return ""
            plain = _to_plain(value)
            if plain is None:
                return ""
            if isinstance(plain, (int, float)):
                return plain
            if isinstance(plain, (list, tuple)):
                return ", ".join(str(_to_plain(v)) for v in plain)
            if isinstance(plain, dict):
                try:
                    return json.dumps({str(k): _to_plain(v) for k, v in plain.items()})
                except Exception:
                    return str(plain)
            return plain

        def _scale_value(value, scale):
            result = _to_plain(value)
            if result is None:
                return ""
            if scale not in (None, 0):
                try:
                    result = result / scale
                except Exception:
                    pass
            return _coerce_scalar(result)

        def _prepare_series(values, expected_len):
            if values is None:
                return None
            if isinstance(values, np.ndarray):
                if values.ndim != 1:
                    return None
                series = values.tolist()
            elif isinstance(values, np.generic):
                series = [values.item()] * max(expected_len, 1)
            elif isinstance(values, str):
                series = [values] * max(expected_len, 1)
            elif isinstance(values, (list, tuple)):
                series = list(values)
            else:
                try:
                    series = list(values)
                except TypeError:
                    series = [values] * max(expected_len, 1)
            if expected_len and len(series) != expected_len:
                if len(series) < expected_len:
                    series = series + [None] * (expected_len - len(series))
                else:
                    series = series[:expected_len]
            return series

        def _clean_sheet_title(name, fallback):
            title = (name or "").strip()
            if not title:
                title = fallback
            title = re.sub(r"[\\/*?:\[\]]", "_", title)
            title = title or fallback
            return title[:31]

        def _format_header(label, unit):
            label = label or ""
            unit = unit or ""
            return f"{label} ({unit})" if unit else label

        def _populate_workbook(workbook):
            results = self.results or {}
            row_count = len(self.var_values or [])
            sheet = workbook.active
            sheet.title = _clean_sheet_title("Run Data", "Run Data")
            headers = [
                "Index",
                _format_header(self.varLabelCtrl.GetValue(), self.unit_scale),
            ]
            series_items = []
            for key, values in results.items():
                series = _prepare_series(values, row_count)
                if series is None:
                    continue
                series_items.append((key, series))
                headers.append(key)
            sheet.append(headers)
            for idx in range(row_count):
                if idx in excluded_union:
                    continue
                row = [idx, _scale_value(self.var_values[idx], self.var_scale)]
                for _, seq in series_items:
                    if idx < len(seq):
                        row.append(_coerce_scalar(seq[idx]))
                    else:
                        row.append("")
                sheet.append(row)

            if self.subrun_results:
                subrun_sheet = workbook.create_sheet(
                    title=_clean_sheet_title("Subrun Data", "Subrun Data")
                )
                result_keys = []
                for run in self.subrun_results:
                    run_results_map = (run or {}).get("results") or {}
                    for key in run_results_map.keys():
                        if key == "image_file":
                            continue
                        if key not in result_keys:
                            result_keys.append(key)
                include_param2 = bool(self.param2_values)
                if not include_param2:
                    include_param2 = any("param2" in (run or {}) for run in self.subrun_results)
                headers = ["Run Index", _format_header(self.paramLabelCtrl.GetValue(), self.param_unit_scale)]
                if include_param2:
                    headers.append(
                        _format_header(self.param2LabelCtrl.GetValue(), self.param2_unit_scale)
                    )
                headers.append(_format_header(self.varLabelCtrl.GetValue(), self.unit_scale))
                headers.append("Image File")
                headers.extend(result_keys)
                subrun_sheet.append(headers)
                global_idx = 0
                for run_idx, run in enumerate(self.subrun_results, 1):
                    run = run or {}
                    run_results = run.get("results") or {}
                    run_var_values = run.get("var_values") or []
                    run_param = run.get("param")
                    run_param2 = run.get("param2")
                    run_param_scale = run.get("param_scale") or self.param_scale or 1.0
                    run_param2_scale = run.get("param2_scale") or self.param2_scale or 1.0
                    run_var_scale = run.get("var_scale") or self.var_scale or 1.0
                    image_files = _prepare_series(
                        run_results.get("image_file"), len(run_var_values)
                    ) or []
                    keyed_results = {
                        key: _prepare_series(run_results.get(key), len(run_var_values))
                        for key in result_keys
                    }
                    for i, value in enumerate(run_var_values):
                        if global_idx in excluded_union:
                            global_idx += 1
                            continue
                        row = [
                            run_idx,
                            _scale_value(run_param, run_param_scale),
                        ]
                        if include_param2:
                            row.append(_scale_value(run_param2, run_param2_scale))
                        row.append(_scale_value(value, run_var_scale))
                        row.append(
                            image_files[i] if i < len(image_files) else ""
                        )
                        for key in result_keys:
                            seq = keyed_results.get(key) or []
                            cell = seq[i] if i < len(seq) else ""
                            row.append(_coerce_scalar(cell))
                        subrun_sheet.append(row)
                        global_idx += 1

            heatmap_candidates = []
            for key, values in results.items():
                if isinstance(values, np.ndarray) and values.ndim >= 2:
                    heatmap_candidates.append((key, values))
            for idx, (key, arr) in enumerate(heatmap_candidates, 1):
                sheet_name = _clean_sheet_title(key, f"Heatmap {idx}")
                heatmap_sheet = workbook.create_sheet(title=sheet_name)
                try:
                    data_arr = np.asarray(arr)
                except Exception:
                    data_arr = np.asarray(arr, dtype=object)
                rows, cols = data_arr.shape[:2]
                x_values = list(self.param_values)
                y_values = list(self.param2_values)
                if len(x_values) != cols:
                    x_values = list(range(cols))
                if len(y_values) != rows:
                    y_values = list(range(rows))
                header = [""] + [
                    _scale_value(val, self.param_scale) for val in x_values
                ]
                heatmap_sheet.append(header)
                for j in range(rows):
                    row = [
                        _scale_value(y_values[j], self.param2_scale)
                        if j < len(y_values)
                        else "",
                    ]
                    for i in range(cols):
                        row.append(_coerce_scalar(data_arr[j, i]))
                    heatmap_sheet.append(row)

            summary_sheet = workbook.create_sheet(
                title=_clean_sheet_title("Fit Summary", "Fit Summary")
            )
            summary_headers = [
                "Plot #",
                "Column",
                "Function",
                "Parameter",
                "Value",
                "Error",
                "Unit",
                "R^2",
                "Chi^2",
                "Derived Key",
                "Derived Value",
            ]
            summary_sheet.append(summary_headers)
            for plot_num in sorted(self.plot_data.keys()):
                pdata = self.plot_data.get(plot_num) or {}
                column = pdata.get("column")
                func_name = pdata.get("func_name")
                params = pdata.get("fit_param_details") or []
                r2 = pdata.get("r2")
                chi2 = pdata.get("chi2")
                derived_items = list((pdata.get("derived") or {}).items())
                info_written = False
                for param in params:
                    row = [
                        plot_num if not info_written else "",
                        column if not info_written else "",
                        func_name if not info_written else "",
                        param.get("name") or "",
                        _coerce_scalar(param.get("value")),
                        _coerce_scalar(param.get("error")),
                        param.get("unit") or "",
                        r2 if not info_written else "",
                        chi2 if not info_written else "",
                        "",
                        "",
                    ]
                    summary_sheet.append(row)
                    info_written = True
                for key, value in derived_items:
                    row = [
                        plot_num if not info_written else "",
                        column if not info_written else "",
                        func_name if not info_written else "",
                        "",
                        "",
                        "",
                        "",
                        r2 if not info_written else "",
                        chi2 if not info_written else "",
                        key,
                        _coerce_scalar(value),
                    ]
                    summary_sheet.append(row)
                    info_written = True
                if not info_written:
                    row = [
                        plot_num,
                        column,
                        func_name,
                        "",
                        "",
                        "",
                        "",
                        r2,
                        chi2,
                        "",
                        "",
                    ]
                    summary_sheet.append(row)

        workbook_path = os.path.join(folder, f"{base_name}.xlsx")
        try:
            wb = Workbook()
        except Exception as err:
            print(f"Failed to initialise Excel workbook '{workbook_path}': {err}")
        else:
            try:
                _populate_workbook(wb)
            except Exception as err:
                print(f"Failed to populate Excel workbook '{workbook_path}': {err}")
            else:
                try:
                    wb.save(workbook_path)
                except Exception as err:
                    print(f"Failed to save Excel workbook '{workbook_path}': {err}")

        def _collect_sequence(control, values, scale, *, skip_excluded=False):
            lines = []
            if control is not None:
                try:
                    text = control.GetValue()
                except Exception:
                    text = None
                else:
                    lines = [
                        str(item)
                        for item in self._serialise_sequence_from_text(text)
                    ]
            if not lines and values is not None:
                for idx, value in enumerate(values):
                    if skip_excluded and idx in excluded_union:
                        continue
                    scaled = _scale_value(value, scale)
                    lines.append("" if scaled is None else str(scaled))
            return lines

        gather_settings = getattr(self, "_gather_user_settings", None)
        try:
            fitting_settings = (
                gather_settings(include_start_file=True)
                if callable(gather_settings)
                else {}
            )
        except Exception as err:
            print(f"Failed to collect fitting settings for session bundle: {err}")
            fitting_settings = {}

        parent = getattr(self, "parent", None)
        parent_settings_snapshot = None
        parent_folder = None
        parent_folder_display = None
        if parent is not None:
            parent_folder = getattr(parent, "path", None)
            path_ctrl = getattr(parent, "imageFolderPath", None)
            if path_ctrl is not None:
                try:
                    parent_folder_display = path_ctrl.GetValue()
                except Exception:
                    parent_folder_display = parent_folder
            collector = getattr(parent, "_collect_settings", None)
            if callable(collector):
                try:
                    parent_settings_snapshot = copy.deepcopy(collector())
                except Exception as err:
                    print(
                        "Failed to capture image settings for session bundle: "
                        f"{err}"
                    )

        def _first_image_file():
            results = self.results or {}
            image_series = results.get("image_file")
            series = _prepare_series(image_series, len(self.var_values or [])) or []
            for value in series:
                if value:
                    return str(value)
            for run in self.subrun_results or []:
                run = run or {}
                run_results = run.get("results") or {}
                run_series = _prepare_series(
                    run_results.get("image_file"),
                    len(run.get("var_values") or []),
                )
                for value in run_series or []:
                    if value:
                        return str(value)
            return None

        session_payload = {
            "version": 1,
            "fitting": {
                "settings": fitting_settings,
                "variable_sequence": _collect_sequence(
                    getattr(self, "varListCtrl", None),
                    self.var_values,
                    self.var_scale,
                    skip_excluded=True,
                ),
                "parameter_sequence": _collect_sequence(
                    getattr(self, "paramListCtrl", None),
                    self.param_values,
                    self.param_scale,
                ),
                "parameter2_sequence": _collect_sequence(
                    getattr(self, "param2ListCtrl", None),
                    self.param2_values,
                    self.param2_scale,
                ),
                "excluded_indices": sorted(
                    {
                        int(idx)
                        for idx in excluded_union
                        if isinstance(idx, (int, float, str))
                    }
                ),
            },
        }

        run_exclusions = {}
        for run_idx, indices in (self.run_point_exclusions or {}).items():
            try:
                run_key = int(run_idx)
            except Exception:
                continue
            try:
                values = sorted({int(val) for val in indices})
            except Exception:
                values = sorted({val for val in indices if val is not None})
            if not values:
                continue
            run_exclusions[str(run_key)] = values
        if run_exclusions:
            session_payload["fitting"]["run_exclusions"] = run_exclusions
            self._session_run_exclusions = {
                int(key): set(vals) for key, vals in run_exclusions.items()
            }
        else:
            self._session_run_exclusions = {}

        image_payload = {}
        if parent_folder:
            image_payload["folder"] = parent_folder
        if parent_folder_display and parent_folder_display != parent_folder:
            image_payload["folder_display"] = parent_folder_display
        if parent_settings_snapshot is not None:
            image_payload["settings"] = parent_settings_snapshot
        image_file_hint = _first_image_file()
        if image_file_hint:
            image_payload["image_file"] = image_file_hint
        if image_payload:
            session_payload["image"] = image_payload

        session_path = os.path.join(folder, f"{base_name}.json")
        try:
            with open(session_path, "w", encoding="utf-8") as fh:
                json.dump(session_payload, fh, indent=2, sort_keys=True)
        except Exception as err:
            print(f"Failed to write session JSON '{session_path}': {err}")

        try:
            self._save_user_settings()
        except Exception as err:
            print(
                "Failed to persist fitting window settings before copy: "
                f"{err}"
            )

        if parent is not None:
            save_settings = getattr(parent, "_save_settings", None)
            if callable(save_settings):
                try:
                    save_settings()
                except Exception as err:
                    print(
                        "Failed to persist image UI settings before copy: "
                        f"{err}"
                    )

    def on_close(self, event):
        if self._has_unsaved_changes and self.results:
            dialog = wx.MessageDialog(
                self,
                "You have unsaved results. Would you like to save before exiting?",
                "Unsaved Results",
                style=wx.ICON_WARNING | wx.YES_NO | wx.CANCEL,
            )
            response = dialog.ShowModal()
            dialog.Destroy()
            if response == wx.ID_CANCEL:
                if event.CanVeto():
                    event.Veto()
                return
            if response == wx.ID_YES:
                self.on_save()
                if self._has_unsaved_changes:
                    if event.CanVeto():
                        event.Veto()
                    return
        self._save_user_settings()
        self._restore_parent_context()
        if hasattr(self.parent, "fitWindow"):
            self.parent.fitWindow = None
        if hasattr(self.parent, "fitWindowBtn"):
            button = getattr(self.parent, "fitWindowBtn", None)
            if button is not None:
                try:
                    button.Enable()
                except Exception:
                    pass
        event.Skip()

    def _load_user_settings(self):
        try:
            with open(self._settings_path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
        except FileNotFoundError:
            return {}
        except Exception as err:
            print(f"Failed to load fitting window settings: {err}")
            return {}
        return data if isinstance(data, dict) else {}

    def _apply_user_settings(self, settings):
        if not isinstance(settings, dict):
            return

        variable = settings.get("variable")
        if isinstance(variable, dict):
            label = variable.get("label")
            if isinstance(label, str):
                self.varLabelCtrl.ChangeValue(label)
            unit_type = variable.get("unit_type")
            if isinstance(unit_type, str):
                idx = self.unitTypeChoice.FindString(unit_type)
                if idx != wx.NOT_FOUND:
                    self.unitTypeChoice.SetSelection(idx)
                    self._update_unit_scale_choices()
                    unit_scale = variable.get("unit_scale")
                    if isinstance(unit_scale, str):
                        scale_idx = self.unitScaleChoice.FindString(unit_scale)
                        if scale_idx != wx.NOT_FOUND:
                            self.unitScaleChoice.SetSelection(scale_idx)
                            self.on_unit_scale_change(None)
            offset = variable.get("offset")
            if isinstance(offset, (int, float, str)):
                self.varOffsetCtrl.ChangeValue(str(offset))

        parameter = settings.get("parameter")
        if isinstance(parameter, dict):
            label = parameter.get("label")
            if isinstance(label, str):
                self.paramLabelCtrl.ChangeValue(label)
            unit_type = parameter.get("unit_type")
            if isinstance(unit_type, str):
                idx = self.paramUnitTypeChoice.FindString(unit_type)
                if idx != wx.NOT_FOUND:
                    self.paramUnitTypeChoice.SetSelection(idx)
                    self._update_param_unit_scale_choices()
                    unit_scale = parameter.get("unit_scale")
                    if isinstance(unit_scale, str):
                        scale_idx = self.paramUnitScaleChoice.FindString(unit_scale)
                        if scale_idx != wx.NOT_FOUND:
                            self.paramUnitScaleChoice.SetSelection(scale_idx)
                            self.on_param_unit_scale_change(None)
            offset = parameter.get("offset")
            if isinstance(offset, (int, float, str)):
                self.paramOffsetCtrl.ChangeValue(str(offset))

        parameter2 = settings.get("parameter2")
        if isinstance(parameter2, dict):
            label = parameter2.get("label")
            if isinstance(label, str):
                self.param2LabelCtrl.ChangeValue(label)
            unit_type = parameter2.get("unit_type")
            if isinstance(unit_type, str):
                idx = self.param2UnitTypeChoice.FindString(unit_type)
                if idx != wx.NOT_FOUND:
                    self.param2UnitTypeChoice.SetSelection(idx)
                    self._update_param2_unit_scale_choices()
                    unit_scale = parameter2.get("unit_scale")
                    if isinstance(unit_scale, str):
                        scale_idx = self.param2UnitScaleChoice.FindString(unit_scale)
                        if scale_idx != wx.NOT_FOUND:
                            self.param2UnitScaleChoice.SetSelection(scale_idx)
                            self.on_param2_unit_scale_change(None)
            offset = parameter2.get("offset")
            if isinstance(offset, (int, float, str)):
                self.param2OffsetCtrl.ChangeValue(str(offset))

        if "start_file" in settings:
            start_file = settings.get("start_file")
            ctrl = getattr(self, "startFileCtrl", None)
            if ctrl is not None:
                try:
                    ctrl.ChangeValue("" if start_file is None else str(start_file))
                except Exception:
                    pass

        self._refresh_plot_labels()

    def _gather_user_settings(self, *, include_start_file=False):
        settings = {
            "variable": {
                "label": self.varLabelCtrl.GetValue(),
                "unit_type": self.unitTypeChoice.GetStringSelection(),
                "unit_scale": self.unitScaleChoice.GetStringSelection(),
                "offset": self.varOffsetCtrl.GetValue(),
            },
            "parameter": {
                "label": self.paramLabelCtrl.GetValue(),
                "unit_type": self.paramUnitTypeChoice.GetStringSelection(),
                "unit_scale": self.paramUnitScaleChoice.GetStringSelection(),
                "offset": self.paramOffsetCtrl.GetValue(),
            },
            "parameter2": {
                "label": self.param2LabelCtrl.GetValue(),
                "unit_type": self.param2UnitTypeChoice.GetStringSelection(),
                "unit_scale": self.param2UnitScaleChoice.GetStringSelection(),
                "offset": self.param2OffsetCtrl.GetValue(),
            },
        }
        if include_start_file:
            settings["start_file"] = self.startFileCtrl.GetValue()
        return settings

    def _save_user_settings(self):
        settings = self._gather_user_settings()
        try:
            with open(self._settings_path, "w", encoding="utf-8") as fh:
                json.dump(settings, fh, indent=2, sort_keys=True)
        except Exception as err:
            print(f"Failed to save fitting window settings: {err}")
        else:
            self._user_settings = settings

    def _generate_fit_domain(self, x_values, axis_scale):
        """Return x-values used when plotting fitted curves."""

        try:
            arr = np.asarray(x_values, dtype=float)
        except Exception:
            return np.asarray([], dtype=float)
        if arr.size == 0:
            return arr
        finite = arr[np.isfinite(arr)]
        scale_mode = (axis_scale or SCALE_OPTIONS[0]).lower()
        if scale_mode == "log":
            finite = finite[finite > 0]
        if finite.size == 0:
            return np.asarray([], dtype=float)
        xmin = float(np.min(finite))
        xmax = float(np.max(finite))
        if xmin == xmax:
            return np.full(300, xmin, dtype=float)
        if scale_mode == "log" and xmin > 0 and xmax > 0:
            return np.geomspace(xmin, xmax, 300)
        return np.linspace(xmin, xmax, 300)

    def _plot_dataset(
        self,
        ax,
        data,
        *,
        plot_num=None,
        set_title=False,
        color="black",
        label=None,
        marker=None,
        markerfacecolor=None,
        markeredgecolor=None,
    ):
        def _is_none_color(value):
            return isinstance(value, str) and value.lower() == "none"

        hollow_marker = markerfacecolor is not None and _is_none_color(markerfacecolor)
        marker_kwargs = {}
        if markerfacecolor is not None:
            marker_kwargs["markerfacecolor"] = markerfacecolor
            if hollow_marker:
                marker_kwargs["fillstyle"] = "none"
        if markeredgecolor is not None:
            marker_kwargs["markeredgecolor"] = markeredgecolor

        self._deregister_artists(data.get("artists"))
        x_all, y_all = data["x"], data["y"]
        y_err_all = data.get("y_err")
        x_scale = data.get("x_scale", self.var_scale)
        if y_err_all is not None:
            y_err_all = np.asarray(y_err_all)
        excluded = data.get("excluded", set())
        mask = np.ones(len(x_all), dtype=bool)
        if excluded:
            groups = data.get("orig_indices")
            if groups:
                exc = set(excluded)
                for i, grp in enumerate(groups):
                    if exc.intersection(np.atleast_1d(grp)):
                        mask[i] = False
            else:
                mask[list(excluded)] = False
        x = x_all[mask]
        y = y_all[mask]
        y_err = y_err_all[mask] if y_err_all is not None else None
        display_idx = np.arange(len(x_all))[mask]
        axis_scale_mode = data.get("axis_scale_mode") or self._normalise_axis_scale(
            data.get("axis_scale")
        )
        finite_mask = np.isfinite(x) & np.isfinite(y)
        if y_err is not None:
            finite_mask &= np.isfinite(y_err)
        if axis_scale_mode in ("semi-log", "log"):
            finite_mask &= y > 0
        if axis_scale_mode == "log":
            finite_mask &= x > 0
        if not np.all(finite_mask):
            x = x[finite_mask]
            y = y[finite_mask]
            if y_err is not None:
                y_err = y_err[finite_mask]
            display_idx = display_idx[finite_mask]
        func_name = data["func_name"]
        data["residual_x"] = None
        data["residual_y"] = None
        data["residual_axis_label"] = None
        data["residual_series_label"] = None
        start_subtitle = None
        if set_title:
            start_file_name = self.startFileCtrl.GetValue().strip()
            if start_file_name:
                start_subtitle = os.path.basename(start_file_name)
        if not func_name:
            y_unit, y_factor = self.col_scales.get(data["column"], ("", 1.0))
            display_factor = y_factor if y_factor is not None else 1.0
            data["y_unit"] = y_unit
            plot_x = x / x_scale
            plot_y = y * display_factor
            plot_err = y_err * display_factor if y_err is not None else None
            marker_style = marker or "o"
            data["plot_err"] = np.asarray(plot_err) if plot_err is not None else None
            if plot_err is not None:
                cont = ax.errorbar(
                    plot_x,
                    plot_y,
                    yerr=plot_err,
                    fmt=marker_style,
                    color=color,
                    ecolor=color,
                    capsize=3,
                    picker=True,
                    label=label,
                )
                line = cont.lines[0]
                if markerfacecolor is not None:
                    line.set_markerfacecolor(markerfacecolor)
                if markeredgecolor is not None:
                    line.set_markeredgecolor(markeredgecolor)
                if hollow_marker:
                    line.set_fillstyle("none")
                caplines = getattr(cont, "caplines", [])
                barcols = getattr(cont, "barlinecols", [])
                if isinstance(barcols, (list, tuple)):
                    barlist = list(barcols)
                elif barcols:
                    barlist = [barcols]
                else:
                    barlist = []
                artists = [line] + list(caplines) + barlist
            else:
                (line,) = ax.plot(
                    plot_x,
                    plot_y,
                    marker_style,
                    color=color,
                    picker=True,
                    label=label,
                    **marker_kwargs,
                )
                if hollow_marker:
                    line.set_fillstyle("none")
                artists = [line]
            self._make_artists_pickable(artists)
            self._register_artists(plot_num, artists)
            if set_title:
                title_escaped = data["column"].replace(" ", r"\ ")
                if start_subtitle:
                    ax.set_title(
                        f"$\\bf{{{title_escaped}}}$\n{start_subtitle}",
                        fontsize=self._get_font_size("title"),
                    )
                else:
                    ax.set_title(
                        f"$\\bf{{{title_escaped}}}$",
                        fontsize=self._get_font_size("title"),
                    )
            x_lbl = data.get("x_label", self.varLabelCtrl.GetValue())
            x_unit = data.get("x_unit", self.unit_scale)
            y_lbl = data.get("y_label") or data["column"]
            data["x_label"] = x_lbl
            data["x_unit_display"] = x_unit
            data["y_label"] = y_lbl
            data["y_unit"] = y_unit
            data["x_unit_family"] = self._resolve_unit_family(
                data.get("x_unit"), data.get("x_unit_family")
            )
            data["y_unit_family"] = self._resolve_unit_family(y_unit)
            x_axis = f"{x_lbl} ({x_unit})" if x_unit else x_lbl
            y_axis = f"{y_lbl} ({y_unit})" if y_unit else y_lbl
            if x_axis:
                ax.set_xlabel(x_axis, fontsize=self._get_font_size("label"))
            else:
                ax.set_xlabel("")
            if y_axis:
                ax.set_ylabel(y_axis, fontsize=self._get_font_size("label"))
            else:
                ax.set_ylabel("")
            data["popt"] = None
            data["perr"] = None
            data["r2"] = None
            data["derived"] = {}
            data["artist"] = line
            data["artists"] = artists
            data["display_idx"] = display_idx
            data["plot_x"] = np.asarray(plot_x)
            data["plot_y"] = np.asarray(plot_y)
            data["plot_x_fit"] = None
            data["plot_y_fit"] = None
            data["y_display_factor"] = display_factor
            data["fit_param_details"] = []
            data["fit_title"] = None
            data["fit_formula"] = None
            return data["column"]
        fit_info = FIT_FUNCTIONS[func_name]
        y_unit, y_factor = self.col_scales.get(
            data["column"], (fit_info.get("y_unit"), 1.0)
        )
        display_factor = y_factor if y_factor is not None else 1.0
        data["y_unit"] = y_unit
        data["y_unit_family"] = self._resolve_unit_family(y_unit)
        popt = None
        pcov = None
        r2 = None
        chi2 = None
        perr = None
        if len(x) == 0:
            data["plot_x"] = np.array([])
            data["plot_y"] = np.array([])
            data["plot_x_fit"] = None
            data["plot_y_fit"] = None
            data["y_display_factor"] = display_factor
            data["fit_param_details"] = []
            data["fit_title"] = fit_info.get("title")
            data["fit_formula"] = fit_info.get("formula")
            return None
        x_fit = self._generate_fit_domain(x, axis_scale_mode)
        try:
            if func_name in TEMP_PSD_FIT_KEYS:
                t_sq = np.square(x)
                sig_sq = np.square(y)
                if y_err is not None:
                    sig_sq_err = 2 * y * y_err
                    w = np.where(sig_sq_err > 0, 1 / sig_sq_err, 1)
                    (m, b), pcov = np.polyfit(t_sq, sig_sq, 1, w=w, cov=True)
                else:
                    sig_sq_err = None
                    (m, b), pcov = np.polyfit(t_sq, sig_sq, 1, cov=True)
                popt = (m, b)
                y_fit = m * t_sq + b
                x_fit_sq = np.square(x_fit)
                y_fit_plot = m * x_fit_sq + b
                plot_x = t_sq / (x_scale ** 2)
                display_factor = 1e12
                plot_y = sig_sq * display_factor
                plot_x_fit = x_fit_sq / (x_scale ** 2)
                plot_y_fit = y_fit_plot * display_factor
                plot_err = (
                    sig_sq_err * display_factor if sig_sq_err is not None else None
                )
                sigma_for_chi2 = sig_sq_err
                target = sig_sq
                y_unit = fit_info.get("y_unit")
            elif func_name == "Linear":
                x_transformed = np.asarray(x, dtype=float)
                y_transformed = np.asarray(y, dtype=float)
                y_err_transformed = None
                x_fit_transformed = np.asarray(x_fit, dtype=float)
                if axis_scale_mode in ("semi-log", "log"):
                    with np.errstate(divide="ignore", invalid="ignore"):
                        y_transformed = np.log10(y_transformed)
                    if y_err is not None:
                        with np.errstate(divide="ignore", invalid="ignore"):
                            y_err_transformed = y_err / (np.log(10.0) * y)
                        y_err_transformed = np.asarray(y_err_transformed, dtype=float)
                        y_err_transformed[~np.isfinite(y_err_transformed)] = 0.0
                    if axis_scale_mode == "log":
                        with np.errstate(divide="ignore", invalid="ignore"):
                            x_transformed = np.log10(x_transformed)
                        if x_fit_transformed.size:
                            with np.errstate(divide="ignore", invalid="ignore"):
                                x_fit_transformed = np.log10(x_fit_transformed)
                else:
                    y_err_transformed = (
                        np.asarray(y_err, dtype=float) if y_err is not None else None
                    )
                weights = None
                if y_err_transformed is not None:
                    weights = np.where(
                        (y_err_transformed > 0) & np.isfinite(y_err_transformed),
                        1.0 / y_err_transformed,
                        1.0,
                    )
                    (m, b), pcov = np.polyfit(
                        x_transformed, y_transformed, 1, w=weights, cov=True
                    )
                else:
                    (m, b), pcov = np.polyfit(x_transformed, y_transformed, 1, cov=True)
                popt = (m, b)
                y_fit_transformed = m * x_transformed + b
                if x_fit_transformed.size:
                    y_fit_plot_transformed = m * x_fit_transformed + b
                else:
                    y_fit_plot_transformed = np.asarray([], dtype=float)
                plot_x = x / x_scale
                plot_y = y * display_factor
                plot_x_fit = x_fit / x_scale if x_fit.size else None
                plot_err = y_err * display_factor if y_err is not None else None
                if axis_scale_mode in ("semi-log", "log"):
                    y_fit = y_fit_transformed
                    target = y_transformed
                    sigma_for_chi2 = y_err_transformed
                    plot_y_fit_values = np.power(10.0, y_fit_plot_transformed)
                else:
                    y_fit = y_fit_transformed
                    target = y_transformed
                    sigma_for_chi2 = y_err_transformed
                    plot_y_fit_values = y_fit_plot_transformed
                if plot_y_fit_values.size:
                    plot_y_fit = plot_y_fit_values * display_factor
                else:
                    plot_y_fit = None
            else:
                p0 = data.get("p0", fit_info["p0"])
                if y_err is not None:
                    sigma = np.copy(y_err)
                    sigma[sigma <= 0] = 1.0
                    popt, pcov = curve_fit(
                        fit_info["func"], x, y, p0=p0, sigma=sigma, absolute_sigma=True
                    )
                else:
                    popt, pcov = curve_fit(fit_info["func"], x, y, p0=p0)
                y_fit = fit_info["func"](x, *popt)
                y_fit_plot = fit_info["func"](x_fit, *popt)
                plot_x, plot_y = x / x_scale, y * display_factor
                plot_x_fit, plot_y_fit = x_fit / x_scale, y_fit_plot * display_factor
                plot_err = (
                    y_err * display_factor if y_err is not None else None
                )
                sigma_for_chi2 = y_err
                target = y
            ss_res = np.sum((y_fit - target) ** 2)
            ss_tot = np.sum((target - np.mean(target)) ** 2)
            r2 = 1 - ss_res / ss_tot if ss_tot != 0 else 0
            if sigma_for_chi2 is not None:
                sigma = np.copy(sigma_for_chi2)
                sigma[sigma <= 0] = 1.0
                resid = (y_fit - target) / sigma
            else:
                resid = y_fit - target
            dof = len(resid) - len(popt) if popt is not None else len(resid)
            chi2_val = np.sum(resid ** 2)
            if dof > 0:
                chi2_val /= dof
            if self.paramBinCheck.GetValue():
                chi2 = chi2_val
        except Exception:
            plot_x, plot_y = x / x_scale, y * display_factor
            plot_x_fit = plot_y_fit = None
            plot_err = y_err * display_factor if y_err is not None else None
        perr = np.sqrt(np.diag(pcov)) if pcov is not None else None
        data["y_display_factor"] = display_factor
        data["plot_err"] = np.asarray(plot_err) if plot_err is not None else None
        fit_params = (
            self._scale_fit_parameters(func_name, popt, perr, data)
            if popt is not None
            else []
        )
        data["fit_param_details"] = fit_params
        data["fit_title"] = fit_info.get("title")
        data["fit_formula"] = fit_info.get("formula")
        container = None
        marker_style = marker or "o"
        if plot_err is not None:
            cont = ax.errorbar(
                plot_x,
                plot_y,
                yerr=plot_err,
                fmt=marker_style,
                color=color,
                ecolor=color,
                capsize=3,
                picker=True,
                label=label,
            )
            container = cont
            line = cont.lines[0]
            if markerfacecolor is not None:
                line.set_markerfacecolor(markerfacecolor)
            if markeredgecolor is not None:
                line.set_markeredgecolor(markeredgecolor)
            if hollow_marker:
                line.set_fillstyle("none")
            caplines = getattr(cont, "caplines", [])
            barcols = getattr(cont, "barlinecols", [])
            if isinstance(barcols, (list, tuple)):
                barlist = list(barcols)
            elif barcols:
                barlist = [barcols]
            else:
                barlist = []
            artists = [line] + list(caplines) + barlist
        else:
            (line,) = ax.plot(
                plot_x,
                plot_y,
                marker_style,
                color=color,
                picker=True,
                label=label,
                **marker_kwargs,
            )
            if hollow_marker:
                line.set_fillstyle("none")
            artists = [line]
        self._make_artists_pickable(artists)
        self._register_artists(plot_num, artists)
        label_text = fit_info["formula"]
        title_main = fit_info["title"]
        subtitles = []
        if set_title and start_subtitle:
            subtitles.append(start_subtitle)
        if fit_params:
            display_popt = [item["value"] for item in fit_params]
        elif popt is not None:
            display_popt = list(popt)
        else:
            display_popt = []
        if popt is not None and plot_x_fit is not None:
            ax.plot(plot_x_fit, plot_y_fit, color=color)
            label_text += f"\nR^2={r2:.2f}"
            if self.paramBinCheck.GetValue() and chi2 is not None:
                label_text += f"\n\u03C7^2/\u03BD={chi2:.2f}"
            if fit_params:
                param_info = ", ".join(
                    f"{item['name']}={item['value']:.3g}" for item in fit_params
                )
            else:
                param_info = ", ".join(
                    f"{n}={v:.3g}" for n, v in zip(fit_info["param_names"], display_popt)
                )
        else:
            label_text += "\nFit failed"
            param_info = None

        derived_vals = {}
        derive_fn = fit_info.get("derived")
        if popt is not None and derive_fn:
            try:
                base_var_vals = np.asarray(self.var_values)
            except Exception:
                base_var_vals = np.asarray(self.var_values, dtype=float)
            base_len = base_var_vals.shape[0] if base_var_vals.ndim == 1 else 0
            source_results = self.results if isinstance(self.results, dict) else {}
            base_results = {}
            for key, values in source_results.items():
                try:
                    arr = np.asarray(values)
                except Exception:
                    arr = None
                if arr is not None and arr.shape and arr.shape[0] == base_len:
                    base_results[key] = arr
                elif isinstance(values, (list, tuple)) and len(values) == base_len:
                    base_results[key] = np.asarray(values)
                else:
                    base_results[key] = values

            analysis_var = base_var_vals
            analysis_results = base_results
            groups = data.get("orig_indices") if self.paramBinCheck.GetValue() else None
            aggregated_mask = None
            if groups:
                aggregated = self._build_param_bin_analysis_context(
                    groups, base_var_vals, base_results
                )
                if aggregated:
                    agg_var, agg_results = aggregated
                    if np.asarray(agg_var).size:
                        analysis_var = np.asarray(agg_var)
                        analysis_results = dict(agg_results)
                        aggregated_mask = np.asarray(mask, dtype=bool)

            mask_to_apply = None
            if aggregated_mask is not None and aggregated_mask.size == analysis_var.shape[0]:
                mask_to_apply = aggregated_mask
            elif excluded and analysis_var.ndim == 1:
                mask_to_apply = np.ones(analysis_var.shape[0], dtype=bool)
                for idx in excluded:
                    try:
                        pos = int(idx)
                    except Exception:
                        continue
                    if 0 <= pos < mask_to_apply.size:
                        mask_to_apply[pos] = False

            if mask_to_apply is not None and analysis_var.ndim == 1:
                analysis_var = analysis_var[mask_to_apply]
                filtered_results = {}
                mask_len = mask_to_apply.size
                for key, values in analysis_results.items():
                    if isinstance(values, np.ndarray):
                        if values.ndim >= 1 and values.shape[0] == mask_len:
                            filtered_results[key] = values[mask_to_apply]
                        else:
                            filtered_results[key] = values
                    elif isinstance(values, (list, tuple)) and len(values) == mask_len:
                        filtered_results[key] = [
                            values[i] for i in range(mask_len) if mask_to_apply[i]
                        ]
                    else:
                        filtered_results[key] = values
                analysis_results = filtered_results

            derived_vals = derive_fn(
                popt, data, analysis_var, analysis_results, self.parent, pcov
            )
        if derived_vals:
            unit = data.get("x_unit", self.unit_scale)
            factor = data.get("x_scale", self.var_scale)
            if func_name == MOT_LIFETIME_DISPLAY_NAME:
                lifetime_key = "Lifetime (s)"
                lifetime_err_key = "Lifetime (s)_err"
                if lifetime_key in derived_vals:
                    derived_vals[f"Lifetime ({unit})"] = derived_vals.pop(lifetime_key) / factor
                if lifetime_err_key in derived_vals:
                    derived_vals[f"Lifetime ({unit})_err"] = derived_vals.pop(lifetime_err_key) / factor
                if "Decay rate (1/s)" in derived_vals:
                    derived_vals[f"Decay rate (1/{unit})"] = derived_vals.pop("Decay rate (1/s)") * factor
            elif func_name in ("Gaussian", "Lorentzian") and "FWHM" in derived_vals:
                derived_vals[f"FWHM ({unit})"] = derived_vals.pop("FWHM") / factor
            elif func_name == MOT_RINGDOWN_DISPLAY_NAME:
                if "Damping time (s)" in derived_vals:
                    derived_vals[f"Damping time ({unit})"] = derived_vals.pop("Damping time (s)") / factor
                if "Frequency (Hz)" in derived_vals:
                    derived_vals[f"Frequency (1/{unit})"] = derived_vals.pop("Frequency (Hz)") * factor
        derived_info = None
        if popt is not None and derived_vals:
            derived_lines = []
            if func_name in TEMP_PSD_FIT_KEYS:
                temp_x = derived_vals.get("Temp_x (µK)")
                temp_x_err = derived_vals.get("Temp_x_err (µK)")
                temp_y = derived_vals.get("Temp_y (µK)")
                temp_y_err = derived_vals.get("Temp_y_err (µK)")
                psd = derived_vals.get("PSD")
                psd_err = derived_vals.get("PSD_err")
                density = derived_vals.get("Density (cm^-3)")
                density_err = derived_vals.get("Density (cm^-3)_err")
                if None not in (temp_x, temp_x_err, temp_y, temp_y_err):
                    derived_lines.append(
                        f"x-Temp = {temp_x:.3g} +/- {temp_x_err:.3g} (uK), "
                        f"y-Temp = {temp_y:.3g} +/- {temp_y_err:.3g} (uK)"
                    )
                if psd is not None and psd_err is not None:
                    derived_lines.append(f"PSD = {psd:.3g} +/- {psd_err:.3g}")
                if density is not None and density_err is not None:
                    derived_lines.append(
                        f"Density = {density:.3g} +/- {density_err:.3g} cm^-3"
                    )
            else:
                formatted = []
                for k, v in derived_vals.items():
                    if self._is_error_key(k):
                        continue
                    err_key = k + "_err"
                    if err_key in derived_vals:
                        formatted.append(
                            f"{k}={derived_vals[k]:.3g}\u00B1{derived_vals[err_key]:.3g}"
                        )
                    else:
                        formatted.append(f"{k}={v:.3g}")
                if formatted:
                    derived_lines.append(
                        ", ".join(formatted)
                    )
            if derived_lines:
                derived_info = "\n".join(derived_lines)
        if not set_title:
            extra_lines = []
            if param_info:
                extra_lines.append(param_info)
            if derived_info:
                extra_lines.append(derived_info)
            if extra_lines:
                label_text += "\n" + "\n".join(extra_lines)
        if set_title:
            subtitle = "\n".join(subtitles)
            title_escaped = title_main.replace(" ", r"\ ")
            if subtitle:
                ax.set_title(
                    f"$\\bf{{{title_escaped}}}$\n{subtitle}",
                    fontsize=self._get_font_size("title"),
                )
            else:
                ax.set_title(
                    f"$\\bf{{{title_escaped}}}$",
                    fontsize=self._get_font_size("title"),
                )

        def fmt_axis(lbl, unit):
            return f"{lbl} ({unit})" if lbl and unit else lbl or ""

        x_lbl = fit_info.get("x_label") or data.get("x_label", self.varLabelCtrl.GetValue())
        y_lbl = fit_info.get("y_label") or data.get("y_label") or data["column"]
        x_unit_val = (
            data.get("x_unit", self.unit_scale) or fit_info.get("x_unit") or ""
        )
        x_unit_disp = (
            f"{x_unit_val}^2" if func_name in TEMP_PSD_FIT_KEYS and x_unit_val else x_unit_val
        )
        data["x_label"] = x_lbl
        data["y_label"] = y_lbl
        data["x_unit"] = x_unit_val
        data["x_unit_display"] = x_unit_disp
        data["y_unit"] = y_unit
        data["x_unit_family"] = self._resolve_unit_family(
            x_unit_val, data.get("x_unit_family")
        )
        data["y_unit_family"] = self._resolve_unit_family(y_unit)
        x_axis_label = fmt_axis(x_lbl, x_unit_disp)
        y_axis_label = fmt_axis(y_lbl, y_unit)
        if x_axis_label:
            ax.set_xlabel(x_axis_label, fontsize=self._get_font_size("label"))
        else:
            ax.set_xlabel("")
        if y_axis_label:
            ax.set_ylabel(y_axis_label, fontsize=self._get_font_size("label"))
        else:
            ax.set_ylabel("")
        if set_title and label_text:
            xpos, ypos, ha, va = self._choose_text_position(plot_x, plot_y)
            ax.text(
                xpos,
                ypos,
                label_text,
                transform=ax.transAxes,
                va=va,
                ha=ha,
            )
        data["popt"] = popt
        data["perr"] = perr
        data["r2"] = r2
        data["chi2"] = chi2
        data["derived"] = derived_vals
        data["artist"] = line
        legend_label = label
        hide_from_legend = isinstance(label, str) and "residual" in label.lower()
        if hide_from_legend:
            legend_label = "_nolegend_"
        elif label is not None:
            legend_label = f"{label}: {label_text}" if label_text else label
        if container is not None:
            container.set_label(legend_label)
        else:
            line.set_label(legend_label)
        data["artists"] = artists
        data["display_idx"] = display_idx
        data["plot_x"] = np.asarray(plot_x)
        data["plot_y"] = np.asarray(plot_y)
        data["plot_x_fit"] = (
            np.asarray(plot_x_fit) if plot_x_fit is not None else None
        )
        data["plot_y_fit"] = (
            np.asarray(plot_y_fit) if plot_y_fit is not None else None
        )
        if popt is not None and y_fit is not None:
            try:
                plot_x_display = np.asarray(plot_x, dtype=float)
                if func_name in TEMP_PSD_FIT_KEYS:
                    y_display = np.asarray(plot_y, dtype=float)
                else:
                    y_display = np.asarray(y, dtype=float) * display_factor
                if axis_scale_mode in ("semi-log", "log") and (
                    func_name == "Linear"
                ):
                    fit_display = (
                        np.power(10.0, np.asarray(y_fit, dtype=float)) * display_factor
                    )
                else:
                    fit_display = np.asarray(y_fit, dtype=float) * display_factor
                residual_values = y_display - fit_display
                valid_mask = np.isfinite(plot_x_display) & np.isfinite(residual_values)
                if np.any(valid_mask):
                    data["residual_x"] = plot_x_display[valid_mask]
                    data["residual_y"] = residual_values[valid_mask]
                column_name = data.get("column") or "Data"
                y_unit_label = data.get("y_unit") or ""
                if y_unit_label:
                    axis_label = f"Residual ({y_unit_label})"
                else:
                    axis_label = "Residual"
                data["residual_axis_label"] = axis_label
                # Residuals should not create legend entries in the plot.
                data["residual_series_label"] = None
            except Exception:
                pass
        return legend_label

    def _make_artists_pickable(self, artists):
        """Ensure that every plotted artist responds to pick events."""

        tolerance = getattr(self, "_pick_tolerance", DEFAULT_PICK_TOLERANCE)
        for artist in artists or ():
            if artist is None:
                continue
            try:
                if hasattr(artist, "set_pickradius"):
                    artist.set_pickradius(tolerance)
            except Exception:
                pass
            try:
                if hasattr(artist, "set_picker"):
                    artist.set_picker(tolerance)
            except Exception:
                pass

    def _register_artists(self, plot_num, artists):
        if plot_num is None or not artists:
            return
        mapping = getattr(self, "_artist_plot_map", None)
        if mapping is None:
            self._artist_plot_map = mapping = {}
        for artist in artists:
            if artist is not None:
                mapping[artist] = plot_num

    def _deregister_artists(self, artists):
        mapping = getattr(self, "_artist_plot_map", None)
        if not mapping or not artists:
            return
        for artist in artists:
            mapping.pop(artist, None)

    def _apply_axis_scale(self, ax, scale_type, plot_x=None, plot_y=None):
        scale = (scale_type or SCALE_OPTIONS[0]).lower()

        def has_positive(values):
            if values is None:
                return False
            arr = np.asarray(values, dtype=float)
            if arr.size == 0:
                return False
            arr = arr[np.isfinite(arr)]
            if arr.size == 0:
                return False
            return np.any(arr > 0)

        if scale == "semi-log":
            ax.set_xscale("linear")
            ax.set_yscale("log" if has_positive(plot_y) else "linear")
        elif scale == "log":
            ax.set_xscale("log" if has_positive(plot_x) else "linear")
            ax.set_yscale("log" if has_positive(plot_y) else "linear")
        else:
            ax.set_xscale("linear")
            ax.set_yscale("linear")

    def _combine_plot_arrays(self, *arrays):
        """Return a concatenated array of finite values from inputs."""

        chunks = []
        for arr in arrays:
            if arr is None:
                continue
            try:
                data = np.asarray(arr, dtype=float)
            except Exception:
                continue
            if data.size == 0:
                continue
            data = data[np.isfinite(data)]
            if data.size:
                chunks.append(data)
        if not chunks:
            return None
        return np.concatenate(chunks)

    def _align_overlay_ticks(self, ax_left, ax_right):
        if ax_left is None or ax_right is None:
            return
        try:
            ax_right.set_ylim(ax_left.get_ylim())
        except Exception:
            pass
        try:
            major = ax_left.get_yticks()
            ax_right.set_yticks(major)
        except Exception:
            pass
        try:
            minor = ax_left.yaxis.get_minorticklocs()
        except Exception:
            minor = []
        if minor is not None and len(minor) > 0:
            try:
                ax_right.yaxis.set_minor_locator(mticker.FixedLocator(minor))
            except Exception:
                pass
        else:
            try:
                ax_right.yaxis.set_minor_locator(mticker.NullLocator())
            except Exception:
                pass

    def _dimension_scale_and_unit(self, descriptor, data):
        """Return scaling factor and unit string for a parameter descriptor."""

        if not descriptor:
            return 1.0, ""
        descriptor = descriptor.lower()
        x_scale = data.get("x_scale") or self.var_scale or 1.0
        y_factor = data.get("y_display_factor")
        if y_factor is None:
            y_factor = 1.0

        def _clean_unit(unit):
            if unit is None:
                return ""
            if isinstance(unit, str):
                return unit.strip()
            if isinstance(unit, (list, tuple)):
                parts = []
                for part in unit:
                    cleaned = _clean_unit(part)
                    if cleaned:
                        parts.append(cleaned)
                if not parts:
                    return ""
                return "·".join(parts)
            if isinstance(unit, set):
                cleaned_parts = []
                for part in unit:
                    cleaned = _clean_unit(part)
                    if cleaned:
                        cleaned_parts.append(cleaned)
                parts = sorted(cleaned_parts)
                if not parts:
                    return ""
                if len(parts) == 1:
                    return parts[0]
                return "·".join(parts)
            try:
                text = str(unit).strip()
            except Exception:
                text = ""
            return text

        x_unit = _clean_unit(data.get("x_unit_display") or data.get("x_unit") or "")
        y_unit = _clean_unit(data.get("y_unit") or "")
        func_name = data.get("func_name")
        axis_scale_mode = data.get("axis_scale_mode") or self._normalise_axis_scale(
            data.get("axis_scale")
        )
        log_linear_fit = func_name == "Linear" and axis_scale_mode in ("semi-log", "log")

        def _log_unit_text(unit_value, fallback_label):
            base = _clean_unit(unit_value)
            if base:
                return f"log10({base})"
            fallback = fallback_label or "value"
            fallback_clean = _clean_unit(fallback) if fallback else "value"
            if not fallback_clean:
                fallback_clean = "value"
            return f"log10({fallback_clean})"

        def ratio_unit(numerator, denominator, power=1):
            numerator = _clean_unit(numerator)
            denominator = _clean_unit(denominator)
            if isinstance(numerator, (set, list, tuple)):
                numerator = _clean_unit(list(numerator))
            if isinstance(denominator, (set, list, tuple)):
                denominator = _clean_unit(list(denominator))
            if not isinstance(numerator, str):
                try:
                    numerator = str(numerator)
                except Exception:
                    numerator = ""
            if not isinstance(denominator, str):
                try:
                    denominator = str(denominator)
                except Exception:
                    denominator = ""

            numerator = numerator.strip()
            denominator = denominator.strip()

            def _format_number(value):
                if isinstance(value, (int, float)):
                    if isinstance(value, float) and not value.is_integer():
                        return f"{value:.3g}"
                    return str(int(value))
                cleaned = _clean_unit(value)
                if isinstance(cleaned, str):
                    return cleaned.strip()
                try:
                    return str(cleaned).strip()
                except Exception:
                    return ""

            if isinstance(power, (int, float)):
                pos_power = _format_number(power)
                neg_power = _format_number(-power)
            else:
                pos_power = _format_number(power)
                neg_power = f"-{pos_power}" if pos_power else ""

            if denominator:
                if numerator:
                    if isinstance(power, (int, float)) and power == 1:
                        return f"{numerator}/{denominator}"
                    suffix = f"^{pos_power}" if pos_power and pos_power != "1" else ""
                    return f"{numerator}/{denominator}{suffix}"
                suffix = f"^{neg_power}" if neg_power else ""
                return f"{denominator}{suffix}"
            return numerator

        if descriptor == "y":
            if log_linear_fit:
                unit_text = _log_unit_text(y_unit, data.get("y_label"))
                return 1.0, unit_text
            return y_factor, y_unit
        if descriptor == "x":
            return (1.0 / x_scale) if x_scale else 1.0, x_unit
        if descriptor == "x^2":
            factor = (1.0 / (x_scale ** 2)) if x_scale else 1.0
            unit = f"{x_unit}^2" if x_unit else ""
            return factor, unit
        if descriptor == "1/x":
            unit = ratio_unit("", x_unit)
            return x_scale if x_scale else 1.0, unit
        if descriptor == "1/x^2":
            unit = ratio_unit("", x_unit, power=2)
            factor = (x_scale ** 2) if x_scale else 1.0
            return factor, unit
        if descriptor == "y/x":
            if log_linear_fit:
                if axis_scale_mode == "log":
                    numerator = _log_unit_text(y_unit, data.get("y_label") or "y")
                    denominator = _log_unit_text(
                        x_unit, data.get("x_label") or "x"
                    )
                    return 1.0, f"{numerator}/{denominator}"
                numerator = _log_unit_text(y_unit, data.get("y_label"))
                if x_unit:
                    unit = f"{numerator}/{x_unit}"
                else:
                    unit = numerator
                factor = (1.0 / x_scale) if x_scale else 1.0
                return factor, unit
            unit = ratio_unit(y_unit, x_unit)
            factor = (y_factor * x_scale) if x_scale else y_factor
            return factor, unit
        if descriptor == "y/x^2":
            unit = ratio_unit(y_unit, x_unit, power=2)
            factor = (y_factor * (x_scale ** 2)) if x_scale else y_factor
            return factor, unit
        if descriptor == "y*x":
            unit = f"{y_unit}·{x_unit}" if y_unit and x_unit else y_unit or x_unit
            factor = y_factor / x_scale if x_scale else y_factor
            return factor, unit
        if descriptor == "rad":
            return 1.0, "rad"
        if descriptor in ("dimensionless", "unitless"):
            return 1.0, ""
        return 1.0, ""

    def _scale_fit_parameters(self, func_name, popt, perr, data):
        """Return scaled parameter dictionaries with values, errors, and units."""

        if popt is None:
            return []
        info = FIT_FUNCTIONS.get(func_name, {})
        names = info.get("param_names", [])
        descriptors = FIT_PARAMETER_DIMENSIONS.get(func_name, [])
        results = []
        for idx, name in enumerate(names):
            descriptor = descriptors[idx] if idx < len(descriptors) else None
            scale, unit = self._dimension_scale_and_unit(descriptor, data)
            value = popt[idx]
            error = None
            if perr is not None and idx < len(perr):
                error = perr[idx]
            display_value = value * scale
            display_error = error * abs(scale) if (error is not None) else None
            results.append(
                {
                    "name": name,
                    "value": display_value,
                    "error": display_error,
                    "unit": unit,
                    "descriptor": descriptor,
                }
            )
        return results

    def _build_residual_series(self, data, color=None):
        if not data:
            return None
        x = data.get("residual_x")
        y = data.get("residual_y")
        if x is None or y is None:
            return None
        try:
            x_arr = np.asarray(x, dtype=float)
            y_arr = np.asarray(y, dtype=float)
        except Exception:
            return None
        if x_arr.size == 0 or y_arr.size == 0:
            return None
        series = {
            "x": x_arr,
            "y": y_arr,
            "axis_label": data.get("residual_axis_label") or "Residuals",
            "label": data.get("residual_series_label"),
        }
        if color:
            series["color"] = color
        return series

    def _render_residual_axis(self, axis, series_list, x_label):
        if axis is None:
            return
        axis.clear()
        self._configure_plot_axis(axis, force_square=False)
        axis.tick_params(labeltop=False)
        valid_series = []
        for series in series_list or []:
            if not series:
                continue
            x_vals = series.get("x")
            y_vals = series.get("y")
            if x_vals is None or y_vals is None:
                continue
            try:
                x_arr = np.asarray(x_vals, dtype=float)
                y_arr = np.asarray(y_vals, dtype=float)
            except Exception:
                continue
            if x_arr.size == 0 or y_arr.size == 0:
                continue
            valid_series.append((x_arr, y_arr, series.get("color"), series.get("label")))
        axis.axhline(0, color="0.5", linestyle="--", linewidth=1)
        if valid_series:
            for idx, (x_arr, y_arr, color, label) in enumerate(valid_series):
                plot_color = color or f"C{idx}"
                axis.plot(x_arr, y_arr, "o", color=plot_color, markersize=4, label=label)
            if any(label for *_, label in valid_series):
                axis.legend(loc="upper right", fontsize=self._get_font_size("tick"))
            axis.relim()
            axis.autoscale_view()
            axis.margins(y=0.2)
        else:
            axis.text(
                0.5,
                0.5,
                "Residuals unavailable",
                transform=axis.transAxes,
                ha="center",
                va="center",
                fontsize=self._get_font_size("tick"),
            )
        labels = [
            (series or {}).get("axis_label")
            for series in (series_list or [])
            if (series or {}).get("axis_label")
        ]
        if labels:
            axis_label = labels[0] if len(set(labels)) == 1 else "Residuals"
        else:
            axis_label = "Residuals"
        axis.set_ylabel(axis_label, fontsize=self._get_font_size("label"))
        if x_label:
            axis.set_xlabel(x_label, fontsize=self._get_font_size("label"))
        else:
            axis.set_xlabel("")
        axis.tick_params(labelbottom=True)
        if not valid_series:
            axis.set_ylim(-1, 1)

    def _is_error_key(self, name):
        if not isinstance(name, str):
            return False
        stripped = name.strip()
        if "(" in stripped and stripped.endswith(")"):
            stripped = stripped.split("(", 1)[0].strip()
        lowered = stripped.lower()
        suffixes = ("_err", " err", "_std", " std", "_stderr", " stderr", "_error", " error")
        return any(lowered.endswith(suffix.strip()) for suffix in suffixes)

    def _find_matching_error_key(self, name, mapping):
        suffixes = ["_err", "_Err", " err", " Err", "_std", "_Std", " std", " Std"]
        for suffix in suffixes:
            candidate = f"{name}{suffix}"
            if candidate in mapping:
                return candidate
        if "(" in name:
            base, rest = name.split("(", 1)
            base = base.strip()
            for suffix in suffixes:
                candidate = f"{base}{suffix} ({rest}"
                if candidate in mapping:
                    return candidate
        return None

    def _format_measurement(self, value, error=None, unit=""):
        if value is None or not np.isfinite(value):
            return None
        display_error = None
        if error is not None and np.isfinite(error) and error > 0:
            display_error = error
        if display_error is not None:
            text = f"{value:.3g} ± {display_error:.2g}"
        else:
            text = f"{value:.3g}"
        if unit:
            unit_text = str(unit).strip()
            if unit_text:
                return f"{text} ({unit_text})"
        return text

    def _build_caption_text_props(self, bold=False, underline=False):
        props = {
            "fontsize": max(self._get_font_size("caption"), 8),
            "ha": "left",
            "va": "baseline",
            "color": "black",
        }
        if bold:
            props["fontweight"] = "bold"
        if underline:
            props["underline"] = True
        return props

    def _generate_user_caption_lines(self, text):
        if not text:
            return []
        lines = []
        for raw in text.splitlines():
            stripped = raw.strip()
            if not stripped:
                continue
            line = CaptionLine()
            line.append(stripped)
            lines.append(line)
        return lines

    def _wrap_caption_lines(self, lines, renderer, max_width):
        if not lines:
            return []

        wrapped = []

        def iter_tokens(text):
            for match in re.finditer(r"\S+\s*|\s+", text):
                token = match.group(0)
                if token:
                    yield token

        for line in lines:
            tokens = []
            for span in line.spans:
                if not span.text:
                    continue
                for token in iter_tokens(span.text):
                    tokens.append((token, span.bold, span.underline))

            if not tokens:
                continue

            current_spans = []
            current_width = 0.0

            for token, bold, underline in tokens:
                props = self._build_caption_text_props(bold=bold, underline=underline)
                width = None
                if renderer is not None and max_width is not None:
                    try:
                        area = TextArea(token, textprops=props)
                        width = area.get_extent(renderer)[0]
                    except Exception:
                        width = None

                exceeds = (
                    max_width is not None
                    and width is not None
                    and current_width + width > max_width
                    and current_spans
                )

                if exceeds:
                    wrapped.append(CaptionLine(current_spans))
                    current_spans = []
                    current_width = 0.0

                current_spans.append(CaptionSpan(token, bold=bold, underline=underline))
                if width is not None:
                    current_width += width

            if current_spans:
                wrapped.append(CaptionLine(current_spans))

        return wrapped

    def _merge_caption_lines(self, *groups):
        merged = []
        for group in groups:
            for line in group or []:
                text = line.plain_text().strip()
                if not text:
                    continue
                merged.append(line)
        return merged


    def _build_caption_separator_line(self):
        line = CaptionLine()
        line.append("-" * 40)
        return line


    def _format_derived_entries(self, data):
        derived = data.get("derived") or {}
        if not derived:
            return []
        func_name = data.get("func_name")
        entries = []
        for key, value in derived.items():
            if self._is_error_key(key):
                continue
            try:
                numeric_value = float(value)
            except Exception:
                continue
            if not np.isfinite(numeric_value):
                continue
            error_key = self._find_matching_error_key(key, derived)
            error_val = derived.get(error_key) if error_key else None
            if error_val is not None:
                try:
                    error_val = float(error_val)
                except Exception:
                    error_val = None
            if error_val is not None and not np.isfinite(error_val):
                error_val = None
            base_name = key
            unit = ""
            if "(" in key and key.strip().endswith(")"):
                base, rest = key.split("(", 1)
                base_name = base.strip()
                unit = rest.rstrip(")").strip()
            descriptor = None
            if not unit:
                descriptor = DERIVED_VALUE_DIMENSIONS.get((func_name, base_name)) or DERIVED_VALUE_DIMENSIONS.get((func_name, key))
            if descriptor:
                scale, inferred_unit = self._dimension_scale_and_unit(descriptor, data)
            else:
                scale, inferred_unit = 1.0, ""
            display_value = numeric_value * scale
            display_error = error_val * abs(scale) if (error_val is not None) else None
            final_unit = inferred_unit or unit
            measurement = self._format_measurement(
                display_value, display_error, final_unit
            )
            if measurement:
                entries.append((base_name, measurement))
        return entries

    def _format_formula_text(self, func_name, formula):
        if not formula:
            return ""
        if func_name in PLAIN_FORMULAE:
            return PLAIN_FORMULAE[func_name]
        text = formula.strip()
        if text.startswith("$") and text.endswith("$") and len(text) >= 2:
            text = text[1:-1]
        replacements = {
            r"\\,": " ",
            r"\\ ": " ",
            r"\\left": "",
            r"\\right": "",
        }
        for src, dest in replacements.items():
            text = text.replace(src, dest)
        text = text.replace("\\", "")
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def _build_dataset_caption_lines(self, data, heading=None):
        lines = []
        func_name = data.get("func_name")
        if not func_name:
            return lines
        info = FIT_FUNCTIONS.get(func_name, {})
        title = info.get("title", func_name)
        formula = info.get("formula")
        plain_title = title or func_name or ""

        if heading:
            heading_line = CaptionLine()
            heading_line.append(heading, bold=True)
            lines.append(heading_line)

        if data.get("popt") is None:
            line = CaptionLine()
            if plain_title:
                line.append(plain_title, bold=True, underline=True)
                line.append(" did not converge.")
            else:
                line.append("Fit did not converge.")
            lines.append(line)
            return lines

        display_formula = self._format_formula_text(func_name, formula)

        if display_formula:
            line = CaptionLine()
            line.append(display_formula, bold=True)
            lines.append(line)
        elif plain_title:
            line = CaptionLine()
            line.append(plain_title, bold=True)
            lines.append(line)

        params = data.get("fit_param_details") or []
        if params:
            formatted_params = []
            for item in params:
                formatted = self._format_measurement(
                    item.get("value"), item.get("error"), item.get("unit") or ""
                )
                if formatted:
                    name = item.get("name") or ""
                    formatted_params.append((name, formatted))
            if formatted_params:
                label_line = CaptionLine()
                label_line.append("Parameters:", bold=True, underline=True)
                lines.append(label_line)
                for name, measurement in formatted_params:
                    if not measurement:
                        continue
                    entry = CaptionLine()
                    if name:
                        entry.append(f"{name}:", bold=True)
                        entry.append(" ")
                    entry.append(measurement)
                    lines.append(entry)

        derived_items = self._format_derived_entries(data)
        if derived_items:
            label_line = CaptionLine()
            label_line.append("Derived:", bold=True, underline=True)
            lines.append(label_line)
            for name, measurement in derived_items:
                if not measurement:
                    continue
                entry = CaptionLine()
                if name:
                    entry.append(f"{name}:", bold=True)
                    entry.append(" ")
                entry.append(measurement)
                lines.append(entry)

        return lines


    def _compose_single_caption(self, plot_num):
        data = self.plot_data.get(plot_num)
        if not data:
            return []
        auto_lines = self._build_dataset_caption_lines(data)
        user_text = data.get("caption", "")
        user_lines = self._generate_user_caption_lines(user_text)
        return self._merge_caption_lines(auto_lines, user_lines)

    def _compose_overlay_caption(self, indices, axis_sides):
        grouped_lines = []
        for idx in indices:
            data = self.plot_data[idx]
            base_label = data.get("y_label") or data.get("column") or f"Plot {idx}"
            if base_label and not base_label.startswith(f"Plot {idx}"):
                heading_text = f"Plot {idx}: {base_label}"
            else:
                heading_text = base_label or f"Plot {idx}"
            unit = data.get("y_unit") or ""
            unit = unit.strip() if isinstance(unit, str) else unit
            if unit:
                heading = f"{heading_text} ({unit})"
            else:
                heading = heading_text
            auto_lines = self._build_dataset_caption_lines(data, heading=heading)
            text_caption = (data.get("caption") or "").strip()
            user_lines = (
                self._generate_user_caption_lines(text_caption) if text_caption else []
            )
            merged_group = self._merge_caption_lines(auto_lines, user_lines)
            if merged_group:
                grouped_lines.append(merged_group)

        return grouped_lines

    def _apply_caption(self, plot_ax, caption_ax, lines):
        self._configure_caption_axis(caption_ax)
        if not lines:
            return 0

        canvas = plot_ax.figure.canvas
        renderer = None
        if canvas is not None:
            get_renderer = getattr(canvas, "get_renderer", None)
            if callable(get_renderer):
                try:
                    renderer = get_renderer()
                except Exception:
                    renderer = None
        if renderer is None:
            renderer = getattr(plot_ax.figure, "_cachedRenderer", None)
        if renderer is None and canvas is not None:
            try:
                canvas.draw()
                get_renderer = getattr(canvas, "get_renderer", None)
                if callable(get_renderer):
                    renderer = get_renderer()
            except Exception:
                renderer = None

        max_width = None
        if renderer is not None:
            plot_width = None
            try:
                plot_bbox = plot_ax.get_window_extent(renderer=renderer)
                plot_width = plot_bbox.width
            except Exception:
                plot_width = None
            if plot_width and plot_width > 0:
                max_width = plot_width * 0.8
            if max_width is None or max_width <= 0:
                try:
                    caption_bbox = caption_ax.get_window_extent(renderer=renderer)
                    max_width = caption_bbox.width * 0.8
                except Exception:
                    max_width = None

        if not lines:
            return 0

        if lines and isinstance(lines[0], list):
            groups = [group for group in lines if group]
        else:
            groups = [lines]
        if not groups:
            return 0

        column_width = None
        if max_width is not None and len(groups) > 0:
            column_width = max_width / len(groups)

        column_boxes = []
        max_lines = 0
        for group in groups:
            wrapped = self._wrap_caption_lines(group, renderer, column_width)
            if not wrapped:
                continue
            line_boxes = []
            for line in wrapped:
                span_boxes = []
                for span in line.spans:
                    if not span.text:
                        continue
                    props = self._build_caption_text_props(
                        bold=span.bold, underline=span.underline
                    )
                    try:
                        span_boxes.append(TextArea(span.text, textprops=props))
                    except Exception:
                        continue
                if not span_boxes:
                    continue
                line_boxes.append(
                    HPacker(children=span_boxes, align="baseline", pad=0, sep=0)
                )
            if not line_boxes:
                continue
            max_lines = max(max_lines, len(line_boxes))
            column_boxes.append(
                VPacker(children=line_boxes, align="top", pad=0, sep=4)
            )

        if not column_boxes:
            return 0

        if len(column_boxes) == 1:
            container = column_boxes[0]
        else:
            container = HPacker(children=column_boxes, align="top", pad=0, sep=30)

        anchored = AnchoredOffsetbox(
            loc="upper center",
            child=container,
            frameon=False,
            bbox_to_anchor=(0.5, 0.99),
            bbox_transform=caption_ax.transAxes,
            borderpad=0.0,
        )
        caption_ax.add_artist(anchored)
        return max_lines


class MultiFitDetailFrame(wx.Frame):
    """Display the data and fit contributing to a multi-fit aggregate point."""

    def __init__(self, fitting_window, runs, fit_key, display_name, column):
        title = display_name
        if column:
            title = f"{display_name} – {column}"
        super().__init__(fitting_window, title=title)
        self.fitting_window = fitting_window
        self.runs = runs
        self.fit_key = fit_key
        self.display_name = display_name
        self.column = column
        self.toolbars = []
        self.figures = []
        self.canvases = []
        self._run_artist_metadata = []

        main_sizer = wx.BoxSizer(wx.VERTICAL)
        if len(runs) == 1:
            panel = self._create_run_panel(self, runs[0])
            main_sizer.Add(panel, 1, wx.EXPAND)
        else:
            notebook = wx.Notebook(self)
            for run in runs:
                panel = self._create_run_panel(notebook, run)
                notebook.AddPage(panel, self._format_run_label(run))
            main_sizer.Add(notebook, 1, wx.EXPAND)

        self.SetSizer(main_sizer)
        self.SetSize((1150, 780))
        self.SetMinSize(wx.Size(900, 680))
        self.Layout()

    def _create_run_panel(self, parent, run):
        panel = wx.Panel(parent)
        sizer = wx.BoxSizer(wx.VERTICAL)
        fig = Figure(figsize=(10, 5))
        canvas = FigureCanvas(panel, -1, fig)
        summary_lines, metadata = self._populate_plot(fig, run)
        self._apply_figure_subtitle(fig, summary_lines)
        canvas.draw()
        toolbar = NavigationToolbar2Wx(canvas)
        toolbar.Realize()
        self.toolbars.append(toolbar)
        self.figures.append(fig)
        self.canvases.append(canvas)
        self._run_artist_metadata.append(metadata or [])
        sizer.Add(canvas, 1, wx.EXPAND)
        sizer.Add(toolbar, 0, wx.EXPAND)
        panel.SetSizer(sizer)
        panel.Layout()
        return panel

    def _populate_plot(self, fig, run):
        if self.fit_key in TEMP_PSD_FIT_KEYS:
            return self._populate_linear_temp_plot(fig, run)
        if self.fit_key == MOT_LIFETIME_DISPLAY_NAME:
            return self._populate_mot_lifetime_plot(fig, run)
        fig.text(
            0.5,
            0.5,
            "Detailed plotting is not available for this multi-fit.",
            ha="center",
            va="center",
            wrap=True,
        )
        return ["Detailed plotting is not available for this multi-fit."], []

    def _populate_linear_temp_plot(self, fig, run):
        metadata = []
        pick_radius = float(
            getattr(self.fitting_window, "_pick_tolerance", DEFAULT_PICK_TOLERANCE)
        )
        pick_radius = max(pick_radius, DEFAULT_PICK_TOLERANCE) * 1.25
        filtered_var = run.get("filtered_var_values", run.get("var_values", []))
        filtered_results = run.get("filtered_results", run.get("results", {}))
        analysis = get_temp_psd_analysis(
            filtered_var,
            filtered_results,
            self.fitting_window.parent,
        )
        derived = analysis.get("derived", {}) if analysis else {}
        analysis_valid = bool(analysis and "t_squared" in analysis)
        summary = self._build_summary_lines(run, derived)
        if not analysis_valid:
            summary.append("Not enough data to compute the fit.")

        included_indices = run.get("included_indices")
        if included_indices is None:
            included_indices = list(range(len(filtered_var)))
        excluded_indices = run.get("excluded_indices") or []

        time_scale = run.get("var_scale") or self.fitting_window.var_scale or 1.0
        time_unit = run.get("var_unit") or self.fitting_window.unit_scale or ""
        time_label = run.get("var_label") or self.fitting_window.varLabelCtrl.GetValue()
        time_label = time_label or "Variable"
        denom = time_scale ** 2 if time_scale else 1.0

        if analysis_valid:
            t_sq = np.asarray(analysis["t_squared"], dtype=float)
            sx_sq = np.asarray(analysis["sigma_x_squared"], dtype=float)
            sy_sq = np.asarray(analysis["sigma_y_squared"], dtype=float)
            err_x = analysis.get("sigma_x_squared_err")
            err_y = analysis.get("sigma_y_squared_err")
            slope_x = analysis.get("slope_x")
            slope_y = analysis.get("slope_y")
            intercept_x = analysis.get("intercept_x")
            intercept_y = analysis.get("intercept_y")
            t_sq_disp = t_sq / denom if denom else t_sq
        else:
            t_sq = np.asarray([], dtype=float)
            sx_sq = np.asarray([], dtype=float)
            sy_sq = np.asarray([], dtype=float)
            err_x = err_y = None
            slope_x = slope_y = None
            intercept_x = intercept_y = None
            t_sq_disp = t_sq

        t_sq_fit = None
        t_sq_fit_disp = None
        fit_x_disp = fit_x_coords = None
        fit_y_disp = fit_y_coords = None
        if analysis_valid and t_sq.size:
            t_min = float(np.min(t_sq))
            t_max = float(np.max(t_sq))
            if np.isclose(t_min, t_max):
                span = abs(t_min) if t_min else 1.0
                t_sq_fit = np.linspace(t_min - 0.1 * span, t_max + 0.1 * span, 200)
            else:
                t_sq_fit = np.linspace(t_min, t_max, 200)
            t_sq_fit_disp = t_sq_fit / denom if denom else t_sq_fit
        unit_x, factor_x = self.fitting_window.col_scales.get("x-True Width", ("", 1.0))
        unit_y, factor_y = self.fitting_window.col_scales.get("y-True Width", ("", 1.0))
        sx_sq_disp = sx_sq * (factor_x ** 2)
        sy_sq_disp = sy_sq * (factor_y ** 2)
        err_x_disp = None
        if err_x is not None:
            try:
                err_x_disp = np.asarray(err_x, dtype=float) * (factor_x ** 2)
            except Exception:
                err_x_disp = None
        err_y_disp = None
        if err_y is not None:
            try:
                err_y_disp = np.asarray(err_y, dtype=float) * (factor_y ** 2)
            except Exception:
                err_y_disp = None
        if analysis_valid and slope_x is not None and intercept_x is not None and t_sq_fit is not None:
            fit_x_disp = (slope_x * t_sq_fit + intercept_x) * (factor_x ** 2)
            fit_x_coords = t_sq_fit_disp
        if analysis_valid and slope_y is not None and intercept_y is not None and t_sq_fit is not None:
            fit_y_disp = (slope_y * t_sq_fit + intercept_y) * (factor_y ** 2)
            fit_y_coords = t_sq_fit_disp

        axes = fig.subplots(1, 2, sharex=True)
        fig.subplots_adjust(wspace=0.3)
        for axis in axes:
            axis.set_box_aspect(1)
        ax_x, ax_y = axes
        line_x = self._plot_with_errorbars(
            ax_x,
            t_sq_disp,
            sx_sq_disp,
            err_x_disp,
            fit_x_coords,
            fit_x_disp,
            color="tab:blue",
        )
        if line_x is not None:
            line_x.set_pickradius(pick_radius)
            try:
                line_x.set_picker(pick_radius)
            except Exception:
                line_x.set_picker(True)
            entry = {
                "artist": line_x,
                "measurement_indices": included_indices[: len(t_sq_disp)],
                "status": "included",
                "component": "sigma_x",
                "axis": ax_x,
                "x_values": np.asarray(t_sq_disp, dtype=float),
                "y_values": np.asarray(sx_sq_disp, dtype=float),
            }
            if err_x_disp is not None:
                try:
                    entry["error_values"] = np.asarray(err_x_disp, dtype=float)
                except Exception:
                    pass
            metadata.append(entry)
        line_y = self._plot_with_errorbars(
            ax_y,
            t_sq_disp,
            sy_sq_disp,
            err_y_disp,
            fit_y_coords,
            fit_y_disp,
            color="tab:orange",
        )
        if line_y is not None:
            line_y.set_pickradius(pick_radius)
            try:
                line_y.set_picker(pick_radius)
            except Exception:
                line_y.set_picker(True)
            entry = {
                "artist": line_y,
                "measurement_indices": included_indices[: len(t_sq_disp)],
                "status": "included",
                "component": "sigma_y",
                "axis": ax_y,
                "x_values": np.asarray(t_sq_disp, dtype=float),
                "y_values": np.asarray(sy_sq_disp, dtype=float),
            }
            if err_y_disp is not None:
                try:
                    entry["error_values"] = np.asarray(err_y_disp, dtype=float)
                except Exception:
                    pass
            metadata.append(entry)

        all_var = np.asarray(run.get("var_values", []), dtype=float)
        sigma_x_all = np.asarray(run.get("results", {}).get("x-True Width", []), dtype=float)
        sigma_y_all = np.asarray(run.get("results", {}).get("y-True Width", []), dtype=float)
        if excluded_indices and all_var.size:
            try:
                excluded_indices = [int(i) for i in excluded_indices if int(i) < all_var.size]
            except Exception:
                excluded_indices = []
        if excluded_indices:
            t_sq_all_disp = np.square(all_var) / denom
            sx_sq_all_disp = np.square(sigma_x_all) * (factor_x ** 2)
            sy_sq_all_disp = np.square(sigma_y_all) * (factor_y ** 2)
            if excluded_indices and len(t_sq_all_disp) >= max(excluded_indices) + 1:
                excl_t = t_sq_all_disp[excluded_indices]
                excl_x = sx_sq_all_disp[excluded_indices]
                excl_y = sy_sq_all_disp[excluded_indices]
                if excl_t.size:
                    ex_line_x = ax_x.plot(
                        excl_t,
                        excl_x,
                        "x",
                        color="tab:gray",
                        alpha=0.6,
                        markersize=6,
                    )[0]
                    ex_line_x.set_picker(pick_radius)
                    ex_line_x.set_pickradius(pick_radius)
                    metadata.append(
                        {
                            "artist": ex_line_x,
                            "measurement_indices": excluded_indices[: len(excl_t)],
                            "status": "excluded",
                            "component": "sigma_x",
                            "axis": ax_x,
                            "x_values": np.asarray(excl_t, dtype=float),
                            "y_values": np.asarray(excl_x, dtype=float),
                        }
                    )
                if excl_t.size:
                    ex_line_y = ax_y.plot(
                        excl_t,
                        excl_y,
                        "x",
                        color="tab:gray",
                        alpha=0.6,
                        markersize=6,
                    )[0]
                    ex_line_y.set_picker(pick_radius)
                    ex_line_y.set_pickradius(pick_radius)
                    metadata.append(
                        {
                            "artist": ex_line_y,
                            "measurement_indices": excluded_indices[: len(excl_t)],
                            "status": "excluded",
                            "component": "sigma_y",
                            "axis": ax_y,
                            "x_values": np.asarray(excl_t, dtype=float),
                            "y_values": np.asarray(excl_y, dtype=float),
                        }
                    )

        x_axis = self._format_axis_label(time_label, time_unit, squared=True)
        x_axis_label = x_axis
        if x_axis_label:
            ax_x.set_xlabel(x_axis_label, fontsize=AXIS_LABEL_FONTSIZE)
            ax_y.set_xlabel(x_axis_label, fontsize=AXIS_LABEL_FONTSIZE)
        else:
            ax_x.set_xlabel("")
            ax_y.set_xlabel("")
        yx_label = self._format_axis_label("σ_x", unit_x, squared=True)
        yy_label = self._format_axis_label("σ_y", unit_y, squared=True)
        if yx_label:
            ax_x.set_ylabel(yx_label, fontsize=AXIS_LABEL_FONTSIZE)
        else:
            ax_x.set_ylabel("")
        if yy_label:
            ax_y.set_ylabel(yy_label, fontsize=AXIS_LABEL_FONTSIZE)
        else:
            ax_y.set_ylabel("")
        ax_x.set_title("x-axis fit")
        ax_y.set_title("y-axis fit")

        if not analysis_valid:
            fig.text(
                0.5,
                0.5,
                "Not enough data to compute the temperature/PSD fit.",
                ha="center",
                va="center",
                wrap=True,
            )

        return summary, metadata

    def _populate_mot_lifetime_plot(self, fig, run):
        result = fit_multi_run_mot_lifetime(
            run.get("var_values", []), run.get("results", {})
        )
        derived = {}
        if not result:
            fig.text(
                0.5,
                0.5,
                "Not enough data to compute the MOT lifetime fit.",
                ha="center",
                va="center",
                wrap=True,
            )
            summary = self._build_summary_lines(run, derived)
            summary.append("Not enough data to compute the fit.")
            return summary

        times = np.asarray(result["times"], dtype=float)
        atom_counts = np.asarray(result["atom_counts"], dtype=float)
        popt = result.get("popt")
        pcov = result.get("pcov")
        if popt is None or times.size == 0:
            fig.text(
                0.5,
                0.5,
                "Not enough data to compute the MOT lifetime fit.",
                ha="center",
                va="center",
                wrap=True,
            )
            summary = self._build_summary_lines(run, derived)
            summary.append("Not enough data to compute the fit.")
            return summary

        def model(t):
            return popt[0] * np.exp(-t / popt[1]) + popt[2]

        fitted = model(times)
        ss_res = np.sum((fitted - atom_counts) ** 2)
        ss_tot = np.sum((atom_counts - np.mean(atom_counts)) ** 2)
        r2 = 1 - ss_res / ss_tot if ss_tot != 0 else float("nan")

        time_scale = run.get("var_scale") or self.fitting_window.var_scale or 1.0
        time_unit = run.get("var_unit") or self.fitting_window.unit_scale or ""
        time_label = run.get("var_label") or self.fitting_window.varLabelCtrl.GetValue()
        time_label = time_label or "Time"

        denom = time_scale if time_scale else 1.0
        times_disp = times / denom

        t_min = float(np.min(times))
        t_max = float(np.max(times))
        if np.isclose(t_min, t_max):
            span = abs(t_min) if t_min else 1.0
            t_fit = np.linspace(t_min - 0.1 * span, t_max + 0.1 * span, 200)
        else:
            t_fit = np.linspace(t_min, t_max, 200)
        t_fit_disp = t_fit / denom
        fit_disp = model(t_fit)

        y_unit, y_factor = self.fitting_window.col_scales.get("Atom Number", ("", 1.0))
        if y_factor is None:
            y_factor = 1.0
        atom_disp = atom_counts * y_factor
        fit_line_disp = fit_disp * y_factor
        ax_main = fig.subplots()

        tau_disp = popt[1] / denom if denom else popt[1]
        n0_disp = popt[0] * y_factor
        offset_disp = popt[2] * y_factor
        fit_label_parts = [
            f"τ={tau_disp:.3g}{' ' + time_unit if time_unit else ''}",
            f"N₀={n0_disp:.3g}{' ' + y_unit if y_unit else ''}",
            f"C={offset_disp:.3g}{' ' + y_unit if y_unit else ''}",
        ]
        if np.isfinite(r2):
            fit_label_parts.append(f"R²={r2:.3f}")
        fit_label = "Fit: " + ", ".join(fit_label_parts)

        ax_main.plot(t_fit_disp, fit_line_disp, color="tab:red", label=fit_label)
        ax_main.scatter(times_disp, atom_disp, color="tab:blue", label="Data")
        ax_main.set_ylabel(
            "Atom Number" + (f" ({y_unit})" if y_unit else ""),
            fontsize=AXIS_LABEL_FONTSIZE,
        )
        ax_main.set_title("MOT lifetime fit")
        ax_main.legend(loc="best")
        ax_main.set_xlabel(
            self._format_axis_label(time_label, time_unit),
            fontsize=AXIS_LABEL_FONTSIZE,
        )
        ax_main.grid(True, alpha=0.2)

        derived = FIT_FUNCTIONS[MOT_LIFETIME_DISPLAY_NAME]["derived"](
            popt,
            None,
            run.get("var_values", []),
            run.get("results", {}),
            self.fitting_window.parent,
            pcov,
        )
        derived = dict(derived or {})
        unit = time_unit
        scale = denom if denom else 1.0
        lifetime_key = "Lifetime (s)"
        if lifetime_key in derived:
            lifetime_val = derived.pop(lifetime_key)
            label = f"Lifetime ({unit})" if unit else "Lifetime"
            derived[label] = lifetime_val / scale
            err_key = "Lifetime (s)_err"
            if err_key in derived:
                err_val = derived.pop(err_key)
                derived[f"{label}_err"] = err_val / scale
        rate_key = "Decay rate (1/s)"
        if rate_key in derived:
            rate_val = derived.pop(rate_key)
            rate_label = f"Decay rate (1/{unit})" if unit else "Decay rate (1/s)"
            derived[rate_label] = rate_val * scale

        n0_err = None
        offset_err = None
        if pcov is not None:
            try:
                n0_err = float(np.sqrt(pcov[0, 0]) * y_factor)
            except Exception:
                n0_err = None
            try:
                offset_err = float(np.sqrt(pcov[2, 2]) * y_factor)
            except Exception:
                offset_err = None
        n0_label = f"N0 ({y_unit})" if y_unit else "N0"
        derived[n0_label] = n0_disp
        if n0_err is not None and np.isfinite(n0_err):
            derived[f"{n0_label}_err"] = n0_err
        offset_label = f"Offset ({y_unit})" if y_unit else "Offset"
        derived[offset_label] = offset_disp
        if offset_err is not None and np.isfinite(offset_err):
            derived[f"{offset_label}_err"] = offset_err

        original_column = self.column
        highlight_map = {}
        if unit:
            highlight_map["Lifetime (s)"] = f"Lifetime ({unit})"
            highlight_map["Decay rate (1/s)"] = f"Decay rate (1/{unit})"
        if y_unit:
            highlight_map["N0"] = f"N0 ({y_unit})"
            highlight_map["Offset"] = f"Offset ({y_unit})"
        self.column = highlight_map.get(original_column, original_column)
        summary = self._build_summary_lines(run, derived)
        self.column = original_column
        if np.isfinite(r2):
            summary.append(f"R² = {r2:.3f}")
        return summary

    def _plot_with_errorbars(self, ax, x, y, err, fit_x=None, fit_y=None, color="tab:blue"):
        line = None
        if err is not None and len(err) == len(y):
            container = ax.errorbar(x, y, yerr=err, fmt="o", capsize=3, color=color)
            if container and container.lines:
                line = container.lines[0]
        else:
            plotted = ax.plot(x, y, "o", color=color)
            if plotted:
                line = plotted[0]
        if fit_x is not None and fit_y is not None:
            ax.plot(fit_x, fit_y, color="tab:red")
        ax.grid(True, alpha=0.2)
        if line is not None:
            line.set_picker(True)
        return line

    def _build_summary_lines(self, run, derived):
        lines = []
        primary = self._format_value_line(run, "param", "param_label", "param_unit", "param_scale")
        if primary:
            lines.append(primary)
        secondary = self._format_value_line(
            run, "param2", "param2_label", "param2_unit", "param2_scale"
        )
        if secondary:
            lines.append(secondary)
        derived_lines = self._format_derived_lines(derived)
        lines.extend(derived_lines)
        return lines

    def _format_value_line(self, run, value_key, label_key, unit_key, scale_key):
        if value_key not in run:
            return None
        value = run.get(value_key)
        label = run.get(label_key) or ""
        if value is None or label == "":
            return None
        scale = run.get(scale_key, 1.0) or 1.0
        unit = run.get(unit_key) or ""
        display_val = value / scale if scale else value
        unit_str = f" {unit}" if unit else ""
        return f"{label} = {display_val:.6g}{unit_str}"

    def _matches_error_key(self, name):
        checker = getattr(self.fitting_window, "_is_error_key", None)
        if callable(checker):
            try:
                return bool(checker(name))
            except Exception:
                pass
        return self._default_is_error_key(name)

    def _default_is_error_key(self, name):
        if not isinstance(name, str):
            return False
        stripped = name.strip()
        if "(" in stripped and stripped.endswith(")"):
            stripped = stripped.split("(", 1)[0].strip()
        lowered = stripped.lower()
        suffixes = ("_err", " err", "_std", " std", "_stderr", " stderr", "_error", " error")
        return any(lowered.endswith(suffix.strip()) for suffix in suffixes)

    def _is_error_key(self, name):
        return self._matches_error_key(name)

    def _find_error_key(self, name, mapping):
        suffixes = [" Std", " Err", " std", " err", "_Std", "_Err", "_std", "_err"]
        for suffix in suffixes:
            candidate = f"{name}{suffix}"
            if candidate in mapping:
                return candidate
        if "(" in name:
            base, rest = name.split("(", 1)
            base = base.strip()
            for suffix in suffixes:
                candidate = f"{base}{suffix} ({rest}"
                if candidate in mapping:
                    return candidate
        return None

    def _format_derived_lines(self, derived):
        if not derived:
            return []
        lines = []
        highlight = None
        for key, value in derived.items():
            if self._matches_error_key(key):
                continue
            if not np.isscalar(value) or not np.isfinite(value):
                continue
            err_key = self._find_error_key(key, derived)
            if err_key and np.isscalar(derived[err_key]) and np.isfinite(derived[err_key]):
                line = f"{key} = {value:.6g} ± {derived[err_key]:.2g}"
            else:
                line = f"{key} = {value:.6g}"
            if key == self.column:
                highlight = line
            else:
                lines.append(line)
        if highlight:
            return [highlight] + lines
        return lines

    def _format_axis_label(self, label, unit, squared=False):
        text = label
        if squared and label and not label.endswith("²"):
            text = f"{label}²"
        if squared and unit:
            return f"{text} ({unit}²)"
        if unit:
            return f"{text} ({unit})"
        return text

    def _apply_figure_subtitle(self, fig, lines):
        if not lines:
            if fig.axes:
                try:
                    fig.tight_layout()
                except Exception:
                    pass
            return
        subtitle = "\n".join(lines)
        bottom_margin = 0.1 + 0.025 * len(lines)
        bottom_margin = min(max(bottom_margin, 0.12), 0.35)
        if fig.axes:
            try:
                fig.tight_layout(rect=(0, bottom_margin, 1, 0.94))
            except Exception:
                pass
        text_y = max(0.02, bottom_margin - 0.01)
        fig.text(
            0.5,
            text_y,
            subtitle,
            ha="center",
            va="top",
            wrap=True,
            fontsize=9,
        )

    def _format_run_label(self, run):
        idx = run.get("index")
        prefix = f"#{idx + 1}" if idx is not None else "Run"
        label = run.get("param_label")
        value = run.get("param")
        if label and value is not None:
            scale = run.get("param_scale", 1.0) or 1.0
            unit = run.get("param_unit") or ""
            disp = value / scale if scale else value
            unit_str = f" {unit}" if unit else ""
            return f"{prefix} {label}={disp:.3g}{unit_str}"
        return prefix

    def _refresh_plots(self):
        if len(self._run_artist_metadata) > len(self.runs):
            self._run_artist_metadata = self._run_artist_metadata[: len(self.runs)]
        for idx, fig in enumerate(self.figures):
            if idx >= len(self.runs):
                break
            canvas = self.canvases[idx] if idx < len(self.canvases) else None
            run = self.runs[idx]
            fig.clf()
            summary, metadata = self._populate_plot(fig, run)
            if idx < len(self._run_artist_metadata):
                self._run_artist_metadata[idx] = metadata or []
            else:
                self._run_artist_metadata.append(metadata or [])
            self._apply_figure_subtitle(fig, summary)
            if canvas is not None:
                canvas.draw()


class PointDetailFrame(MultiFitDetailFrame):
    """Inspector window for a picked data point with image thumbnails."""

    def __init__(
        self,
        fitting_window,
        key,
        orig_indices,
        runs,
        fit_key,
        display_name,
        column,
        point_info,
        image_paths,
    ):
        self.key = key
        self.orig_indices = tuple(orig_indices)
        self.point_info = dict(point_info or {})
        self.image_entries = [dict(entry) for entry in (image_paths or [])]
        self._thumbnails = []
        self.info_label = None
        self.exclude_checkbox = None
        self.scrolled_window = None
        self.thumbnail_sizer = None
        self._canvas_connections = {}
        self._artist_metadata = []
        self.base_runs = [copy.deepcopy(run) for run in (runs or [])]
        self.measurement_panel = None
        self.measurement_instructions = None
        self.measurement_list_sizer = None
        self._measurement_entries = {}
        self._hover_annotations = {}
        self._hover_last = (None, None, None)
        self.accept_button = None
        self._baseline_exclusions = {}
        self._preview_exclusions = {}
        self._baseline_point_excluded = False
        self._preview_point_excluded = False
        tolerance = getattr(fitting_window, "_hover_snap_distance_sq", None)
        self._hover_snap_distance_sq = (
            float(tolerance)
            if tolerance is not None
            else HOVER_SNAP_DISTANCE_SQ
        )

        super().__init__(fitting_window, runs, fit_key, display_name, column)

        existing_sizer = self.GetSizer()
        content_window = None
        if existing_sizer and existing_sizer.GetItemCount():
            item = existing_sizer.GetItem(0)
            if item:
                content_window = item.GetWindow()
                if content_window:
                    existing_sizer.Detach(content_window)
        if existing_sizer:
            existing_sizer.Clear(delete_windows=False)

        root_sizer = wx.BoxSizer(wx.VERTICAL)
        self.summary_box = wx.StaticBoxSizer(wx.StaticBox(self, label="Point Summary"), wx.VERTICAL)
        self.detail_box = wx.StaticBoxSizer(wx.StaticBox(self, label="Fit Details"), wx.VERTICAL)
        self.images_box = wx.StaticBoxSizer(wx.StaticBox(self, label="Source Images"), wx.VERTICAL)
        root_sizer.Add(self.summary_box, 0, wx.EXPAND | wx.ALL, 10)
        root_sizer.Add(self.detail_box, 1, wx.EXPAND | wx.LEFT | wx.RIGHT, 10)
        root_sizer.Add(self.images_box, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 10)
        self.SetSizer(root_sizer)
        min_height = int(round((FILMSTRIP_MIN_HEIGHT + 30) * FILMSTRIP_PANEL_SCALE))
        self.images_box.SetMinSize(wx.Size(-1, min_height))

        button_panel = wx.Panel(self)
        button_sizer = wx.BoxSizer(wx.HORIZONTAL)
        button_sizer.AddStretchSpacer()
        self.accept_button = wx.Button(button_panel, label="Accept")
        self.accept_button.Disable()
        self.accept_button.Bind(wx.EVT_BUTTON, self._on_accept)
        button_sizer.Add(self.accept_button, 0)
        button_panel.SetSizer(button_sizer)
        root_sizer.Add(button_panel, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)

        if content_window:
            try:
                content_window.Reparent(self)
            except Exception:
                pass
            self.detail_box.Add(content_window, 1, wx.EXPAND)

        base_title = self.GetTitle() or "Point Detail"
        plot_num = self.point_info.get("plot")
        if plot_num is not None:
            self.SetTitle(f"Plot {plot_num} – {base_title}")

        self._insert_header_panel()
        self._insert_measurement_panel()
        self._insert_filmstrip_panel()
        self.update_context(runs, self.point_info, self.image_entries, fit_key, display_name, column)
        self.SetSize((1150, 900))
        self.SetMinSize(wx.Size(900, 760))
        self.Layout()

    def update_context(self, runs, point_info, image_paths, fit_key, display_name, column):
        self.base_runs = [copy.deepcopy(run) for run in (runs or [])]
        self.point_info = dict(point_info or {})
        self.image_entries = [dict(entry) for entry in (image_paths or [])]
        if display_name:
            title = display_name
            if column:
                title = f"{display_name} – {column}"
            plot_num = self.point_info.get("plot")
            if plot_num is not None:
                title = f"Plot {plot_num} – {title}"
            self.SetTitle(title)
        if fit_key is not None:
            self.fit_key = fit_key
        self.column = column
        parent = self.fitting_window
        self._baseline_point_excluded = bool(self.point_info.get("excluded"))
        self._preview_point_excluded = self._baseline_point_excluded
        self.point_info["excluded"] = self._preview_point_excluded
        self._baseline_exclusions = {}
        if parent:
            for idx in self.orig_indices:
                if parent and hasattr(parent, "run_point_exclusions"):
                    base = parent.run_point_exclusions.get(idx, set())
                else:
                    base = set()
                try:
                    self._baseline_exclusions[idx] = {int(val) for val in base}
                except Exception:
                    self._baseline_exclusions[idx] = set(base)
        else:
            for run in self.base_runs:
                run_idx = run.get("index")
                if run_idx is not None:
                    self._baseline_exclusions[run_idx] = set()
        self._preview_exclusions = {
            idx: set(values) for idx, values in self._baseline_exclusions.items()
        }
        self._apply_preview(refresh=True)
        self._update_header()
        self._populate_filmstrip()
        self._refresh_measurement_panel()
        self._update_accept_state()
        self.Layout()

    def _insert_header_panel(self):
        target = getattr(self, "summary_box", None)
        if target is None:
            target = self.GetSizer()
        header_panel = wx.Panel(self)
        header_sizer = wx.BoxSizer(wx.VERTICAL)
        self.info_label = wx.StaticText(header_panel, label="")
        info_font = self.info_label.GetFont()
        info_font.SetWeight(wx.FONTWEIGHT_BOLD)
        self.info_label.SetFont(info_font)
        header_sizer.Add(self.info_label, 0)
        self.exclude_checkbox = wx.CheckBox(
            header_panel, label="Exclude this data point from fit"
        )
        header_sizer.Add(self.exclude_checkbox, 0, wx.TOP, 6)
        self.exclude_checkbox.Bind(wx.EVT_CHECKBOX, self._on_exclude_changed)
        header_panel.SetSizer(header_sizer)
        header_panel.Layout()
        if target:
            target.Add(header_panel, 0, wx.EXPAND | wx.BOTTOM, 8)
        else:
            fallback = wx.BoxSizer(wx.VERTICAL)
            fallback.Add(header_panel, 0, wx.EXPAND | wx.ALL, 10)
            self.SetSizer(fallback)

    def _insert_measurement_panel(self):
        target = getattr(self, "summary_box", None)
        if target is None:
            target = self.GetSizer()
        panel = wx.Panel(self)
        static_box = wx.StaticBox(panel, label="Selected Measurements")
        box = wx.StaticBoxSizer(static_box, wx.VERTICAL)
        instructions = wx.StaticText(
            panel,
            label="Click a subplot point to inspect it. Use Accept to apply changes.",
        )
        instructions.Wrap(520)
        box.Add(instructions, 0, wx.BOTTOM, 6)
        list_sizer = wx.BoxSizer(wx.VERTICAL)
        box.Add(list_sizer, 0, wx.EXPAND)
        panel.SetSizer(box)
        panel.Hide()
        if target:
            target.Add(panel, 0, wx.EXPAND | wx.BOTTOM, 6)
        self.measurement_panel = panel
        self.measurement_instructions = instructions
        self.measurement_list_sizer = list_sizer

    def _insert_filmstrip_panel(self):
        target = getattr(self, "images_box", None)
        if target is None:
            target = self.GetSizer()
        container = wx.Panel(self)
        container_sizer = wx.BoxSizer(wx.VERTICAL)
        self.scrolled_window = wx.ScrolledWindow(
            container, style=wx.HSCROLL | wx.VSCROLL | wx.BORDER_SIMPLE
        )
        self.scrolled_window.SetScrollRate(10, 10)
        self.thumbnail_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.scrolled_window.SetSizer(self.thumbnail_sizer)
        min_height = int(round(FILMSTRIP_MIN_HEIGHT * FILMSTRIP_PANEL_SCALE))
        self.scrolled_window.SetMinSize(wx.Size(-1, min_height))
        container_sizer.Add(self.scrolled_window, 1, wx.EXPAND)
        container.SetSizer(container_sizer)
        container.Layout()
        if target:
            target.Add(container, 1, wx.EXPAND)
        else:
            fallback = wx.BoxSizer(wx.VERTICAL)
            fallback.Add(container, 1, wx.EXPAND | wx.ALL, 10)
            self.SetSizer(fallback)

    def _apply_preview(self, refresh=True):
        parent = self.fitting_window
        preview_runs = None
        if parent and hasattr(parent, "collect_detail_runs_preview"):
            try:
                preview_runs = parent.collect_detail_runs_preview(
                    self.orig_indices, self._preview_exclusions
                )
            except Exception:
                preview_runs = None
        if not preview_runs:
            preview_runs = [copy.deepcopy(run) for run in self.base_runs]
        self.runs = preview_runs
        if refresh:
            self._refresh_plots()
            self._bind_pick_events()

    def _has_pending_changes(self):
        if self._preview_point_excluded != self._baseline_point_excluded:
            return True
        baseline = self._baseline_exclusions or {}
        preview = self._preview_exclusions or {}
        indices = set(baseline.keys()) | set(preview.keys())
        for idx in indices:
            base = set(baseline.get(idx, set()))
            desired = set(preview.get(idx, set()))
            if base != desired:
                return True
        return False

    def _update_accept_state(self):
        if not self.accept_button:
            return
        self.accept_button.Enable(self._has_pending_changes())

    def _on_accept(self, event):
        parent = self.fitting_window
        if parent:
            if hasattr(parent, "set_run_exclusions_bulk"):
                for run_idx, exclusions in self._preview_exclusions.items():
                    try:
                        parent.set_run_exclusions_bulk(run_idx, exclusions)
                    except Exception:
                        continue
            if (
                self._preview_point_excluded != self._baseline_point_excluded
                and hasattr(parent, "apply_point_exclusion")
            ):
                try:
                    parent.apply_point_exclusion(
                        self.key, self.orig_indices, self._preview_point_excluded
                    )
                except Exception:
                    pass
            if hasattr(parent, "note_subpoint_exclusions"):
                try:
                    parent.note_subpoint_exclusions(self.key, self.orig_indices)
                except Exception:
                    pass
        self.Close()

    def _resolve_group_members(self, run_index, measurement_idx):
        run = next((r for r in self.runs if r.get("index") == run_index), None)
        if not run:
            return (measurement_idx,)
        groups = run.get("variable_groups") or {}
        if measurement_idx in groups:
            try:
                return tuple(int(val) for val in groups[measurement_idx])
            except Exception:
                return tuple(groups[measurement_idx])
        for members in groups.values():
            try:
                seq = tuple(int(val) for val in members)
            except Exception:
                seq = tuple(members)
            if measurement_idx in seq:
                return seq
        return (measurement_idx,)

    def _disconnect_pick_events(self):
        for canvas, ids in list(self._canvas_connections.items()):
            if not isinstance(ids, (list, tuple)):
                ids = [ids]
            for cid in ids:
                try:
                    canvas.mpl_disconnect(cid)
                except Exception:
                    pass
        self._canvas_connections.clear()
        self._hover_annotations.clear()
        self._hover_last = (None, None, None)

    def _update_artist_metadata(self):
        self._artist_metadata = []
        tolerance = getattr(self.fitting_window, "_pick_tolerance", DEFAULT_PICK_TOLERANCE)
        try:
            tolerance = float(tolerance)
        except Exception:
            tolerance = float(DEFAULT_PICK_TOLERANCE)
        if tolerance <= 0:
            tolerance = float(DEFAULT_PICK_TOLERANCE)
        for idx, entries in enumerate(self._run_artist_metadata):
            run = self.runs[idx] if idx < len(self.runs) else {}
            run_index = run.get("index")
            for entry in entries or []:
                artist = entry.get("artist")
                if artist is None:
                    continue
                success = False
                if hasattr(artist, "set_pickradius"):
                    try:
                        artist.set_pickradius(tolerance)
                    except Exception:
                        pass
                if hasattr(artist, "set_picker"):
                    try:
                        artist.set_picker(tolerance)
                        success = True
                    except Exception:
                        try:
                            artist.set_picker(True)
                            success = True
                        except Exception:
                            success = False
                if not success:
                    continue
                meta = dict(entry)
                meta.setdefault("run_index", run_index)
                meta.setdefault("run_position", idx)
                axis = meta.get("axis")
                if axis is None:
                    axis = getattr(artist, "axes", None)
                meta["axis"] = axis
                if "x_values" not in meta or meta.get("x_values") is None:
                    getter = getattr(artist, "get_xdata", None)
                    if callable(getter):
                        try:
                            meta["x_values"] = np.asarray(getter(), dtype=float)
                        except Exception:
                            meta["x_values"] = None
                if "y_values" not in meta or meta.get("y_values") is None:
                    getter = getattr(artist, "get_ydata", None)
                    if callable(getter):
                        try:
                            meta["y_values"] = np.asarray(getter(), dtype=float)
                        except Exception:
                            meta["y_values"] = None
                self._artist_metadata.append(meta)

    def _bind_pick_events(self):
        self._disconnect_pick_events()
        self._update_artist_metadata()
        for canvas in self.canvases:
            ids = []
            try:
                ids.append(canvas.mpl_connect("pick_event", self._on_canvas_pick))
            except Exception:
                continue
            try:
                ids.append(canvas.mpl_connect("motion_notify_event", self._on_canvas_motion))
            except Exception:
                pass
            try:
                ids.append(canvas.mpl_connect("figure_leave_event", self._on_canvas_leave))
            except Exception:
                pass
            if ids:
                self._canvas_connections[canvas] = ids

    def _get_hover_annotation(self, ax):
        annotation = self._hover_annotations.get(ax)
        if (
            annotation is None
            or annotation.axes is not ax
            or annotation not in getattr(ax, "texts", [])
        ):
            font_size = 10.0
            if self.fitting_window and hasattr(self.fitting_window, "_get_font_size"):
                try:
                    font_size = float(self.fitting_window._get_font_size("tick"))
                except Exception:
                    font_size = 10.0
            annotation = ax.annotate(
                "",
                xy=(0, 0),
                xytext=(10, 10),
                textcoords="offset points",
                bbox={"boxstyle": "round,pad=0.3", "fc": "white", "alpha": 0.85},
                arrowprops={"arrowstyle": "->", "color": "black"},
                fontsize=font_size,
            )
            annotation.set_visible(False)
            self._hover_annotations[ax] = annotation
        return annotation

    def _format_hover_text(self, meta, x_value, y_value, error=None):
        formatter = getattr(self.fitting_window, "_format_hover_text", None)
        if callable(formatter):
            try:
                return formatter(meta, float(x_value), float(y_value), error)
            except Exception:
                pass
        lines = [f"x: {float(x_value):.3g}", f"y: {float(y_value):.3g}"]
        if error is None:
            lines.append("err: n/a")
        elif isinstance(error, tuple):
            lower, upper = error
            lower_text = "n/a"
            upper_text = "n/a"
            if lower is not None and np.isfinite(lower):
                lower_text = f"{abs(lower):.3g}"
            if upper is not None and np.isfinite(upper):
                upper_text = f"{abs(upper):.3g}"
            lines.append(f"err: -{lower_text}/+{upper_text}")
        else:
            if np.isfinite(error):
                lines.append(f"err: {abs(error):.3g}")
            else:
                lines.append("err: n/a")
        return "\n".join(lines)

    def _hide_hover(self, axis=None):
        changed = False
        if axis is None:
            for annotation in self._hover_annotations.values():
                if annotation is not None and annotation.get_visible():
                    annotation.set_visible(False)
                    changed = True
            self._hover_last = (None, None, None)
        else:
            annotation = self._hover_annotations.get(axis)
            if annotation is not None and annotation.get_visible():
                annotation.set_visible(False)
                changed = True
            if self._hover_last[0] is axis:
                self._hover_last = (None, None, None)
        if changed:
            if axis is not None and hasattr(axis, "figure"):
                canvas = getattr(axis.figure, "canvas", None)
                if canvas is not None:
                    canvas.draw_idle()
            else:
                for canvas in self.canvases:
                    if canvas is not None:
                        canvas.draw_idle()
                        break

    def _on_canvas_motion(self, event):
        ax = getattr(event, "inaxes", None)
        if ax is None:
            self._hide_hover()
            return
        xdata = getattr(event, "xdata", None)
        ydata = getattr(event, "ydata", None)
        if xdata is None or ydata is None:
            self._hide_hover(axis=ax)
            return
        try:
            mouse_disp = ax.transData.transform((xdata, ydata))
        except Exception:
            self._hide_hover(axis=ax)
            return
        best = None
        for meta in self._artist_metadata:
            if meta.get("axis") is not ax:
                continue
            artist = meta.get("artist")
            if artist is None:
                continue
            x_values = meta.get("x_values")
            y_values = meta.get("y_values")
            error_values = meta.get("error_values")
            try:
                xs = np.asarray(x_values, dtype=float)
                ys = np.asarray(y_values, dtype=float)
            except Exception:
                continue
            if xs.size == 0 or ys.size == 0:
                continue
            length = min(xs.size, ys.size)
            if length <= 0:
                continue
            xs = xs[:length]
            ys = ys[:length]
            err_values = None
            err_lower = None
            err_upper = None
            if error_values is not None:
                try:
                    err_arr = np.asarray(error_values, dtype=float)
                    if err_arr.ndim == 1:
                        err_values = err_arr[:length]
                    elif err_arr.ndim == 2 and err_arr.shape[0] == 2:
                        err_lower = err_arr[0][:length]
                        err_upper = err_arr[1][:length]
                except Exception:
                    err_values = None
                    err_lower = None
                    err_upper = None
            valid = np.isfinite(xs) & np.isfinite(ys)
            if not np.any(valid):
                continue
            xs = xs[valid]
            ys = ys[valid]
            if err_values is not None:
                err_values = err_values[valid]
            if err_lower is not None and err_upper is not None:
                err_lower = err_lower[valid]
                err_upper = err_upper[valid]
            if xs.size == 0:
                continue
            try:
                points_disp = ax.transData.transform(np.column_stack((xs, ys)))
            except Exception:
                continue
            deltas = points_disp - mouse_disp
            dist_sq = np.sum(deltas**2, axis=1)
            if dist_sq.size == 0 or not np.any(np.isfinite(dist_sq)):
                continue
            idx = int(np.nanargmin(dist_sq))
            distance = float(dist_sq[idx])
            if distance > self._hover_snap_distance_sq:
                continue
            err_value = None
            if err_values is not None and idx < err_values.size:
                err_value = float(err_values[idx])
            elif (
                err_lower is not None
                and err_upper is not None
                and idx < err_lower.size
                and idx < err_upper.size
            ):
                err_value = (float(err_lower[idx]), float(err_upper[idx]))
            if best is None or distance < best[0]:
                best = (distance, meta, idx, xs[idx], ys[idx], err_value)
        if best is None:
            self._hide_hover(axis=ax)
            return
        _, meta, idx, hover_x, hover_y, hover_err = best
        annotation = self._get_hover_annotation(ax)
        annotation.xy = (float(hover_x), float(hover_y))
        annotation.set_text(self._format_hover_text(meta, hover_x, hover_y, hover_err))
        if not annotation.get_visible():
            annotation.set_visible(True)
        self._hover_last = (ax, meta.get("artist"), idx)
        canvas = getattr(ax.figure, "canvas", None)
        if canvas is not None:
            canvas.draw_idle()

    def _on_canvas_leave(self, event):
        self._hide_hover()

    def _create_measurement_entry(self, key):
        if not self.measurement_panel or not self.measurement_list_sizer:
            return None
        panel = wx.Panel(self.measurement_panel)
        row = wx.BoxSizer(wx.HORIZONTAL)
        checkbox = wx.CheckBox(panel)
        checkbox.Bind(wx.EVT_CHECKBOX, partial(self._on_measurement_checkbox, key=key))
        row.Add(checkbox, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 6)
        label = wx.StaticText(panel, label="")
        label.Wrap(520)
        row.Add(label, 1, wx.ALIGN_CENTER_VERTICAL)
        panel.SetSizer(row)
        panel.Layout()
        self.measurement_list_sizer.Add(panel, 0, wx.TOP | wx.EXPAND, 4)
        entry = {"panel": panel, "checkbox": checkbox, "label": label}
        self._measurement_entries[key] = entry
        return entry

    def _remove_measurement_entry(self, key):
        entry = self._measurement_entries.pop(key, None)
        if not entry:
            return
        panel = entry.get("panel")
        if panel and self.measurement_list_sizer:
            try:
                self.measurement_list_sizer.Detach(panel)
            except Exception:
                pass
            panel.Destroy()

    def _update_measurement_entry(self, key, entry, preferred_meta=None):
        checkbox = entry.get("checkbox")
        label_ctrl = entry.get("label")
        run_index, measurement_idx, component = key
        selection = {
            "run_index": run_index,
            "measurement_idx": measurement_idx,
            "component": component,
        }
        meta, position = self._find_metadata_entry(selection, preferred_meta)
        if meta is None:
            return False
        text = self._format_measurement_label(selection, position)
        if label_ctrl:
            label_ctrl.SetLabel(text)
            label_ctrl.Wrap(520)
        if checkbox:
            excluded = False
            preview = self._preview_exclusions.get(run_index, set())
            try:
                excluded = int(measurement_idx) in {int(val) for val in preview}
            except Exception:
                excluded = measurement_idx in preview
            checkbox.SetEvtHandlerEnabled(False)
            checkbox.SetValue(excluded)
            checkbox.SetEvtHandlerEnabled(True)
        entry["meta"] = meta
        return True

    def _finalize_measurement_panel_visibility(self):
        if not self.measurement_panel:
            return
        has_entries = bool(self._measurement_entries)
        if self.measurement_instructions:
            if has_entries:
                message = (
                    "Toggle the checkboxes below to include or exclude measurements,"
                    " then press Accept to apply the changes."
                )
            else:
                message = "Click a subplot point to inspect it. Use Accept to apply changes."
            self.measurement_instructions.SetLabel(message)
            self.measurement_instructions.Wrap(520)
        if has_entries:
            if not self.measurement_panel.IsShown():
                self.measurement_panel.Show()
        else:
            if self.measurement_panel.IsShown():
                self.measurement_panel.Hide()
        self.measurement_panel.Layout()
        self.Layout()

    def _set_measurement_selection(self, meta, measurement_idx):
        if meta is None:
            return False
        run_index = meta.get("run_index")
        if run_index is None:
            return False
        try:
            run_idx_val = int(run_index)
            meas_idx_val = int(measurement_idx)
        except Exception:
            return False
        component = meta.get("component")
        key = (run_idx_val, meas_idx_val, component)
        entry = self._measurement_entries.get(key)
        if entry is None:
            entry = self._create_measurement_entry(key)
        if entry is None:
            return False
        if not self._update_measurement_entry(key, entry, meta):
            self._remove_measurement_entry(key)
            self._finalize_measurement_panel_visibility()
            return False
        if self.measurement_list_sizer:
            panel = entry.get("panel")
            if panel:
                try:
                    self.measurement_list_sizer.Detach(panel)
                except Exception:
                    panel = None
                if panel:
                    self.measurement_list_sizer.Insert(0, panel, 0, wx.TOP | wx.EXPAND, 4)
        self._finalize_measurement_panel_visibility()
        return True

    def _refresh_measurement_panel(self, meta=None):
        if not self.measurement_panel or not self.measurement_list_sizer:
            return
        for key in list(self._measurement_entries.keys()):
            entry = self._measurement_entries.get(key)
            if entry is None:
                continue
            if not self._update_measurement_entry(key, entry, meta):
                self._remove_measurement_entry(key)
        self._finalize_measurement_panel_visibility()

    def _find_metadata_entry(self, selection, preferred_meta=None):
        def matches(entry):
            if not entry:
                return False
            if entry.get("run_index") != selection.get("run_index"):
                return False
            indices = entry.get("measurement_indices") or []
            try:
                indices = [int(idx) for idx in indices]
            except Exception:
                return False
            if selection.get("measurement_idx") not in indices:
                return False
            component = selection.get("component")
            if component and entry.get("component") and entry.get("component") != component:
                return False
            return True

        candidates = []
        if matches(preferred_meta):
            candidates.append(preferred_meta)
        candidates.extend(self._artist_metadata)
        for entry in candidates:
            if not matches(entry):
                continue
            indices = entry.get("measurement_indices") or []
            try:
                idx_list = [int(idx) for idx in indices]
                position = idx_list.index(selection.get("measurement_idx"))
            except (ValueError, TypeError):
                position = 0
            return entry, position
        return None, None

    def _find_run_by_index(self, run_index):
        for run in self.runs:
            if run.get("index") == run_index:
                return run
        return None

    def _measurement_group_size(self, run, measurement_idx):
        if not run:
            return 1
        groups = run.get("variable_groups") or {}
        members = groups.get(measurement_idx)
        if not members:
            return 1
        try:
            return len(tuple(members))
        except TypeError:
            return 1

    def _resolve_var_scale(self, run):
        scale = None
        if run:
            scale = run.get("var_scale")
        if isinstance(scale, (int, float)) and np.isfinite(scale) and scale != 0:
            return float(scale)
        parent_scale = None
        if self.fitting_window:
            parent_scale = getattr(self.fitting_window, "var_scale", None)
        if isinstance(parent_scale, (int, float)) and np.isfinite(parent_scale) and parent_scale != 0:
            return float(parent_scale)
        return 1.0

    def _format_measurement_label(self, selection, position):
        run_idx = selection.get("run_index")
        measurement_idx = selection.get("measurement_idx")
        try:
            idx_value = int(measurement_idx)
        except (TypeError, ValueError):
            idx_value = None
        component = selection.get("component")
        run = self._find_run_by_index(run_idx)
        run_label = f"Run {run_idx + 1}" if isinstance(run_idx, int) else "Run"
        component_names = {"sigma_x": "σₓ²", "sigma_y": "σ_y²"}
        comp_label = component_names.get(component, component or "Measurement")
        if position is not None:
            display_index = position + 1
        else:
            display_index = (idx_value if idx_value is not None else 0) + 1
        label = f"{run_label} – {comp_label} #{display_index}"
        if run:
            scale = self._resolve_var_scale(run)
            var_label = run.get("var_label") or "Variable"
            var_unit = run.get("var_unit") or ""
            unit_str = f" {var_unit}" if var_unit else ""

            def _format_var_value(raw_value):
                try:
                    numeric = float(raw_value)
                except (TypeError, ValueError):
                    return f"{raw_value}{unit_str}"
                if not np.isfinite(numeric):
                    return f"{numeric:.6g}{unit_str}"
                display_val = numeric / scale if scale else numeric
                if np.isfinite(display_val):
                    return f"{display_val:.6g}{unit_str}"
                return f"{numeric:.6g}{unit_str}"

            values = run.get("filtered_var_values") or []
            included = run.get("included_indices") or []
            if idx_value is not None and idx_value in included:
                try:
                    idx_pos = included.index(idx_value)
                except ValueError:
                    idx_pos = -1
                if 0 <= idx_pos < len(values):
                    var_value = values[idx_pos]
                    value_text = _format_var_value(var_value)
                    label += f" – {var_label}={value_text}"
            else:
                all_values = run.get("var_values") or []
                if idx_value is not None and 0 <= idx_value < len(all_values):
                    var_value = all_values[idx_value]
                    value_text = _format_var_value(var_value)
                    label += f" – {var_label}={value_text}"
            group_size = self._measurement_group_size(run, idx_value)
            if group_size > 1:
                label += f" ({group_size} points averaged)"
        return label

    def _on_canvas_pick(self, event):
        artist = getattr(event, "artist", None)
        indices = getattr(event, "ind", None)
        if artist is None or not indices:
            return
        meta = next((m for m in self._artist_metadata if m.get("artist") is artist), None)
        if not meta:
            return
        measurement_indices = meta.get("measurement_indices") or []
        run_index = meta.get("run_index")
        if run_index is None:
            return
        for idx in indices:
            try:
                point_idx = int(idx)
            except Exception:
                continue
            if point_idx < 0 or point_idx >= len(measurement_indices):
                continue
            try:
                measurement_idx = int(measurement_indices[point_idx])
            except Exception:
                continue
            if self._set_measurement_selection(meta, measurement_idx):
                break

    def _update_header(self):
        if not self.info_label:
            return
        plot_num = self.point_info.get("plot")
        pieces = []
        if plot_num is not None:
            pieces.append(f"Plot {plot_num}")
        for key in ("x_text", "y_text"):
            text = self.point_info.get(key)
            if text:
                pieces.append(text)
        summary = " • ".join(pieces) if pieces else "Selected data point"
        self.info_label.SetLabel(summary)
        excluded = bool(getattr(self, "_preview_point_excluded", self.point_info.get("excluded")))
        self.point_info["excluded"] = excluded
        if self.exclude_checkbox:
            self.exclude_checkbox.SetEvtHandlerEnabled(False)
            self.exclude_checkbox.SetValue(excluded)
            self.exclude_checkbox.SetEvtHandlerEnabled(True)

    def _populate_filmstrip(self):
        if not self.scrolled_window or not self.thumbnail_sizer:
            return
        self.thumbnail_sizer.Clear(True)
        self._thumbnails = []
        if not self.image_entries:
            message = wx.StaticText(
                self.scrolled_window,
                label="No source images are associated with this point.",
            )
            self.thumbnail_sizer.Add(message, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 10)
            self.scrolled_window.Layout()
            self.scrolled_window.FitInside()
            return
        for entry in self.image_entries:
            path = entry.get("path")
            panel = wx.Panel(self.scrolled_window)
            vbox = wx.BoxSizer(wx.VERTICAL)
            bmp = self._load_thumbnail(path)
            if bmp is None:
                bmp = self._missing_thumbnail()
            if bmp is not None:
                self._thumbnails.append(bmp)
                bmp_ctrl = wx.StaticBitmap(panel, bitmap=bmp)
                vbox.Add(bmp_ctrl, 0, wx.BOTTOM | wx.ALIGN_CENTER_HORIZONTAL, 4)
            name = os.path.basename(path) if path else "Unknown"
            name_label = wx.StaticText(panel, label=name or "(unnamed)")
            vbox.Add(name_label, 0, wx.ALIGN_CENTER_HORIZONTAL)
            caption = self._format_image_caption(entry)
            if caption:
                caption_label = wx.StaticText(panel, label=caption)
                caption_label.Wrap(180)
                vbox.Add(caption_label, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.TOP, 2)
            timestamp = self._format_timestamp(path)
            ts_label = wx.StaticText(panel, label=timestamp)
            ts_font = ts_label.GetFont()
            ts_font.SetPointSize(max(ts_font.GetPointSize() - 1, 7))
            ts_label.SetFont(ts_font)
            vbox.Add(ts_label, 0, wx.ALIGN_CENTER_HORIZONTAL)
            tooltip_parts = [path or ""]
            if caption:
                tooltip_parts.append(caption)
            panel.SetToolTip("\n".join(part for part in tooltip_parts if part))
            panel.SetSizer(vbox)
            panel.Layout()
            self.thumbnail_sizer.Add(panel, 0, wx.ALL, 6)
        self.thumbnail_sizer.Layout()
        self.scrolled_window.FitInside()
        self.scrolled_window.Layout()

    def _get_parent_flip_state(self):
        parent = getattr(self.fitting_window, "parent", None)
        if parent is None:
            return False, False

        def _read_flag(control_attr, fallback_attr):
            control = getattr(parent, control_attr, None)
            if control is not None and hasattr(control, "GetValue"):
                try:
                    return bool(control.GetValue())
                except Exception:
                    return False
            fallback = getattr(parent, fallback_attr, None)
            if isinstance(fallback, bool):
                return fallback
            return False

        flip_h = _read_flag("flipHCheck", "flip_horizontal")
        flip_v = _read_flag("flipVCheck", "flip_vertical")
        return flip_h, flip_v

    def _load_thumbnail(self, path, target_height=FILMSTRIP_TARGET_HEIGHT):
        if not path or not os.path.exists(path):
            return None
        image = None
        ext = os.path.splitext(path)[1].lower()
        standard_formats = {".png", ".jpg", ".jpeg", ".bmp", ".gif"}
        flip_h, flip_v = self._get_parent_flip_state()
        if ext in standard_formats and wx.Image.CanRead(path):
            try:
                with wx.LogNull():
                    image = wx.Image(path)
            except Exception:
                image = None
        if image is None or not image.IsOk():
            bmp = self._load_absorption_bitmap(path, target_height, flip_h, flip_v)
            if bmp is not None:
                return bmp
            return None
        if flip_h or flip_v:
            try:
                if flip_h:
                    image = image.Mirror(True)
                if flip_v:
                    image = image.Mirror(False)
            except Exception:
                pass
        width = max(image.GetWidth(), 1)
        height = max(image.GetHeight(), 1)
        if height <= 0:
            return None
        target_height = max(target_height, 40)
        target_width = int(width * (target_height / float(height)))
        if target_width <= 0:
            target_width = target_height
        image = image.Scale(target_width, target_height, wx.IMAGE_QUALITY_HIGH)
        return wx.Bitmap(image)

    def _load_absorption_bitmap(self, path, target_height, flip_h=False, flip_v=False):
        ext = os.path.splitext(path)[1].lower()
        supported = {".fits", ".fit", ".fts", ".aia", ".tif", ".tiff"}
        parent = self.fitting_window.parent if self.fitting_window else None
        file_type = getattr(parent, "fileType", "") or ext.lstrip(".")
        normalized = file_type.lower()
        if normalized in {"tiff", "tif"}:
            normalized = "tif"
        elif normalized in {"fit", "fts"}:
            normalized = "fits"
        if ext not in supported and normalized not in {"fits", "aia", "tif"}:
            return None
        try:
            absorb, _ = readData(path, normalized, [False, None])
        except Exception:
            return None
        try:
            arr = np.asarray(absorb, dtype=float)
        except Exception:
            return None
        if arr.size == 0:
            return None
        finite = arr[np.isfinite(arr)]
        if finite.size == 0:
            return None
        try:
            low, high = np.percentile(finite, (5, 95))
        except Exception:
            low, high = np.nanmin(finite), np.nanmax(finite)
        if not np.isfinite(low) or not np.isfinite(high) or np.isclose(low, high):
            low = np.nanmin(finite)
            high = np.nanmax(finite)
        if not np.isfinite(low) or not np.isfinite(high) or np.isclose(low, high):
            high = low + 1.0
        with np.errstate(invalid="ignore"):
            norm = np.clip((arr - low) / (high - low), 0.0, 1.0)
        data = (norm * 255).astype(np.uint8)
        if flip_h:
            data = np.fliplr(data)
        if flip_v:
            data = np.flipud(data)
        height, width = data.shape
        if height <= 0 or width <= 0:
            return None
        image = wx.Image(width, height)
        rgb = np.dstack([data] * 3)
        try:
            image.SetData(rgb.tobytes())
        except Exception:
            return None
        target_height = max(target_height, 40)
        target_width = int(width * (target_height / float(height)))
        if target_width <= 0:
            target_width = target_height
        image = image.Scale(target_width, target_height, wx.IMAGE_QUALITY_HIGH)
        return wx.Bitmap(image)

    def _missing_thumbnail(self, size=None):
        if size is None:
            size = max(int(round(160 * FILMSTRIP_SIZE_SCALE)), 60)
        art = wx.ArtProvider.GetBitmap(
            wx.ART_MISSING_IMAGE, wx.ART_OTHER, wx.Size(size, size)
        )
        if art and art.IsOk():
            if art.GetHeight() != size or art.GetWidth() != size:
                img = art.ConvertToImage()
                img = img.Scale(size, size, wx.IMAGE_QUALITY_HIGH)
                return wx.Bitmap(img)
            return art
        bmp = wx.Bitmap(size, size)
        dc = wx.MemoryDC(bmp)
        dc.SetBackground(wx.Brush(wx.Colour(240, 240, 240)))
        dc.Clear()
        dc.SelectObject(wx.NullBitmap)
        return bmp

    def _format_timestamp(self, path):
        if not path or not os.path.exists(path):
            return "File not found"
        try:
            ts = datetime.datetime.fromtimestamp(os.path.getmtime(path))
        except Exception:
            return "Timestamp unavailable"
        return ts.strftime("%Y-%m-%d %H:%M:%S")

    def _format_image_caption(self, entry):
        if not entry:
            return ""
        x_val = entry.get("x")
        y_val = entry.get("y")
        if x_val is None and y_val is None:
            return ""

        def _format_component(value):
            if value is None:
                return "?"
            if isinstance(value, (int, float)):
                if not np.isfinite(value):
                    return "?"
                return f"{value:.3g}"
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                return str(value)
            if not np.isfinite(numeric):
                return "?"
            return f"{numeric:.3g}"

        x_text = _format_component(x_val)
        y_text = _format_component(y_val)
        y_unit = entry.get("y_unit") or ""
        if y_unit:
            y_text = f"{y_text} {y_unit}"
        return f"({x_text}, {y_text})"

    def _on_exclude_changed(self, event):
        value = bool(self.exclude_checkbox.GetValue()) if self.exclude_checkbox else False
        self._preview_point_excluded = value
        self.point_info["excluded"] = value
        self._update_accept_state()

    def _on_measurement_checkbox(self, event, key):
        try:
            run_idx, measurement_idx, _ = key
        except ValueError:
            return
        try:
            run_idx = int(run_idx)
            measurement_idx = int(measurement_idx)
        except Exception:
            return
        checkbox = event.GetEventObject()
        desired = bool(checkbox.GetValue()) if checkbox else False
        members = self._resolve_group_members(run_idx, measurement_idx)
        preview = self._preview_exclusions.setdefault(
            run_idx, set(self._baseline_exclusions.get(run_idx, set()))
        )
        members = tuple(int(idx) for idx in members)
        if desired:
            preview.update(members)
        else:
            for idx in members:
                preview.discard(idx)
        self._apply_preview(refresh=True)
        self._refresh_measurement_panel()
        self._update_accept_state()

    def sync_exclude_state(self, value):
        if not self.exclude_checkbox:
            return
        desired = bool(value)
        current = bool(self.exclude_checkbox.GetValue())
        if current != desired:
            self.exclude_checkbox.SetEvtHandlerEnabled(False)
            self.exclude_checkbox.SetValue(desired)
            self.exclude_checkbox.SetEvtHandlerEnabled(True)
        self.point_info["excluded"] = desired
        self._baseline_point_excluded = desired
        self._preview_point_excluded = desired
        self._refresh_measurement_panel()
        self._update_accept_state()

    def Destroy(self):
        self._disconnect_pick_events()
        return super().Destroy()
